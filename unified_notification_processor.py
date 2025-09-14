#!/usr/bin/env python3
"""
Unified Notification Processor
Logic layer that reads master notification rules and creates sent_noti_now.xlsx
for Power Automate to process every 10 minutes.

This script:
1. Reads master_notification_rules.xlsx
2. Applies time conditions and trigger logic
3. Creates/updates sent_noti_now.xlsx with notifications due NOW
4. Ensures consistent format: EmployeeID, ContactEmail, Message, NotificationType, Priority
"""

import pandas as pd
import sqlite3
from datetime import datetime, date, timedelta
from pathlib import Path
import json
import logging
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class UnifiedNotificationProcessor:
    """Processes master notification rules and generates current notifications"""
    
    def __init__(self, network_folder="network_folder_simplified", db_path="vendor_timesheet.db"):
        self.network_folder = Path(network_folder)
        self.network_folder.mkdir(exist_ok=True)
        self.db_path = db_path
        self.master_rules_file = "master_notification_rules.xlsx"
        self.output_file = self.network_folder / "sent_noti_now.xlsx"
        self.context_file = self.network_folder / "notification_context.json"
        
        # Standard output format columns as requested
        self.output_columns = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 'Priority']
        
    def load_master_rules(self) -> pd.DataFrame:
        """Load master notification rules from Excel"""
        try:
            if not Path(self.master_rules_file).exists():
                logger.error(f"❌ Master rules file not found: {self.master_rules_file}")
                return pd.DataFrame()
            
            df = pd.read_excel(self.master_rules_file)
            logger.info(f"✅ Loaded {len(df)} notification rules from master file")
            return df
        except Exception as e:
            logger.error(f"❌ Error loading master rules: {str(e)}")
            return pd.DataFrame()
    
    def get_database_connection(self):
        """Get database connection"""
        if not Path(self.db_path).exists():
            logger.error(f"❌ Database not found: {self.db_path}")
            return None
        return sqlite3.connect(self.db_path)
    
    def check_time_condition(self, rule: Dict, current_time: datetime) -> bool:
        """Check if current time matches rule's trigger conditions"""
        try:
            # Get time intervals
            intervals = rule.get('Time_Intervals', '').strip()
            
            if intervals == 'IMMEDIATE':
                return True  # Event-driven, check trigger condition separately
            
            if not intervals:
                return False
            
            # Parse time intervals
            times = [t.strip() for t in intervals.split(',')]
            current_time_str = current_time.strftime('%H:%M')
            
            # Check if current time matches any interval (with 10-minute window)
            for time_str in times:
                try:
                    rule_time = datetime.strptime(time_str, '%H:%M').time()
                    current_time_only = current_time.time()
                    
                    # Allow 10-minute window
                    rule_datetime = datetime.combine(current_time.date(), rule_time)
                    time_diff = abs((current_time - rule_datetime).total_seconds() / 60)
                    
                    if time_diff <= 10:  # Within 10 minutes
                        return True
                except ValueError:
                    continue
            
            return False
        except Exception as e:
            logger.error(f"Error checking time condition: {str(e)}")
            return False
    
    def check_weekday_condition(self, rule: Dict, current_time: datetime) -> bool:
        """Check weekday conditions"""
        weekdays_only = rule.get('Weekdays_Only', 'NO').upper() == 'YES'
        if weekdays_only and current_time.weekday() >= 5:  # Saturday=5, Sunday=6
            return False
        return True
    
    def check_holiday_condition(self, rule: Dict, current_time: datetime) -> bool:
        """Check holiday exclusion conditions"""
        exclude_holidays = rule.get('Exclude_Holidays', 'NO').upper() == 'YES'
        if exclude_holidays:
            # Simple holiday check - you can enhance this
            if self.is_holiday(current_time.date()):
                return False
        return True
    
    def is_holiday(self, check_date: date) -> bool:
        """Check if given date is a holiday (simplified implementation)"""
        try:
            conn = self.get_database_connection()
            if not conn:
                return False
            
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM holidays WHERE holiday_date = ? LIMIT 1", [check_date.isoformat()])
            result = cursor.fetchone()
            conn.close()
            return result is not None
        except Exception:
            return False
    
    def get_pending_vendors(self) -> List[Dict]:
        """Get vendors who haven't submitted today's status"""
        try:
            conn = self.get_database_connection()
            if not conn:
                return []
            
            today = date.today().isoformat()
            
            query = """
            SELECT 
                v.vendor_id as EmployeeID,
                u.email as ContactEmail,
                v.full_name as EmployeeName,
                v.department,
                v.company,
                v.manager_id,
                m.email as manager_email,
                m.full_name as manager_name
            FROM vendors v
            JOIN users u ON v.user_id = u.id
            LEFT JOIN managers m ON v.manager_id = m.manager_id
            LEFT JOIN daily_statuses ds ON v.id = ds.vendor_id AND ds.status_date = ?
            WHERE u.is_active = 1 AND ds.id IS NULL
            ORDER BY v.vendor_id
            """
            
            df = pd.read_sql_query(query, conn, params=[today])
            conn.close()
            
            return df.to_dict('records')
        except Exception as e:
            logger.error(f"Error getting pending vendors: {str(e)}")
            return []
    
    def get_managers_with_pending_teams(self) -> List[Dict]:
        """Get managers who have team members with pending submissions"""
        try:
            conn = self.get_database_connection()
            if not conn:
                return []
            
            today = date.today().isoformat()
            
            query = """
            SELECT 
                m.manager_id as EmployeeID,
                m.email as ContactEmail,
                m.full_name as ManagerName,
                m.department,
                COUNT(v.id) as pending_count
            FROM managers m
            JOIN vendors v ON m.manager_id = v.manager_id
            JOIN users u ON v.user_id = u.id
            LEFT JOIN daily_statuses ds ON v.id = ds.vendor_id AND ds.status_date = ?
            WHERE u.is_active = 1 AND ds.id IS NULL
            GROUP BY m.manager_id, m.email, m.full_name, m.department
            HAVING pending_count > 0
            ORDER BY pending_count DESC
            """
            
            df = pd.read_sql_query(query, conn, params=[today])
            conn.close()
            
            return df.to_dict('records')
        except Exception as e:
            logger.error(f"Error getting managers with pending teams: {str(e)}")
            return []
    
    def get_admins(self) -> List[Dict]:
        """Get system administrators"""
        try:
            # For now, return a default admin - you can enhance this
            return [{
                'EmployeeID': 'ADMIN_001',
                'ContactEmail': 'admin@company.com',
                'AdminName': 'System Administrator',
                'department': 'IT'
            }]
        except Exception as e:
            logger.error(f"Error getting admins: {str(e)}")
            return []
    
    def process_daily_status_reminders(self, rule: Dict, current_time: datetime) -> List[Dict]:
        """Process daily status reminder notifications"""
        notifications = []
        pending_vendors = self.get_pending_vendors()
        
        for vendor in pending_vendors:
            notifications.append({
                'EmployeeID': vendor['EmployeeID'],
                'ContactEmail': vendor['ContactEmail'],
                'Message': f"{rule['Message_Template']} - {vendor.get('EmployeeName', '')}",
                'NotificationType': rule['Notification_Type'],
                'Priority': rule['Priority']
            })
        
        logger.info(f"📝 Generated {len(notifications)} daily status reminders")
        return notifications
    
    def process_manager_summary(self, rule: Dict, current_time: datetime) -> List[Dict]:
        """Process manager summary notifications"""
        notifications = []
        managers = self.get_managers_with_pending_teams()
        
        for manager in managers:
            pending_count = manager.get('pending_count', 0)
            notifications.append({
                'EmployeeID': manager['EmployeeID'],
                'ContactEmail': manager['ContactEmail'],
                'Message': f"{rule['Message_Template']} - {pending_count} team members pending",
                'NotificationType': rule['Notification_Type'],
                'Priority': rule['Priority']
            })
        
        logger.info(f"👔 Generated {len(notifications)} manager summaries")
        return notifications
    
    def process_late_submissions(self, rule: Dict, current_time: datetime) -> List[Dict]:
        """Process late submission notifications"""
        notifications = []
        
        # Get managers with pending teams (late submissions)
        managers = self.get_managers_with_pending_teams()
        
        for manager in managers:
            pending_count = manager.get('pending_count', 0)
            notifications.append({
                'EmployeeID': manager['EmployeeID'],
                'ContactEmail': manager['ContactEmail'],
                'Message': f"{rule['Message_Template']} - {pending_count} late submissions",
                'NotificationType': rule['Notification_Type'],
                'Priority': rule['Priority']
            })
        
        # Also notify admins
        admins = self.get_admins()
        for admin in admins:
            total_late = len(self.get_pending_vendors())
            if total_late > 0:
                notifications.append({
                    'EmployeeID': admin['EmployeeID'],
                    'ContactEmail': admin['ContactEmail'],
                    'Message': f"{rule['Message_Template']} - {total_late} total late submissions",
                    'NotificationType': rule['Notification_Type'],
                    'Priority': rule['Priority']
                })
        
        logger.info(f"⚠️ Generated {len(notifications)} late submission alerts")
        return notifications
    
    def process_monthly_reports(self, rule: Dict, current_time: datetime) -> List[Dict]:
        """Process monthly report notifications (1st of month only)"""
        notifications = []
        
        # Only on 1st of month at 9 AM
        if current_time.day == 1 and current_time.hour == 9:
            # Notify managers
            managers = self.get_managers_with_pending_teams()
            for manager in managers:
                notifications.append({
                    'EmployeeID': manager['EmployeeID'],
                    'ContactEmail': manager['ContactEmail'],
                    'Message': f"{rule['Message_Template']} for {current_time.strftime('%B %Y')}",
                    'NotificationType': rule['Notification_Type'],
                    'Priority': rule['Priority']
                })
            
            # Notify admins
            admins = self.get_admins()
            for admin in admins:
                notifications.append({
                    'EmployeeID': admin['EmployeeID'],
                    'ContactEmail': admin['ContactEmail'],
                    'Message': f"{rule['Message_Template']} for {current_time.strftime('%B %Y')}",
                    'NotificationType': rule['Notification_Type'],
                    'Priority': rule['Priority']
                })
        
        logger.info(f"📊 Generated {len(notifications)} monthly reports")
        return notifications
    
    def process_holiday_reminders(self, rule: Dict, current_time: datetime) -> List[Dict]:
        """Process holiday reminder notifications"""
        notifications = []
        
        # Simple implementation - you can enhance this
        # For now, just check if it's a Friday (weekend reminder)
        if current_time.weekday() == 4:  # Friday
            all_users = []
            all_users.extend(self.get_pending_vendors())  # Vendors
            all_users.extend(self.get_managers_with_pending_teams())  # Managers
            all_users.extend(self.get_admins())  # Admins
            
            for user in all_users:
                notifications.append({
                    'EmployeeID': user['EmployeeID'],
                    'ContactEmail': user['ContactEmail'],
                    'Message': f"{rule['Message_Template']} - Weekend approaching",
                    'NotificationType': rule['Notification_Type'],
                    'Priority': rule['Priority']
                })
        
        logger.info(f"🏖️ Generated {len(notifications)} holiday reminders")
        return notifications
    
    def process_rule(self, rule: Dict, current_time: datetime) -> List[Dict]:
        """Process a single notification rule"""
        try:
            # Check if rule is active
            if rule.get('Active', 'NO').upper() != 'YES':
                return []
            
            # Check time conditions
            if not self.check_time_condition(rule, current_time):
                return []
            
            # Check weekday conditions
            if not self.check_weekday_condition(rule, current_time):
                return []
            
            # Check holiday conditions
            if not self.check_holiday_condition(rule, current_time):
                return []
            
            # Process based on notification type
            notification_type = rule['Notification_Type']
            
            if notification_type == 'Daily Status Reminders':
                return self.process_daily_status_reminders(rule, current_time)
            elif notification_type == 'Manager Summary':
                return self.process_manager_summary(rule, current_time)
            elif notification_type == 'Late Submission':
                return self.process_late_submissions(rule, current_time)
            elif notification_type == 'Monthly Reports':
                return self.process_monthly_reports(rule, current_time)
            elif notification_type == 'Holiday Reminders':
                return self.process_holiday_reminders(rule, current_time)
            else:
                # For other notification types, implement similar logic
                logger.info(f"⚠️ Notification type '{notification_type}' not yet implemented")
                return []
                
        except Exception as e:
            logger.error(f"Error processing rule {rule.get('Notification_Type', 'Unknown')}: {str(e)}")
            return []
    
    def create_output_excel(self, notifications: List[Dict]):
        """Create sent_noti_now.xlsx with proper table formatting"""
        try:
            # Create DataFrame with standard format
            df = pd.DataFrame(notifications, columns=self.output_columns)
            
            # If no notifications, create empty DataFrame with headers
            if df.empty:
                df = pd.DataFrame(columns=self.output_columns)
            
            # Create Excel with table formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Notifications_Now"
            
            # Write headers
            for col_num, column_title in enumerate(self.output_columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
            
            # Write data
            if not df.empty:
                for row_num, row_data in enumerate(df.values, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = cell_value
            
            # Create table
            table_range = f"A1:{get_column_letter(len(self.output_columns))}{len(df) + 1}"
            table = Table(displayName="NotificationsNow", ref=table_range)
            
            # Add table style
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            
            # Add table to worksheet
            ws.add_table(table)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(self.output_file)
            logger.info(f"✅ Created output file: {self.output_file} with {len(df)} notifications")
            
        except Exception as e:
            logger.error(f"Error creating output Excel: {str(e)}")
    
    def save_context(self, notifications_count: int):
        """Save processing context for tracking"""
        context = {
            'last_run': datetime.now().isoformat(),
            'notifications_generated': notifications_count,
            'output_file': str(self.output_file),
            'status': 'SUCCESS'
        }
        
        try:
            with open(self.context_file, 'w') as f:
                json.dump(context, f, indent=2)
        except Exception as e:
            logger.error(f"Error saving context: {str(e)}")
    
    def process_notifications(self) -> int:
        """Main processing function - returns count of notifications generated"""
        try:
            current_time = datetime.now()
            logger.info(f"🚀 Processing notifications at {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Load master rules
            master_rules = self.load_master_rules()
            if master_rules.empty:
                logger.error("❌ No master rules loaded")
                return 0
            
            # Process all rules
            all_notifications = []
            
            for _, rule in master_rules.iterrows():
                rule_notifications = self.process_rule(rule.to_dict(), current_time)
                all_notifications.extend(rule_notifications)
            
            # Remove duplicates based on EmployeeID and ContactEmail
            unique_notifications = []
            seen = set()
            for notif in all_notifications:
                key = (notif['EmployeeID'], notif['ContactEmail'], notif['NotificationType'])
                if key not in seen:
                    seen.add(key)
                    unique_notifications.append(notif)
            
            # Create output Excel
            self.create_output_excel(unique_notifications)
            
            # Save context
            self.save_context(len(unique_notifications))
            
            logger.info(f"✅ Processing complete: {len(unique_notifications)} notifications ready for Power Automate")
            return len(unique_notifications)
            
        except Exception as e:
            logger.error(f"❌ Error in notification processing: {str(e)}")
            return 0

# Global instance
unified_processor = UnifiedNotificationProcessor()

def run_notification_processing():
    """Public function to run notification processing"""
    return unified_processor.process_notifications()

if __name__ == "__main__":
    # Command line interface
    import argparse
    
    parser = argparse.ArgumentParser(description='Unified Notification Processor')
    parser.add_argument('--test', action='store_true', help='Run in test mode')
    
    args = parser.parse_args()
    
    if args.test:
        print("🧪 Running in test mode...")
    
    notifications_count = run_notification_processing()
    print(f"\n📊 PROCESSING SUMMARY:")
    print(f"   Notifications Generated: {notifications_count}")
    print(f"   Output File: {unified_processor.output_file}")
    print(f"   Ready for Power Automate: {'YES' if notifications_count > 0 else 'NO NOTIFICATIONS DUE'}")
