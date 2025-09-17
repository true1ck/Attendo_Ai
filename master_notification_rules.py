#!/usr/bin/env python3
"""
Master Notification Rules Excel Generator
Creates a master Excel file with all notification rules for the unified Power Automate flow system.
"""

import pandas as pd
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def create_master_notification_rules():
    """Create master Excel file with notification rules for unified Power Automate system"""
    
    print("🚀 Creating Master Notification Rules Excel file...")
    
    # Master notification rules data - exactly as per your requirements
    master_rules = [
        {
            'Notification_Type': 'Daily Status Reminders',
            'Frequency': 'Every 3 hours',
            'Primary_Trigger_Time': '9 AM – 6 PM',
            'Recipients': 'Vendors',
            'Priority': 'Medium',
            'Trigger_Condition': 'PENDING_SUBMISSION',
            'Time_Intervals': '09:00,12:00,15:00,18:00',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'YES',
            'Escalation_Hours': 0,
            'Message_Template': 'Please submit your daily attendance status',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'Manager Summary',
            'Frequency': '2x daily',
            'Primary_Trigger_Time': '12 PM, 2 PM',
            'Recipients': 'Managers',
            'Priority': 'Medium',
            'Trigger_Condition': 'SCHEDULED',
            'Time_Intervals': '12:00,14:00',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'YES',
            'Escalation_Hours': 0,
            'Message_Template': 'Team attendance summary report',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'All-Complete',
            'Frequency': 'Event-driven',
            'Primary_Trigger_Time': 'When team complete',
            'Recipients': 'Managers',
            'Priority': 'Low',
            'Trigger_Condition': 'ALL_TEAM_COMPLETE',
            'Time_Intervals': 'IMMEDIATE',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 0,
            'Message_Template': 'All team members have submitted their status',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'Mismatch Alerts',
            'Frequency': 'Daily + Escalation',
            'Primary_Trigger_Time': '6 PM daily',
            'Recipients': 'Vendors, Managers',
            'Priority': 'High',
            'Trigger_Condition': 'MISMATCH_DETECTED',
            'Time_Intervals': '18:00',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 24,
            'Message_Template': 'Attendance mismatch detected - explanation required',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'Manager Feedback',
            'Frequency': 'Event-driven',
            'Primary_Trigger_Time': 'Within 1 hour',
            'Recipients': 'Vendors',
            'Priority': 'Medium',
            'Trigger_Condition': 'MANAGER_ACTION',
            'Time_Intervals': 'IMMEDIATE',
            'Weekdays_Only': 'NO',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 1,
            'Message_Template': 'Manager feedback on your attendance submission',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'Monthly Reports',
            'Frequency': 'Monthly',
            'Primary_Trigger_Time': '1st at 9 AM',
            'Recipients': 'Managers, Admins',
            'Priority': 'Medium',
            'Trigger_Condition': 'MONTHLY_SCHEDULE',
            'Time_Intervals': '09:00',
            'Weekdays_Only': 'NO',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 0,
            'Message_Template': 'Monthly attendance report is ready',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'System Alerts',
            'Frequency': 'Event-driven',
            'Primary_Trigger_Time': 'Immediate',
            'Recipients': 'Admins',
            'Priority': 'Critical',
            'Trigger_Condition': 'SYSTEM_ERROR',
            'Time_Intervals': 'IMMEDIATE',
            'Weekdays_Only': 'NO',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 0,
            'Message_Template': 'System alert requiring immediate attention',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'Holiday Reminders',
            'Frequency': 'Before holidays',
            'Primary_Trigger_Time': '3 days, 1 day before',
            'Recipients': 'All Users',
            'Priority': 'Low',
            'Trigger_Condition': 'HOLIDAY_APPROACHING',
            'Time_Intervals': '09:00',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 0,
            'Message_Template': 'Upcoming holiday notification',
            'Active': 'YES'
        },
        {
            'Notification_Type': 'Late Submission',
            'Frequency': 'After deadlines',
            'Primary_Trigger_Time': '24h, 48h after',
            'Recipients': 'Managers, Admins',
            'Priority': 'High',
            'Trigger_Condition': 'DEADLINE_PASSED',
            'Time_Intervals': '18:00',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'NO',
            'Escalation_Hours': 24,
            'Message_Template': 'Late submission alert - follow up required',
            'Active': 'YES'
        }
    ]
    
    # Create DataFrame
    df = pd.DataFrame(master_rules)
    
    # Add metadata columns
    df['Created_Date'] = datetime.now().strftime('%Y-%m-%d')
    df['Last_Modified'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    df['Modified_By'] = 'SYSTEM_SETUP'
    
    # Create the Excel file with proper table formatting
    filename = 'master_notification_rules.xlsx'
    filepath = Path(filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Notification_Rules"
    
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    # Create table
    table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    table = Table(displayName="NotificationRules", ref=table_range)
    
    # Add table style
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Add table to worksheet
    ws.add_table(table)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filepath)
    
    print(f"✅ Created master notification rules Excel file: {filepath.absolute()}")
    print(f"📊 Total notification types configured: {len(master_rules)}")
    
    # Display summary
    print("\n📋 MASTER NOTIFICATION RULES SUMMARY:")
    print("=" * 60)
    for i, rule in enumerate(master_rules, 1):
        print(f"{i:2d}. {rule['Notification_Type']:<25} | {rule['Priority']:<8} | {rule['Recipients']}")
    
    return str(filepath.absolute())

if __name__ == "__main__":
    create_master_notification_rules()
