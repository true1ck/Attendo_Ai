#!/usr/bin/env python3
"""
Manager Complete Notifications Table Formatter
==============================================
This utility script formats the 03_manager_all_complete_notifications.xlsx file
as a proper Excel table for Power Automate integration.

Usage:
    python format_manager_notifications_table.py

Features:
- Converts Excel file to proper table format
- Sets appropriate table styling
- Auto-adjusts column widths
- Maintains data integrity
- Power Automate compatible
"""

import sys
from pathlib import Path

# Add scripts directory to path
sys.path.insert(0, 'scripts')

def main():
    """Main function to format the manager complete notifications Excel file"""
    print("=" * 80)
    print("📊 MANAGER COMPLETE NOTIFICATIONS TABLE FORMATTER")
    print("=" * 80)
    
    try:
        from manager_complete_notification_handler import convert_to_table_format
        
        # Default file path
        excel_file = "notification_configs/03_manager_all_complete_notifications.xlsx"
        table_name = "ManagerCompleteNotifications"
        
        # Check if file exists
        if not Path(excel_file).exists():
            print(f"❌ Excel file not found: {excel_file}")
            print("Please ensure the file exists before formatting.")
            return False
        
        print(f"📁 File: {excel_file}")
        print(f"🏷️ Table Name: {table_name}")
        print()
        
        # Convert to table format
        print("🎨 Converting to table format...")
        success = convert_to_table_format(excel_file, table_name)
        
        if success:
            print()
            print("✅ SUCCESS! The Excel file has been formatted as a table.")
            print()
            print("🔧 What was done:")
            print("  ✅ Data converted to Excel table format")
            print("  ✅ Applied TableStyleMedium9 (blue styling)")
            print("  ✅ Auto-adjusted column widths")
            print("  ✅ Row stripes enabled for readability")
            print("  ✅ Power Automate compatibility ensured")
            print()
            print("🚀 Next Steps:")
            print("  1. Open the Excel file to verify table formatting")
            print("  2. Configure Power Automate to read from the table")
            print("  3. Set up notification delivery mechanisms")
            print("  4. Test the complete workflow")
            print()
            print("💡 Power Automate Tips:")
            print("  - Use 'List rows present in a table' action")
            print("  - Reference table name: 'ManagerCompleteNotifications'")
            print("  - Filter by 'NotiStatus' = false for pending notifications")
            
        else:
            print("❌ FAILED to format Excel file as table.")
            print("Please check the file format and try again.")
            
        return success
        
    except ImportError as e:
        print(f"❌ Import error: {str(e)}")
        print("Please ensure the manager_complete_notification_handler.py script is available.")
        return False
    except Exception as e:
        print(f"❌ Unexpected error: {str(e)}")
        return False

def show_current_table_info():
    """Show information about the current table format"""
    try:
        from openpyxl import load_workbook
        
        excel_file = "notification_configs/03_manager_all_complete_notifications.xlsx"
        
        if not Path(excel_file).exists():
            print(f"❌ Excel file not found: {excel_file}")
            return
        
        wb = load_workbook(excel_file)
        ws = wb.active
        
        print("=" * 80)
        print("📋 CURRENT TABLE INFORMATION")
        print("=" * 80)
        print(f"📄 Worksheet: {ws.title}")
        print(f"📊 Data Range: A1 to column {ws.max_column}, row {ws.max_row}")
        print(f"🏷️ Tables Found: {len(ws.tables)}")
        
        if ws.tables:
            for table in ws.tables.values():
                print(f"\n📋 Table Details:")
                print(f"  Name: {table.displayName}")
                print(f"  Range: {table.ref}")
                style_name = table.tableStyleInfo.name if table.tableStyleInfo else 'None'
                print(f"  Style: {style_name}")
                
                if table.tableStyleInfo:
                    print(f"  Row Stripes: {table.tableStyleInfo.showRowStripes}")
                    print(f"  Column Stripes: {table.tableStyleInfo.showColumnStripes}")
                    print(f"  First Column Highlighted: {table.tableStyleInfo.showFirstColumn}")
                    print(f"  Last Column Highlighted: {table.tableStyleInfo.showLastColumn}")
        else:
            print("⚠️ No tables found - file needs formatting")
        
        wb.close()
        print("=" * 80)
        
    except Exception as e:
        print(f"❌ Error reading table info: {str(e)}")

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Format Manager Complete Notifications Excel as Table')
    parser.add_argument('--info', action='store_true', help='Show current table information')
    parser.add_argument('--format', action='store_true', help='Format file as table')
    
    args = parser.parse_args()
    
    if args.info:
        show_current_table_info()
    elif args.format or len(sys.argv) == 1:  # Default action
        success = main()
        sys.exit(0 if success else 1)
    else:
        parser.print_help()
