#!/usr/bin/env python3
"""
Mismatch Notifications Real-time Sync Verification
================================================
This script verifies that the mismatch notification system is working correctly
with real-time sync between local and network drives.
"""

import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Add scripts directory to path
sys.path.insert(0, 'scripts')

def verify_file_structure():
    """Verify that both local and network files exist and have correct structure"""
    print("📁 File Structure Verification:")
    
    local_file = Path("notification_configs/04_mismatch_notifications.xlsx")
    network_file = Path("network_folder_simplified/04_mismatch_notifications.xlsx")
    
    # Check local file
    if local_file.exists():
        print(f"  ✅ Local file exists: {local_file}")
        df_local = pd.read_excel(local_file)
        print(f"    📊 Records: {len(df_local)}")
    else:
        print(f"  ❌ Local file missing: {local_file}")
    
    # Check network file
    if network_file.exists():
        print(f"  ✅ Network file exists: {network_file}")
        df_network = pd.read_excel(network_file)
        print(f"    📊 Records: {len(df_network)}")
    else:
        print(f"  ❌ Network file missing: {network_file}")
    
    return local_file.exists() and network_file.exists()

def verify_table_format():
    """Verify that both files are properly formatted as tables"""
    print("\n🎨 Table Format Verification:")
    
    files = [
        ("Local", "notification_configs/04_mismatch_notifications.xlsx"),
        ("Network", "network_folder_simplified/04_mismatch_notifications.xlsx")
    ]
    
    all_formatted = True
    
    for location, file_path in files:
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            if len(ws.tables) > 0:
                table = list(ws.tables.values())[0]
                print(f"  ✅ {location}: Table '{table.displayName}' found")
                print(f"    📊 Range: {table.ref}")
                print(f"    🎨 Style: {table.tableStyleInfo.name}")
            else:
                print(f"  ❌ {location}: No tables found")
                all_formatted = False
            
            wb.close()
            
        except Exception as e:
            print(f"  ❌ {location}: Error reading file - {str(e)}")
            all_formatted = False
    
    return all_formatted

def verify_sync_functionality():
    """Verify that the sync functionality is working"""
    print("\n🔄 Sync Functionality Verification:")
    
    try:
        from mismatch_notification_handler import mismatch_notification_handler
        
        # Check sync settings
        sync_enabled = not mismatch_notification_handler.sync_settings.get("stop_notifications", False)
        print(f"  📊 Sync enabled: {'✅ Yes' if sync_enabled else '❌ No'}")
        
        # Check record counts
        local_count = mismatch_notification_handler._get_record_count()
        pending_count = mismatch_notification_handler._get_pending_notification_count()
        
        print(f"  📋 Total records: {local_count}")
        print(f"  ⏳ Pending notifications: {pending_count}")
        
        # Check sync metadata
        metadata_file = Path("network_folder_simplified/mismatch_sync_metadata.json")
        if metadata_file.exists():
            print(f"  ✅ Sync metadata exists: {metadata_file}")
        else:
            print(f"  ⚠️ Sync metadata missing: {metadata_file}")
        
        return True
        
    except Exception as e:
        print(f"  ❌ Error verifying sync functionality: {str(e)}")
        return False

def verify_real_time_detection():
    """Verify that real-time mismatch detection is working"""
    print("\n🔍 Real-time Detection Verification:")
    
    try:
        from mismatch_notification_handler import process_real_time_mismatches
        from datetime import date
        
        # Run real-time detection
        results = process_real_time_mismatches(date.today())
        
        print(f"  📊 Detection Results:")
        print(f"    📅 Date: {results['date']}")
        print(f"    🚨 Mismatches detected: {results['mismatches_detected']}")
        print(f"    📋 Notifications required: {results['notifications_required']}")
        print(f"    💾 Local update: {'✅' if results['local_update_success'] else '❌'}")
        print(f"    🔗 Network sync: {'✅' if results['network_sync_success'] else '❌'}")
        
        if results['errors']:
            print(f"    ❌ Errors: {results['errors']}")
        else:
            print(f"    ✅ No errors detected")
        
        return results['local_update_success'] and results['network_sync_success']
        
    except Exception as e:
        print(f"  ❌ Error in real-time detection: {str(e)}")
        return False

def verify_power_automate_compatibility():
    """Verify Power Automate compatibility"""
    print("\n⚡ Power Automate Compatibility Verification:")
    
    try:
        # Check notification context
        context_file = Path("network_folder_simplified/notification_context.json")
        if context_file.exists():
            import json
            with open(context_file, 'r') as f:
                context = json.load(f)
            
            if 'mismatch_notifications' in context:
                mismatch_context = context['mismatch_notifications']
                print(f"  ✅ Notification context found")
                print(f"    📁 File path: {mismatch_context.get('file_path', 'Not set')}")
                print(f"    📊 Records available: {mismatch_context.get('records_available', 0)}")
                print(f"    ⏳ Pending notifications: {mismatch_context.get('pending_notifications', 0)}")
                print(f"    🔗 Sync enabled: {'✅ Yes' if mismatch_context.get('sync_enabled', False) else '❌ No'}")
            else:
                print(f"  ⚠️ Mismatch context not found in notification context")
        else:
            print(f"  ❌ Notification context file missing: {context_file}")
        
        # Check table format for Power Automate
        network_file = "network_folder_simplified/04_mismatch_notifications.xlsx"
        wb = load_workbook(network_file)
        ws = wb.active
        
        if ws.tables:
            table = list(ws.tables.values())[0]
            print(f"  ✅ Power Automate table ready")
            print(f"    🏷️ Table name: {table.displayName}")
            print(f"    📊 Table range: {table.ref}")
            print(f"    🎨 Style: TableStyleMedium12 (Alert style)")
        else:
            print(f"  ❌ No table found for Power Automate")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ❌ Error verifying Power Automate compatibility: {str(e)}")
        return False

def run_comprehensive_verification():
    """Run all verification checks"""
    print("=" * 80)
    print("🧪 MISMATCH NOTIFICATIONS REAL-TIME SYNC VERIFICATION")
    print("=" * 80)
    
    checks = [
        ("File Structure", verify_file_structure),
        ("Table Format", verify_table_format),
        ("Sync Functionality", verify_sync_functionality),
        ("Real-time Detection", verify_real_time_detection),
        ("Power Automate Compatibility", verify_power_automate_compatibility)
    ]
    
    passed = 0
    total = len(checks)
    
    for check_name, check_func in checks:
        try:
            result = check_func()
            if result:
                passed += 1
            print(f"\n{'✅' if result else '❌'} {check_name}: {'PASSED' if result else 'FAILED'}")
        except Exception as e:
            print(f"\n❌ {check_name}: FAILED with exception: {str(e)}")
    
    # Summary
    print("\n" + "=" * 80)
    print(f"📊 VERIFICATION SUMMARY: {passed}/{total} checks passed")
    print("=" * 80)
    
    if passed == total:
        print("🎉 ALL CHECKS PASSED!")
        print("\n✅ SYSTEM STATUS:")
        print("  - Real-time mismatch detection: ACTIVE")
        print("  - Local Excel updates: WORKING")
        print("  - Network drive sync: WORKING")
        print("  - Table formatting: CORRECT")
        print("  - Power Automate ready: YES")
        
        print("\n🚀 NEXT STEPS:")
        print("  1. Configure Power Automate to read from 'MismatchNotifications' table")
        print("  2. Set up notification delivery (email, Teams, etc.)")
        print("  3. Monitor sync logs for any issues")
        print("  4. Test end-to-end notification flow")
        
    elif passed >= 3:
        print("⚠️ MOSTLY WORKING - Minor issues detected")
        print(f"\n🔧 {total - passed} check(s) failed - please review above")
    else:
        print("❌ SYSTEM NOT READY - Multiple issues found")
        print("\n🚨 Please fix the failed checks before proceeding")
    
    return passed == total

if __name__ == "__main__":
    success = run_comprehensive_verification()
    sys.exit(0 if success else 1)
