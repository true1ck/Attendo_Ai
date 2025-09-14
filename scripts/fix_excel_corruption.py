#!/usr/bin/env python3
"""
Fix corrupted Excel files by recreating them with clean Power Automate format
"""

import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def recreate_excel_file(file_path: Path) -> bool:
    """
    Recreate an Excel file with clean Power Automate format to fix corruption issues.
    """
    try:
        # Required columns for Power Automate
        required_columns = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 'Priority']
        
        # Try to read existing data if possible
        existing_data = []
        try:
            wb_old = openpyxl.load_workbook(file_path, data_only=True)
            ws_old = wb_old.active
            
            # Read existing data if available
            if ws_old.max_row > 1:
                for row in ws_old.iter_rows(min_row=2, max_row=ws_old.max_row, values_only=True):
                    if row and any(cell is not None for cell in row[:5]):  # Check first 5 columns
                        # Ensure we have exactly 5 columns
                        row_data = list(row[:5]) if len(row) >= 5 else list(row) + [None] * (5 - len(row))
                        existing_data.append(row_data)
            wb_old.close()
        except Exception as e:
            logger.warning(f"Could not read existing data from {file_path.name}: {e}")
        
        # Create a new clean workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NotificationData"
        
        # Add headers
        for col_idx, header in enumerate(required_columns, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add existing data if available
        if existing_data:
            for row_idx, row_data in enumerate(existing_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    if value is not None:
                        ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            # Add sample data if no existing data
            sample_data = get_sample_data_for_file(file_path.name)
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Only add table if there's data
        if ws.max_row > 1:
            # Create a simple table without complex naming
            table_range = f"A1:E{ws.max_row}"
            table_name = f"NotificationTable"
            
            # Remove any existing tables (shouldn't be any in new workbook)
            for existing_table in list(ws.tables.keys()):
                del ws.tables[existing_table]
            
            # Create new table
            table = Table(displayName=table_name, ref=table_range)
            
            # Apply simple table style
            style = TableStyleInfo(
                name="TableStyleLight9",  # Use a lighter style
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False  # Disable column stripes
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the clean file
        wb.save(file_path)
        wb.close()
        
        logger.info(f"✅ Successfully recreated {file_path.name}")
        return True
        
    except Exception as e:
        logger.error(f"❌ Failed to recreate {file_path.name}: {e}")
        return False

def get_sample_data_for_file(filename: str) -> list:
    """
    Get appropriate sample data based on filename.
    """
    filename_lower = filename.lower()
    
    if 'daily' in filename_lower:
        return [
            ['EMP001', 'john.doe@company.com', 'Please submit your daily timesheet', 'Daily Reminder', 'High'],
            ['EMP002', 'jane.smith@company.com', 'Timesheet submission pending', 'Daily Reminder', 'High'],
        ]
    elif 'manager_summary' in filename_lower:
        return [
            ['MGR001', 'manager1@company.com', 'Team timesheet summary for review', 'Manager Summary', 'Medium'],
        ]
    elif 'mismatch' in filename_lower:
        return [
            ['EMP003', 'bob.jones@company.com', 'Timesheet hours mismatch detected', 'Mismatch Alert', 'High'],
        ]
    elif 'late' in filename_lower:
        return [
            ['EMP004', 'alice.wong@company.com', 'Timesheet submission is overdue', 'Late Submission', 'High'],
        ]
    else:
        # Default sample data
        return [
            ['EMP000', 'admin@company.com', f'Notification from {filename}', 'General', 'Medium'],
        ]

def fix_all_excel_files_in_directory(directory_path: Path) -> dict:
    """
    Fix all Excel files in a directory by recreating them.
    """
    if not directory_path.exists():
        return {'success': False, 'message': f'Directory not found: {directory_path}'}
    
    # Find all Excel files (excluding temp files)
    excel_files = [
        f for f in directory_path.glob("*.xlsx") 
        if not f.name.startswith("~$") 
        and not f.name.startswith("PA_backup_")
    ]
    
    if not excel_files:
        return {'success': True, 'message': 'No Excel files found to fix'}
    
    logger.info(f"🔧 Fixing {len(excel_files)} Excel files in {directory_path}...")
    
    success_count = 0
    failed_files = []
    
    for file_path in excel_files:
        if recreate_excel_file(file_path):
            success_count += 1
        else:
            failed_files.append(file_path.name)
    
    return {
        'success': success_count > 0,
        'message': f'Fixed {success_count}/{len(excel_files)} files',
        'fixed_count': success_count,
        'total_files': len(excel_files),
        'failed_files': failed_files
    }

if __name__ == "__main__":
    import sys
    
    # Fix files in the network folder
    network_folder = Path('G:/Test')
    
    print(f"🔧 Fixing Excel files in: {network_folder}")
    result = fix_all_excel_files_in_directory(network_folder)
    
    print(f"\n📊 Results:")
    print(f"   Success: {result['success']}")
    print(f"   Message: {result['message']}")
    print(f"   Files fixed: {result['fixed_count']}/{result['total_files']}")
    
    if result.get('failed_files'):
        print(f"   Failed files:")
        for file in result['failed_files']:
            print(f"      - {file}")
    
    print("\n✅ Excel file fixing completed!")
    sys.exit(0 if result['success'] else 1)
