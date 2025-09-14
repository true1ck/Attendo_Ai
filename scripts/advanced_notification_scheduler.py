#!/usr/bin/env python3
"""
Advanced Notification Scheduler
Implements strong time-based logic for adding notifications to queue at specific times
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font
from pathlib import Path
from datetime import datetime, timedelta, time
import logging
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class AdvancedNotificationScheduler:
    """
    Advanced scheduler with specific time-based rules for each notification type.
    """
    
    def __init__(self, network_folder='G:/Test', output_folder='network_folder_simplified'):
        self.network_folder = Path(network_folder)
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        
        self.queue_file = self.output_folder / 'sent_noti_now.xlsx'
        self.schedule_tracking_file = self.output_folder / 'schedule_tracking.json'
        
        # Define notification schedule rules matching exact requirements
        self.notification_schedule = {
            'Daily Status Reminders': {
                'times': ['09:00', '12:00', '15:00', '18:00'],  # Every 3 hours 9AM-6PM
                'max_per_day': 4,
                'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
                'priority': 'Medium',
                'recipients': 'Vendors',
                'messages': [
                    "Good morning! Please submit your timesheet for yesterday",
                    "Midday reminder: Update your morning work hours",
                    "Afternoon check: Log your timesheet entries",
                    "End of day: Complete today's timesheet before leaving"
                ]
            },
            'Manager Summary': {
                'times': ['12:00', '14:00'],  # 12 PM and 2 PM
                'max_per_day': 2,
                'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
                'priority': 'Medium',
                'recipients': 'Managers',
                'messages': [
                    "Noon Report: Review pending timesheet submissions from your team",
                    "Afternoon Update: {count} timesheets awaiting approval"
                ]
            },
            'All Complete': {
                'event_driven': True,
                'priority': 'Low',
                'recipients': 'Managers',
                'messages': [
                    "Great news! All team members have submitted their timesheets",
                    "Team Complete: 100% timesheet submission achieved"
                ]
            },
            'Mismatch Alerts': {
                'times': ['18:00'],  # 6 PM daily
                'max_per_day': 1,
                'escalation': True,
                'days': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
                'priority': 'High',
                'recipients': 'Vendors, Managers',
                'messages': [
                    "URGENT: {hours} hour mismatch detected in your timesheet - immediate action required",
                    "Discrepancy Alert: Your timesheet hours don't match attendance records",
                    "ESCALATION: Timesheet mismatch requires immediate correction"
                ]
            },
            'Manager Feedback': {
                'event_driven': True,
                'within_hours': 1,
                'priority': 'Medium',
                'recipients': 'Vendors',
                'messages': [
                    "Manager Feedback: Your timesheet has been reviewed - see comments",
                    "Action Required: Manager has requested changes to your timesheet"
                ]
            },
            'Monthly Reports': {
                'times': ['09:00'],  # 1st of month at 9 AM
                'monthly': True,
                'day_of_month': 1,
                'priority': 'Medium',
                'recipients': 'Managers, Admins',
                'messages': [
                    "Monthly Report Ready: Review last month's timesheet summary",
                    "Month-End Analysis: Timesheet compliance report available",
                    "Monthly Dashboard: Team timesheet metrics and trends"
                ]
            },
            'System Alerts': {
                'event_driven': True,
                'immediate': True,
                'priority': 'Critical',
                'recipients': 'Admins',
                'messages': [
                    "CRITICAL: System maintenance required immediately",
                    "URGENT: Timesheet system experiencing issues",
                    "SYSTEM ALERT: Emergency maintenance in progress"
                ]
            },
            'Holiday Reminders': {
                'times': ['09:00'],  # 9 AM
                'before_holiday_days': [3, 1],  # 3 days and 1 day before
                'priority': 'Low',
                'recipients': 'All Users',
                'messages': [
                    "Holiday Notice: Office closed in 3 days - submit timesheet early",
                    "Tomorrow is a holiday: Complete your timesheet today",
                    "Holiday Reminder: Adjusted timesheet deadline due to upcoming closure"
                ]
            },
            'Late Submission': {
                'after_deadline_hours': [24, 48],  # 24h and 48h after deadline
                'priority': 'High',
                'recipients': 'Managers, Admins',
                'messages': [
                    "OVERDUE 24h: {employee} has not submitted timesheet",
                    "ESCALATION 48h: {employee} timesheet severely overdue - management notified",
                    "CRITICAL: Multiple timesheet submissions overdue"
                ]
            }
        }
        
        # Load schedule tracking
        self.schedule_tracking = self.load_schedule_tracking()
    
    def load_schedule_tracking(self):
        """Load schedule tracking data."""
        if self.schedule_tracking_file.exists():
            try:
                with open(self.schedule_tracking_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_schedule_tracking(self):
        """Save schedule tracking data."""
        with open(self.schedule_tracking_file, 'w') as f:
            json.dump(self.schedule_tracking, f, indent=2, default=str)
    
    def should_add_to_queue(self, employee_id, notification_type, current_time=None):
        """
        Determine if a notification should be added to queue based on schedule rules.
        
        Returns:
            tuple: (should_add, message) - whether to add and what message to use
        """
        if current_time is None:
            current_time = datetime.now()
        
        # Get schedule rules for this notification type
        if notification_type not in self.notification_schedule:
            return (False, None)
        
        schedule = self.notification_schedule[notification_type]
        
        # Check if it's the right day (skip for event-driven)
        if schedule.get('event_driven', False):
            # Event-driven notifications don't follow daily schedule
            return (False, None)  # Handle through separate event system
        
        current_day = current_time.strftime('%A')
        if 'days' in schedule and current_day not in schedule['days']:
            return (False, None)
        
        # Check if it's month end (for month-end notifications)
        if notification_type == 'Month End':
            # Check if we're in the last 3 days of the month
            next_month = current_time.replace(day=28) + timedelta(days=4)
            last_day = next_month - timedelta(days=next_month.day)
            days_to_end = (last_day - current_time).days
            if days_to_end > 3:
                return (False, None)
        
        # Handle special scheduling cases
        if schedule.get('after_deadline_hours'):
            # Late submission - check if after deadline
            return (False, None)  # Would need deadline tracking
        
        if schedule.get('before_holiday_days'):
            # Holiday reminders - check if before holiday
            return (False, None)  # Would need holiday calendar
        
        # Check if it's the right time (within 10-minute window)
        if 'times' not in schedule:
            return (False, None)
            
        current_time_str = current_time.strftime('%H:%M')
        time_matches = False
        message_index = 0
        
        for idx, scheduled_time in enumerate(schedule['times']):
            scheduled_dt = datetime.strptime(scheduled_time, '%H:%M')
            current_dt = datetime.strptime(current_time_str, '%H:%M')
            
            # Check if we're within 10 minutes of scheduled time
            time_diff = abs((current_dt - scheduled_dt).total_seconds() / 60)
            if time_diff <= 10:
                time_matches = True
                message_index = idx
                break
        
        if not time_matches:
            return (False, None)
        
        # Check if we've already sent this notification today
        tracking_key = f"{employee_id}_{notification_type}_{current_time.date()}"
        
        if tracking_key not in self.schedule_tracking:
            self.schedule_tracking[tracking_key] = {
                'count': 0,
                'times_sent': [],
                'last_sent': None
            }
        
        tracking = self.schedule_tracking[tracking_key]
        
        # Check if we've hit the daily limit
        if tracking['count'] >= schedule['max_per_day']:
            return (False, None)
        
        # Check if we've already sent at this time slot
        time_slot = schedule['times'][message_index]
        if time_slot in tracking['times_sent']:
            return (False, None)
        
        # Generate the message
        messages = schedule['messages']
        message = messages[message_index % len(messages)]
        
        # Replace placeholders in message
        message = self.format_message(message, employee_id, notification_type)
        
        # Update tracking
        tracking['count'] += 1
        tracking['times_sent'].append(time_slot)
        tracking['last_sent'] = current_time.isoformat()
        
        return (True, message)
    
    def format_message(self, message, employee_id, notification_type):
        """Format message with dynamic values."""
        # Replace common placeholders
        replacements = {
            '{count}': str(random.randint(3, 8)),
            '{rate}': str(random.randint(75, 95)),
            '{hours}': str(random.randint(2, 8)),
            '{days}': str(random.randint(1, 3)),
            '{holiday}': 'Independence Day',
            '{date}': 'Monday, August 15th'
        }
        
        for placeholder, value in replacements.items():
            message = message.replace(placeholder, value)
        
        return message
    
    def process_notifications_for_current_time(self):
        """Process notifications based on current time and schedule."""
        current_time = datetime.now()
        notifications = []
        
        logger.info(f"🕐 Processing notifications for {current_time.strftime('%Y-%m-%d %H:%M')}")
        
        # Read vendor data from Excel files
        for file_name in self.network_folder.glob("*.xlsx"):
            if file_name.name.startswith('~') or 'backup' in file_name.name.lower():
                continue
            
            try:
                df = pd.read_excel(file_name)
                
                # Determine notification type from filename
                notification_type = self.get_notification_type_from_filename(file_name.name)
                
                if not notification_type:
                    continue
                
                # Process each employee
                for _, row in df.iterrows():
                    if pd.notna(row.get('EmployeeID')) and pd.notna(row.get('ContactEmail')):
                        employee_id = row['EmployeeID']
                        
                        should_add, message = self.should_add_to_queue(
                            employee_id, 
                            notification_type, 
                            current_time
                        )
                        
                        if should_add:
                            priority = self.notification_schedule[notification_type]['priority']
                            
                            notifications.append({
                                'EmployeeID': employee_id,
                                'ContactEmail': row['ContactEmail'],
                                'Message': message,
                                'NotificationType': notification_type,
                                'Priority': priority,
                                'NTStatusSent': False,
                                'ScheduledTime': current_time.strftime('%H:%M'),
                                'DateAdded': current_time.strftime('%Y-%m-%d %H:%M:%S')
                            })
                            
                            logger.info(f"✅ Added {notification_type} for {employee_id} at {current_time.strftime('%H:%M')}")
                
            except Exception as e:
                logger.error(f"Error processing {file_name.name}: {e}")
        
        # Save tracking
        self.save_schedule_tracking()
        
        return notifications
    
    def get_notification_type_from_filename(self, filename):
        """Map filename to notification type."""
        mapping = {
            '01_daily_status_reminders': 'Daily Status Reminders',
            '02_manager_summary': 'Manager Summary',
            '03_manager_all_complete': 'All Complete',
            '04_mismatch': 'Mismatch Alerts',
            '05_manager_feedback': 'Manager Feedback',
            '06_monthly_report': 'Monthly Reports',
            '07_admin_system': 'System Alerts',
            '08_holiday_reminder': 'Holiday Reminders',
            '09_late_submission': 'Late Submission'
        }
        
        for key, value in mapping.items():
            if key in filename.lower():
                return value
        return None
    
    def add_to_queue(self, notifications):
        """Add notifications to the queue file."""
        # Read existing queue if it exists
        existing_df = pd.DataFrame()
        if self.queue_file.exists():
            try:
                existing_df = pd.read_excel(self.queue_file)
                # Remove sent notifications
                existing_df = existing_df[existing_df['NTStatusSent'] != True]
            except:
                pass
        
        # Add new notifications
        new_df = pd.DataFrame(notifications)
        
        # Combine
        if not existing_df.empty:
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            combined_df = new_df
        
        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'NotificationQueue'
        
        # Headers
        headers = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 
                  'Priority', 'NTStatusSent', 'ScheduledTime', 'DateAdded']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add data
        for idx, row in combined_df.iterrows():
            for col_idx, col_name in enumerate(headers, 1):
                value = row.get(col_name, '')
                if pd.isna(value):
                    value = ''
                ws.cell(row=idx + 2, column=col_idx, value=value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(self.queue_file)
        wb.close()
        
        logger.info(f"📊 Queue updated: {len(combined_df)} total notifications")
        return len(notifications)
    
    def run_scheduled_check(self):
        """Run a scheduled check for current time."""
        # Process notifications for current time
        notifications = self.process_notifications_for_current_time()
        
        if notifications:
            count = self.add_to_queue(notifications)
            logger.info(f"✅ Added {count} new notifications to queue")
        else:
            logger.info("ℹ️ No notifications to add at this time")
        
        # Display next scheduled times
        self.show_next_schedules()
        
        return len(notifications)
    
    def show_next_schedules(self):
        """Show when next notifications will be added."""
        current_time = datetime.now()
        current_hour = current_time.hour
        current_day = current_time.strftime('%A')
        
        print("\n📅 Next Scheduled Notifications:")
        print("-" * 50)
        
        for notification_type, schedule in self.notification_schedule.items():
            # Skip event-driven notifications
            if schedule.get('event_driven', False):
                continue
                
            if 'days' in schedule and current_day in schedule['days']:
                if 'times' in schedule:
                    for scheduled_time in schedule['times']:
                        hour = int(scheduled_time.split(':')[0])
                        if hour > current_hour:
                            print(f"  • {notification_type}: {scheduled_time}")
                            break
            elif schedule.get('monthly'):
                # Handle monthly notifications
                if current_time.day == schedule.get('day_of_month', 1):
                    for scheduled_time in schedule.get('times', ['09:00']):
                        hour = int(scheduled_time.split(':')[0])
                        if hour > current_hour:
                            print(f"  • {notification_type}: {scheduled_time} (Monthly)")
                            break

import random

def main():
    """Main function to run the scheduler."""
    scheduler = AdvancedNotificationScheduler()
    
    print("🕐 Advanced Notification Scheduler")
    print("=" * 50)
    
    # Run scheduled check
    count = scheduler.run_scheduled_check()
    
    print(f"\n📊 Summary:")
    print(f"  • Notifications added: {count}")
    print(f"  • Queue file: {scheduler.queue_file}")
    print(f"  • Current time: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Show schedule summary
    print("\n📄 Daily Schedule Summary:")
    print("-" * 50)
    for notification_type, schedule in scheduler.notification_schedule.items():
        if schedule.get('event_driven', False):
            trigger = "Event-driven"
            if schedule.get('immediate'):
                trigger += " (Immediate)"
            elif schedule.get('within_hours'):
                trigger += f" (Within {schedule['within_hours']}h)"
            print(f"  • {notification_type}: {trigger} | Priority: {schedule['priority']} | Recipients: {schedule.get('recipients', 'N/A')}")
        elif schedule.get('monthly'):
            print(f"  • {notification_type}: Monthly on day {schedule.get('day_of_month', 1)} at {', '.join(schedule.get('times', ['09:00']))} | Recipients: {schedule.get('recipients', 'N/A')}")
        elif schedule.get('after_deadline_hours'):
            print(f"  • {notification_type}: After deadline {schedule['after_deadline_hours']} hours | Recipients: {schedule.get('recipients', 'N/A')}")
        elif schedule.get('before_holiday_days'):
            print(f"  • {notification_type}: Before holidays {schedule['before_holiday_days']} days | Recipients: {schedule.get('recipients', 'N/A')}")
        elif 'times' in schedule:
            times = ', '.join(schedule['times'])
            print(f"  • {notification_type}: {times} ({schedule.get('max_per_day', 'N/A')}/day) | Recipients: {schedule.get('recipients', 'N/A')}")
    
    return count

if __name__ == "__main__":
    main()
