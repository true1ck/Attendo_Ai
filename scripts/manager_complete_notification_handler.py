#!/usr/bin/env python3
"""
Manager All Complete Notification Handler
=========================================
This module handles updating 03_manager_all_complete_notifications.xlsx when all employees 
under a manager complete their submissions. It includes:

1. Detection of manager team completion
2. Local Excel file updates
3. Network drive synchronization with sync features intact
4. Integration with existing notification system

Features:
- Stop notifications flag support
- Force update capabilities
- Retry mechanism
- Proper logging and error handling
- Network drive sync with conflict resolution
"""

import os
import pandas as pd
import shutil
import json
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
import sqlite3
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class ManagerCompletionData:
    """Data structure for manager completion information"""
    manager_id: str
    manager_name: str
    team_size: int
    completion_rate: float
    attendance_date: date
    all_completed: bool
    notification_required: bool
    stop_notifications: bool = False
    force_update: bool = False
    priority: str = "HIGH"

class ManagerCompleteNotificationHandler:
    """Handles manager all-complete notifications with Excel and network sync"""
    
    def __init__(self, 
                 db_path: str = "vendor_timesheet.db",
                 local_excel_folder: str = "notification_configs",
                 network_folder: str = "network_folder_simplified"):
        
        self.db_path = Path(db_path)
        self.local_excel_folder = Path(local_excel_folder)
        self.network_folder = Path(network_folder)
        
        # Excel file paths
        self.local_excel_file = self.local_excel_folder / "03_manager_all_complete_notifications.xlsx"
        self.network_excel_file = self.network_folder / "03_manager_all_complete_notifications.xlsx"
        
        # Sync control files
        self.sync_control_file = self.network_folder / "manager_complete_sync_control.json"
        self.notification_context_file = self.network_folder / "notification_context.json"
        
        # Ensure folders exist
        self.local_excel_folder.mkdir(exist_ok=True, parents=True)
        self.network_folder.mkdir(exist_ok=True, parents=True)
        
        # Load sync settings
        self.sync_settings = self._load_sync_settings()
        
    def _load_sync_settings(self) -> Dict[str, Any]:
        """Load sync control settings from network folder"""
        try:
            if self.sync_control_file.exists():
                with open(self.sync_control_file, 'r') as f:
                    settings = json.load(f)
                logger.info(f"✅ Loaded sync settings: {len(settings)} configurations")
                return settings
            else:
                # Create default sync settings
                default_settings = {
                    "global_sync_enabled": True,
                    "stop_notifications": False,
                    "force_update_mode": False,
                    "auto_retry_enabled": True,
                    "max_retry_count": 3,
                    "sync_interval_minutes": 10,
                    "manager_settings": {},
                    "last_sync_time": datetime.now().isoformat()
                }
                self._save_sync_settings(default_settings)
                return default_settings
        except Exception as e:
            logger.error(f"❌ Error loading sync settings: {str(e)}")
            return {"global_sync_enabled": True, "stop_notifications": False}
    
    def _save_sync_settings(self, settings: Dict[str, Any]):
        """Save sync control settings to network folder"""
        try:
            settings["last_updated"] = datetime.now().isoformat()
            with open(self.sync_control_file, 'w') as f:
                json.dump(settings, f, indent=2)
            logger.info("✅ Sync settings saved")
        except Exception as e:
            logger.error(f"❌ Error saving sync settings: {str(e)}")
    
    def detect_manager_completions(self, target_date: date = None) -> List[ManagerCompletionData]:
        """
        Detect managers whose teams have completed all submissions for the target date
        
        Args:
            target_date: Date to check completions for (defaults to today)
            
        Returns:
            List of ManagerCompletionData objects for managers with completed teams
        """
        if target_date is None:
            target_date = date.today()
        
        logger.info(f"🔍 Detecting manager completions for {target_date}")
        
        try:
            conn = sqlite3.connect(self.db_path)
            
            # Query to get manager completion statistics
            query = """
            SELECT 
                m.manager_id,
                m.full_name as manager_name,
                COUNT(v.id) as team_size,
                COUNT(ds.id) as submitted_count,
                SUM(CASE WHEN ds.approval_status = 'APPROVED' THEN 1 ELSE 0 END) as approved_count,
                MAX(ds.submitted_at) as last_submission_time,
                MAX(ds.approved_at) as last_approval_time
            FROM managers m
            LEFT JOIN vendors v ON m.manager_id = v.manager_id
            LEFT JOIN daily_statuses ds ON v.id = ds.vendor_id AND ds.status_date = ?
            WHERE m.manager_id IS NOT NULL
            GROUP BY m.manager_id, m.full_name
            HAVING COUNT(v.id) > 0
            ORDER BY m.manager_id
            """
            
            df = pd.read_sql_query(query, conn, params=[target_date.isoformat()])
            conn.close()
            
            completions = []
            
            for _, row in df.iterrows():
                team_size = int(row['team_size'])
                submitted_count = int(row['submitted_count'] or 0)
                approved_count = int(row['approved_count'] or 0)
                
                # Calculate completion rate
                completion_rate = (submitted_count / team_size * 100) if team_size > 0 else 0
                all_completed = submitted_count == team_size
                
                # Check if notification is required and not already sent
                notification_required = all_completed and not self._already_notified(
                    row['manager_id'], target_date
                )
                
                # Check manager-specific sync settings
                manager_settings = self.sync_settings.get("manager_settings", {}).get(
                    row['manager_id'], {}
                )
                
                completion_data = ManagerCompletionData(
                    manager_id=row['manager_id'],
                    manager_name=row['manager_name'],
                    team_size=team_size,
                    completion_rate=completion_rate,
                    attendance_date=target_date,
                    all_completed=all_completed,
                    notification_required=notification_required,
                    stop_notifications=manager_settings.get('stop_notifications', False),
                    force_update=manager_settings.get('force_update', False),
                    priority="HIGH" if all_completed else "MEDIUM"
                )
                
                if notification_required and not completion_data.stop_notifications:
                    completions.append(completion_data)
                    logger.info(f"✅ Detected completion: {row['manager_name']} - {team_size} members")
            
            logger.info(f"📊 Found {len(completions)} managers with completed teams requiring notifications")
            return completions
            
        except Exception as e:
            logger.error(f"❌ Error detecting manager completions: {str(e)}")
            return []
    
    def _already_notified(self, manager_id: str, target_date: date) -> bool:
        """Check if manager was already notified for the target date"""
        try:
            if not self.local_excel_file.exists():
                return False
            
            df = pd.read_excel(self.local_excel_file)
            
            # Check for existing notification
            existing = df[
                (df['ManagerID'] == manager_id) & 
                (pd.to_datetime(df['AttendanceDate']).dt.date == target_date) &
                (df['NotiStatus'] == True)
            ]
            
            return not existing.empty
            
        except Exception as e:
            logger.error(f"❌ Error checking notification history: {str(e)}")
            return False
    
    def update_local_excel(self, completions: List[ManagerCompletionData]) -> bool:
        """
        Update the local Excel file with completion data
        
        Args:
            completions: List of completion data to update
            
        Returns:
            True if update successful, False otherwise
        """
        if not completions:
            logger.info("ℹ️ No completions to update")
            return True
        
        logger.info(f"📝 Updating local Excel file with {len(completions)} completions")
        
        try:
            # Load existing Excel file or create new one
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                logger.info(f"📖 Loaded existing Excel with {len(df)} records")
            else:
                # Create new DataFrame with proper structure
                df = pd.DataFrame(columns=[
                    'RecordID', 'ManagerID', 'ManagerName', 'NotificationMessage',
                    'CreatedTime', 'NotiStatus', 'NotificationSentTime', 'Priority',
                    'AttendanceDate', 'TeamSize', 'CompletionRate', 'RetryCount'
                ])
                logger.info("📝 Created new Excel file structure")
            
            # Add new completion records
            new_records = []
            current_time = datetime.now()
            
            for completion in completions:
                # Generate unique RecordID
                record_id = f"NOTIF_{completion.manager_id}_{completion.attendance_date.strftime('%Y%m%d')}_{current_time.strftime('%H%M%S')}"
                
                # Create notification message
                notification_message = self._generate_notification_message(completion)
                
                # Create new record
                new_record = {
                    'RecordID': record_id,
                    'ManagerID': completion.manager_id,
                    'ManagerName': completion.manager_name,
                    'NotificationMessage': notification_message,
                    'CreatedTime': current_time,
                    'NotiStatus': False,  # Will be set to True after successful notification
                    'NotificationSentTime': pd.NaT,  # Will be updated when notification sent
                    'Priority': completion.priority,
                    'AttendanceDate': completion.attendance_date,
                    'TeamSize': completion.team_size,
                    'CompletionRate': f"{completion.completion_rate:.1f}%",
                    'RetryCount': 0
                }
                
                new_records.append(new_record)
                logger.info(f"➕ Added record for manager {completion.manager_name}")
            
            # Append new records to DataFrame
            if new_records:
                new_df = pd.DataFrame(new_records)
                df = pd.concat([df, new_df], ignore_index=True)
                
                # Sort by CreatedTime descending (newest first)
                df = df.sort_values('CreatedTime', ascending=False).reset_index(drop=True)
            
            # Save updated DataFrame
            df.to_excel(self.local_excel_file, index=False, engine='openpyxl')
            logger.info(f"✅ Local Excel updated successfully with {len(new_records)} new records")
            
            # Format as table for Power Automate compatibility
            table_success = self._format_excel_as_table(
                self.local_excel_file, 
                "ManagerCompleteNotifications"
            )
            
            if table_success:
                logger.info("✅ Excel file formatted as table")
            else:
                logger.warning("⚠️ Failed to format Excel as table, but data was saved")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error updating local Excel: {str(e)}")
            return False
    
    def _generate_notification_message(self, completion: ManagerCompletionData) -> str:
        """Generate notification message for manager completion"""
        return (f"🎉 Congratulations! Your entire team ({completion.team_size} members) "
                f"has successfully submitted their attendance status for "
                f"{completion.attendance_date.strftime('%B %d, %Y')}. "
                f"100% completion achieved! Excellent leadership! 👏")
    
    def _format_excel_as_table(self, excel_file_path: Path, table_name: str = "ManagerCompleteNotifications") -> bool:
        """Format Excel file as a proper table for Power Automate compatibility"""
        try:
            logger.info(f"🎨 Formatting Excel file as table: {excel_file_path}")
            
            # Load the workbook
            wb = load_workbook(excel_file_path)
            ws = wb.active
            
            # Check if data exists
            if ws.max_row < 2:  # Only header row or empty
                logger.info("ℹ️ No data to format as table")
                return True
            
            # Remove existing tables if any
            tables_to_remove = []
            for table in ws.tables.values():
                tables_to_remove.append(table.name)
            
            for table_name_to_remove in tables_to_remove:
                del ws.tables[table_name_to_remove]
                logger.info(f"🗑️ Removed existing table: {table_name_to_remove}")
            
            # Define the data range
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Create table reference (e.g., "A1:L100")
            table_range = f"A1:{get_column_letter(max_col)}{max_row}"
            
            # Create table
            table = Table(displayName=table_name, ref=table_range)
            
            # Add table style
            style = TableStyleInfo(
                name="TableStyleMedium9",  # Blue table style
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            # Add the table to the worksheet
            ws.add_table(table)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                ws.column_dimensions[column].width = adjusted_width
            
            # Save the workbook
            wb.save(excel_file_path)
            wb.close()
            
            logger.info(f"✅ Excel file formatted as table: {table_name}")
            logger.info(f"📊 Table range: {table_range} ({max_row} rows, {max_col} columns)")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error formatting Excel as table: {str(e)}")
            return False
    
    def sync_to_network_drive(self, preserve_settings: bool = True) -> bool:
        """
        Sync local Excel file to network drive with sync features intact
        
        Args:
            preserve_settings: Whether to preserve existing sync settings
            
        Returns:
            True if sync successful, False otherwise
        """
        logger.info("🔄 Starting network drive sync...")
        
        # Check global sync settings
        if not self.sync_settings.get("global_sync_enabled", True):
            logger.info("⏭️ Sync disabled globally - skipping")
            return True
        
        if self.sync_settings.get("stop_notifications", False):
            logger.info("🛑 Stop notifications flag set - skipping sync")
            return True
        
        try:
            # Ensure network folder exists
            self.network_folder.mkdir(exist_ok=True, parents=True)
            
            # Check for conflicts if network file exists
            conflict_resolved = True
            if self.network_excel_file.exists() and preserve_settings:
                conflict_resolved = self._resolve_sync_conflicts()
            
            if not conflict_resolved:
                logger.warning("⚠️ Sync conflict not resolved - aborting")
                return False
            
            # Copy local file to network drive
            if self.local_excel_file.exists():
                shutil.copy2(self.local_excel_file, self.network_excel_file)
                logger.info("✅ Excel file synced to network drive")
                
                # Format network file as table for Power Automate
                table_success = self._format_excel_as_table(
                    self.network_excel_file, 
                    "ManagerCompleteNotifications"
                )
                
                if table_success:
                    logger.info("✅ Network Excel file formatted as table")
                else:
                    logger.warning("⚠️ Failed to format network Excel as table")
                
                # Update sync metadata
                self._update_sync_metadata()
                
                # Update notification context for Power Automate
                self._update_notification_context()
                
                return True
            else:
                logger.warning("⚠️ Local Excel file not found - nothing to sync")
                return False
                
        except Exception as e:
            logger.error(f"❌ Error syncing to network drive: {str(e)}")
            return False
    
    def _resolve_sync_conflicts(self) -> bool:
        """Resolve conflicts between local and network files"""
        try:
            # Check modification times
            local_mtime = self.local_excel_file.stat().st_mtime if self.local_excel_file.exists() else 0
            network_mtime = self.network_excel_file.stat().st_mtime if self.network_excel_file.exists() else 0
            
            # If force update is enabled, always prefer local
            if self.sync_settings.get("force_update_mode", False):
                logger.info("🔄 Force update mode - using local file")
                return True
            
            # If network file is newer, merge records
            if network_mtime > local_mtime:
                logger.info("🔄 Network file is newer - merging records")
                return self._merge_excel_files()
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error resolving sync conflicts: {str(e)}")
            return False
    
    def _merge_excel_files(self) -> bool:
        """Merge local and network Excel files to avoid data loss"""
        try:
            # Load both files
            local_df = pd.read_excel(self.local_excel_file) if self.local_excel_file.exists() else pd.DataFrame()
            network_df = pd.read_excel(self.network_excel_file) if self.network_excel_file.exists() else pd.DataFrame()
            
            if local_df.empty and network_df.empty:
                return True
            
            # Combine and remove duplicates based on RecordID
            combined_df = pd.concat([network_df, local_df], ignore_index=True)
            merged_df = combined_df.drop_duplicates(subset=['RecordID'], keep='last')
            
            # Sort by CreatedTime descending
            merged_df = merged_df.sort_values('CreatedTime', ascending=False).reset_index(drop=True)
            
            # Save merged result to local file
            merged_df.to_excel(self.local_excel_file, index=False, engine='openpyxl')
            
            logger.info(f"✅ Merged Excel files: {len(merged_df)} total records")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error merging Excel files: {str(e)}")
            return False
    
    def _update_sync_metadata(self):
        """Update sync metadata in network folder"""
        try:
            metadata = {
                "last_sync_time": datetime.now().isoformat(),
                "source": "manager_complete_notification_handler",
                "sync_status": "success",
                "local_file_path": str(self.local_excel_file),
                "network_file_path": str(self.network_excel_file),
                "records_synced": self._get_record_count()
            }
            
            metadata_file = self.network_folder / "manager_complete_sync_metadata.json"
            with open(metadata_file, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            # Update global sync settings
            self.sync_settings["last_sync_time"] = metadata["last_sync_time"]
            self._save_sync_settings(self.sync_settings)
            
        except Exception as e:
            logger.error(f"❌ Error updating sync metadata: {str(e)}")
    
    def _update_notification_context(self):
        """Update notification context for Power Automate integration"""
        try:
            context = {
                "manager_complete_notifications": {
                    "last_updated": datetime.now().isoformat(),
                    "file_path": str(self.network_excel_file),
                    "records_available": self._get_record_count(),
                    "pending_notifications": self._get_pending_notification_count(),
                    "sync_enabled": not self.sync_settings.get("stop_notifications", False)
                }
            }
            
            # Load existing context if it exists
            if self.notification_context_file.exists():
                with open(self.notification_context_file, 'r') as f:
                    existing_context = json.load(f)
                existing_context.update(context)
                context = existing_context
            
            # Save updated context
            with open(self.notification_context_file, 'w') as f:
                json.dump(context, f, indent=2)
            
            logger.info("✅ Notification context updated for Power Automate")
            
        except Exception as e:
            logger.error(f"❌ Error updating notification context: {str(e)}")
    
    def _get_record_count(self) -> int:
        """Get total number of records in local Excel file"""
        try:
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                return len(df)
            return 0
        except:
            return 0
    
    def _get_pending_notification_count(self) -> int:
        """Get number of pending notifications in local Excel file"""
        try:
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                pending = df[df['NotiStatus'] == False]
                return len(pending)
            return 0
        except:
            return 0
    
    def mark_notifications_sent(self, record_ids: List[str]) -> bool:
        """
        Mark notifications as sent in both local and network Excel files
        
        Args:
            record_ids: List of RecordIDs to mark as sent
            
        Returns:
            True if update successful, False otherwise
        """
        logger.info(f"✅ Marking {len(record_ids)} notifications as sent")
        
        try:
            # Update local file
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                
                # Update matching records
                mask = df['RecordID'].isin(record_ids)
                if mask.any():
                    df.loc[mask, 'NotiStatus'] = True
                    df.loc[mask, 'NotificationSentTime'] = datetime.now()
                    
                    # Save updated file
                    df.to_excel(self.local_excel_file, index=False, engine='openpyxl')
                    
                    # Format as table
                    self._format_excel_as_table(self.local_excel_file, "ManagerCompleteNotifications")
                    
                    # Sync to network drive
                    self.sync_to_network_drive()
                    
                    logger.info(f"✅ Updated {mask.sum()} notification records")
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"❌ Error marking notifications as sent: {str(e)}")
            return False
    
    def process_all_manager_completions(self, target_date: date = None) -> Dict[str, Any]:
        """
        Complete workflow: detect completions, update Excel, sync to network
        
        Args:
            target_date: Date to process completions for (defaults to today)
            
        Returns:
            Dictionary with processing results
        """
        if target_date is None:
            target_date = date.today()
        
        logger.info("🚀 Starting manager completion processing workflow")
        
        results = {
            "date": target_date.isoformat(),
            "completions_detected": 0,
            "local_update_success": False,
            "network_sync_success": False,
            "notifications_required": 0,
            "errors": []
        }
        
        try:
            # Step 1: Detect manager completions
            completions = self.detect_manager_completions(target_date)
            results["completions_detected"] = len(completions)
            results["notifications_required"] = len([c for c in completions if c.notification_required])
            
            if not completions:
                logger.info("ℹ️ No manager completions requiring notifications")
                results["local_update_success"] = True
                results["network_sync_success"] = True
                return results
            
            # Step 2: Update local Excel file
            local_success = self.update_local_excel(completions)
            results["local_update_success"] = local_success
            
            if not local_success:
                results["errors"].append("Failed to update local Excel file")
                return results
            
            # Step 3: Sync to network drive
            network_success = self.sync_to_network_drive()
            results["network_sync_success"] = network_success
            
            if not network_success:
                results["errors"].append("Failed to sync to network drive")
            
            # Step 4: Log summary
            logger.info(f"📊 Processing completed:")
            logger.info(f"  - Completions detected: {results['completions_detected']}")
            logger.info(f"  - Notifications required: {results['notifications_required']}")
            logger.info(f"  - Local update: {'✅' if local_success else '❌'}")
            logger.info(f"  - Network sync: {'✅' if network_success else '❌'}")
            
            return results
            
        except Exception as e:
            error_msg = f"Fatal error in manager completion processing: {str(e)}"
            logger.error(f"❌ {error_msg}")
            results["errors"].append(error_msg)
            return results
    
    def get_pending_notifications(self) -> List[Dict[str, Any]]:
        """Get list of pending notifications that need to be sent"""
        try:
            if not self.local_excel_file.exists():
                return []
            
            df = pd.read_excel(self.local_excel_file)
            pending = df[df['NotiStatus'] == False].to_dict('records')
            
            logger.info(f"📋 Found {len(pending)} pending notifications")
            return pending
            
        except Exception as e:
            logger.error(f"❌ Error getting pending notifications: {str(e)}")
            return []
    
    def configure_manager_sync_settings(self, manager_id: str, settings: Dict[str, Any]):
        """Configure sync settings for a specific manager"""
        if "manager_settings" not in self.sync_settings:
            self.sync_settings["manager_settings"] = {}
        
        self.sync_settings["manager_settings"][manager_id] = settings
        self._save_sync_settings(self.sync_settings)
        
        logger.info(f"✅ Updated sync settings for manager {manager_id}")

# Global instance for easy access
manager_complete_handler = ManagerCompleteNotificationHandler()

def process_manager_completions(target_date: date = None) -> Dict[str, Any]:
    """Public function to process manager completions"""
    return manager_complete_handler.process_all_manager_completions(target_date)

def mark_notification_sent(record_ids: List[str]) -> bool:
    """Public function to mark notifications as sent"""
    return manager_complete_handler.mark_notifications_sent(record_ids)

def get_pending_manager_notifications() -> List[Dict[str, Any]]:
    """Public function to get pending notifications"""
    return manager_complete_handler.get_pending_notifications()

def configure_sync_settings(manager_id: str = None, **settings):
    """Public function to configure sync settings"""
    if manager_id:
        manager_complete_handler.configure_manager_sync_settings(manager_id, settings)
    else:
        # Update global settings
        manager_complete_handler.sync_settings.update(settings)
        manager_complete_handler._save_sync_settings(manager_complete_handler.sync_settings)

def convert_to_table_format(excel_file_path: str = None, table_name: str = "ManagerCompleteNotifications") -> bool:
    """Public function to convert existing Excel file to table format"""
    if excel_file_path is None:
        excel_file_path = "notification_configs/03_manager_all_complete_notifications.xlsx"
    
    try:
        file_path = Path(excel_file_path)
        if not file_path.exists():
            print(f"❌ Excel file not found: {excel_file_path}")
            return False
        
        print(f"🎨 Converting Excel file to table format: {excel_file_path}")
        
        # Use the handler's method to format as table
        success = manager_complete_handler._format_excel_as_table(file_path, table_name)
        
        if success:
            print(f"✅ Successfully converted to table format: {table_name}")
            print(f"📊 File ready for Power Automate integration")
        else:
            print(f"❌ Failed to convert to table format")
        
        return success
        
    except Exception as e:
        print(f"❌ Error converting to table format: {str(e)}")
        return False

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Manager Complete Notification Handler')
    parser.add_argument('--date', type=str, help='Date to process (YYYY-MM-DD format)')
    parser.add_argument('--action', choices=['process', 'pending', 'sync', 'status', 'format-table'], 
                        default='process', help='Action to perform')
    parser.add_argument('--manager-id', type=str, help='Specific manager ID to configure')
    parser.add_argument('--stop-notifications', action='store_true', 
                        help='Stop notifications for manager')
    parser.add_argument('--force-update', action='store_true', 
                        help='Force update for manager')
    parser.add_argument('--excel-file', type=str, 
                        help='Excel file path (for format-table action)')
    parser.add_argument('--table-name', type=str, default='ManagerCompleteNotifications',
                        help='Table name for formatting')
    
    args = parser.parse_args()
    
    target_date = None
    if args.date:
        try:
            target_date = datetime.strptime(args.date, '%Y-%m-%d').date()
        except ValueError:
            print("❌ Invalid date format. Use YYYY-MM-DD")
            exit(1)
    
    if args.action == 'process':
        results = process_manager_completions(target_date)
        print("📊 Processing Results:")
        for key, value in results.items():
            print(f"  {key}: {value}")
    
    elif args.action == 'pending':
        pending = get_pending_manager_notifications()
        print(f"📋 Found {len(pending)} pending notifications:")
        for notification in pending:
            print(f"  - {notification['ManagerName']}: {notification['AttendanceDate']}")
    
    elif args.action == 'sync':
        success = manager_complete_handler.sync_to_network_drive()
        print(f"🔄 Network sync: {'✅ Success' if success else '❌ Failed'}")
    
    elif args.action == 'status':
        record_count = manager_complete_handler._get_record_count()
        pending_count = manager_complete_handler._get_pending_notification_count()
        print(f"📊 Status:")
        print(f"  Total records: {record_count}")
        print(f"  Pending notifications: {pending_count}")
        print(f"  Sync enabled: {not manager_complete_handler.sync_settings.get('stop_notifications', False)}")
    
    elif args.action == 'format-table':
        success = convert_to_table_format(args.excel_file, args.table_name)
        print(f"🎨 Table formatting: {'✅ Success' if success else '❌ Failed'}")
    
    # Handle manager-specific configuration
    if args.manager_id:
        settings = {}
        if args.stop_notifications:
            settings['stop_notifications'] = True
        if args.force_update:
            settings['force_update'] = True
        
        if settings:
            configure_sync_settings(args.manager_id, **settings)
            print(f"✅ Updated settings for manager {args.manager_id}")
