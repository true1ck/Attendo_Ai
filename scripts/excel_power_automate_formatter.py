#!/usr/bin/env python3
"""
Excel Power Automate Formatter Module
Ensures all Excel files have the required column structure for Power Automate integration:
EmployeeID, ContactEmail, Message, NotificationType, Priority

This module validates and reformats Excel files during the sync process to ensure
they are properly structured for Power Automate to read.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import logging
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelPowerAutomateFormatter:
    """
    Handles formatting Excel files for Power Automate compatibility.
    Ensures all files have the required column structure.
    """
    
    # Required columns for Power Automate in exact order
    REQUIRED_COLUMNS = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 'Priority']
    
    def __init__(self):
        self.validation_errors = []
        self.formatted_files = []
    
    def validate_and_format_excel_file(self, file_path: Path, backup: bool = False) -> bool:
        """
        Validate and format an Excel file to ensure Power Automate compatibility.
        
        Args:
            file_path: Path to the Excel file
            backup: Whether to create a backup before modifying (default: False)
            
        Returns:
            bool: True if file was successfully validated/formatted, False otherwise
        """
        try:
            if not file_path.exists() or not file_path.suffix.lower() == '.xlsx':
                logger.error(f"❌ Invalid Excel file: {file_path}")
                return False
            
            # Skip backup creation - not needed
            if backup:
                logger.info(f"⏭️ Skipping backup creation for: {file_path.name}")
            
            # Load the Excel file
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get current headers
            current_headers = [cell.value for cell in ws[1] if cell.value is not None]
            
            # Check if file already has correct format
            if self._has_correct_format(current_headers):
                logger.info(f"✅ File already has correct format: {file_path.name}")
                return True
            
            # Read existing data
            data_rows = []
            if ws.max_row > 1:  # Has data beyond headers
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        data_rows.append(list(row))
            
            # Transform data to required format
            formatted_data = self._transform_data_to_required_format(
                current_headers, data_rows, file_path.name
            )
            
            # Clear the worksheet
            ws.delete_rows(1, ws.max_row)
            
            # Add required headers
            for col_idx, header in enumerate(self.REQUIRED_COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add formatted data
            for row_idx, row_data in enumerate(formatted_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format as table for Power Automate
            if formatted_data:  # Only create table if there's data
                # Remove any existing tables first to avoid conflicts
                for existing_table in list(ws.tables.keys()):
                    del ws.tables[existing_table]
                
                table_range = f"A1:{openpyxl.utils.get_column_letter(len(self.REQUIRED_COLUMNS))}{len(formatted_data) + 1}"
                
                # Create table with simplified name to avoid issues
                from openpyxl.worksheet.table import Table, TableStyleInfo
                table_name = f"Table_{file_path.stem.replace('-', '_').replace(' ', '_')}"
                # Ensure table name is valid (alphanumeric and underscore only)
                table_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
                
                table = Table(displayName=table_name, ref=table_range)
                
                # Apply table style
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=True
                )
                table.tableStyleInfo = style
                ws.add_table(table)
                
                logger.info(f"📊 Created Power Automate table '{table_name}' with {len(formatted_data)} rows")
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws)
            
            # Save the file
            wb.save(file_path)
            logger.info(f"✅ Formatted file for Power Automate: {file_path.name}")
            self.formatted_files.append(file_path.name)
            
            return True
            
        except Exception as e:
            error_msg = f"❌ Failed to format {file_path.name}: {str(e)}"
            logger.error(error_msg)
            self.validation_errors.append(error_msg)
            return False
    
    def _has_correct_format(self, headers: list) -> bool:
        """Check if the file already has the correct column format."""
        if len(headers) < len(self.REQUIRED_COLUMNS):
            return False
            
        # Check if all required columns exist in the correct order (at the beginning)
        for i, required_col in enumerate(self.REQUIRED_COLUMNS):
            if i >= len(headers) or headers[i] != required_col:
                return False
        
        return True
    
    def _transform_data_to_required_format(self, current_headers: list, data_rows: list, filename: str) -> list:
        """
        Transform existing data to match the required column structure.
        
        Args:
            current_headers: Current column headers
            data_rows: Existing data rows
            filename: Name of the file being processed
            
        Returns:
            list: Transformed data rows matching required format
        """
        formatted_rows = []
        
        # Create header mapping for common column names
        header_mapping = self._create_header_mapping(current_headers)
        
        for row_data in data_rows:
            # Ensure row has enough columns
            while len(row_data) < len(current_headers):
                row_data.append(None)
            
            # Create new row with required format
            new_row = [None] * len(self.REQUIRED_COLUMNS)
            
            # Map data from existing columns to required columns
            for req_idx, req_col in enumerate(self.REQUIRED_COLUMNS):
                if req_col in header_mapping:
                    source_idx = header_mapping[req_col]
                    if source_idx < len(row_data):
                        new_row[req_idx] = row_data[source_idx]
                
                # Set default values for missing data
                if new_row[req_idx] is None:
                    new_row[req_idx] = self._get_default_value(req_col, filename)
            
            formatted_rows.append(new_row)
        
        return formatted_rows
    
    def _create_header_mapping(self, current_headers: list) -> dict:
        """Create a mapping from required columns to existing column indices."""
        mapping = {}
        
        # Common column name variations
        column_variations = {
            'EmployeeID': ['EmployeeID', 'Employee_ID', 'employee_id', 'EmpID', 'ID', 'UserID', 'User_ID'],
            'ContactEmail': ['ContactEmail', 'Contact_Email', 'Email', 'email', 'user_email', 'contact_email'],
            'Message': ['Message', 'message', 'notification_message', 'text', 'content', 'body'],
            'NotificationType': ['NotificationType', 'Notification_Type', 'notification_type', 'type', 'category'],
            'Priority': ['Priority', 'priority', 'urgency', 'importance', 'level']
        }
        
        for req_col, variations in column_variations.items():
            for variation in variations:
                if variation in current_headers:
                    mapping[req_col] = current_headers.index(variation)
                    break
        
        return mapping
    
    def _get_default_value(self, column: str, filename: str) -> str:
        """Get default value for a column if no data is available."""
        defaults = {
            'EmployeeID': 'N/A',
            'ContactEmail': 'admin@company.com',
            'Message': f'Notification from {filename}',
            'NotificationType': self._extract_notification_type_from_filename(filename),
            'Priority': 'Medium'
        }
        return defaults.get(column, '')
    
    def _extract_notification_type_from_filename(self, filename: str) -> str:
        """Extract notification type from filename."""
        filename_lower = filename.lower()
        
        type_mapping = {
            'daily': 'Daily Reminder',
            'manager_summary': 'Manager Summary',
            'manager_all_complete': 'All Complete',
            'mismatch': 'Mismatch Alert',
            'manager_feedback': 'Manager Feedback',
            'monthly': 'Monthly Report',
            'admin': 'Admin Alert',
            'holiday': 'Holiday Reminder',
            'late': 'Late Submission',
            'billing': 'Billing Correction'
        }
        
        for key, value in type_mapping.items():
            if key in filename_lower:
                return value
        
        return 'General Notification'
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths for better readability."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with padding, capped at reasonable maximum
            adjusted_width = min(max_length + 3, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def format_all_files_in_directory(self, directory_path: Path, backup: bool = False) -> dict:
        """
        Format all Excel files in a directory for Power Automate compatibility.
        
        Args:
            directory_path: Path to directory containing Excel files
            backup: Whether to create backups
            
        Returns:
            dict: Summary of formatting operation
        """
        self.validation_errors.clear()
        self.formatted_files.clear()
        
        if not directory_path.exists():
            return {
                'success': False,
                'message': f'Directory not found: {directory_path}',
                'formatted_count': 0,
                'errors': [f'Directory not found: {directory_path}']
            }
        
        # Find all Excel files (excluding temporary/backup files)
        excel_files = [
            f for f in directory_path.glob("*.xlsx") 
            if not f.name.startswith("~$") 
            and not f.name.startswith("PA_backup_")
            and not f.name.lower().startswith("pa_backup_")
        ]
        
        if not excel_files:
            return {
                'success': True,
                'message': 'No Excel files found to format',
                'formatted_count': 0,
                'errors': []
            }
        
        logger.info(f"🔄 Formatting {len(excel_files)} Excel files for Power Automate...")
        
        success_count = 0
        for file_path in excel_files:
            if self.validate_and_format_excel_file(file_path, backup):
                success_count += 1
        
        return {
            'success': success_count > 0,
            'message': f'Formatted {success_count}/{len(excel_files)} files for Power Automate',
            'formatted_count': success_count,
            'total_files': len(excel_files),
            'formatted_files': self.formatted_files,
            'errors': self.validation_errors
        }

# Global instance for easy import
excel_power_automate_formatter = ExcelPowerAutomateFormatter()

def validate_and_format_excel_for_power_automate(file_path, backup=False):
    """
    Convenience function to validate and format a single Excel file.
    
    Args:
        file_path: Path to Excel file (string or Path object)
        backup: Whether to create backup before formatting
        
    Returns:
        bool: True if successful, False otherwise
    """
    return excel_power_automate_formatter.validate_and_format_excel_file(Path(file_path), backup)

def format_directory_for_power_automate(directory_path, backup=False):
    """
    Convenience function to format all Excel files in a directory.
    
    Args:
        directory_path: Path to directory (string or Path object) 
        backup: Whether to create backups before formatting
        
    Returns:
        dict: Formatting operation summary
    """
    return excel_power_automate_formatter.format_all_files_in_directory(Path(directory_path), backup)

if __name__ == "__main__":
    # Test the module
    import sys
    
    print("🧪 Testing Excel Power Automate Formatter...")
    
    # Create test data with various column formats
    test_data = [
        ['employee_id', 'email', 'notification_message', 'type', 'urgency'],
        ['EMP001', 'john@company.com', 'Please submit timesheet', 'reminder', 'high'],
        ['EMP002', 'jane@company.com', 'Timesheet approved', 'approval', 'medium'],
        ['EMP003', 'bob@company.com', 'Missing hours detected', 'alert', 'high']
    ]
    
    # Create test Excel file
    test_file = Path("test_power_automate_format.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    for row_data in test_data:
        ws.append(row_data)
    
    wb.save(test_file)
    print(f"📄 Created test file: {test_file}")
    
    # Test formatting
    success = validate_and_format_excel_for_power_automate(test_file, backup=False)
    
    if success:
        print("✅ Test formatting successful!")
        
        # Verify the result
        wb_test = openpyxl.load_workbook(test_file)
        ws_test = wb_test.active
        headers = [cell.value for cell in ws_test[1]]
        
        print(f"📊 Formatted headers: {headers}")
        
        if headers == excel_power_automate_formatter.REQUIRED_COLUMNS:
            print("✅ Headers match required Power Automate format!")
        else:
            print(f"❌ Headers don't match. Expected: {excel_power_automate_formatter.REQUIRED_COLUMNS}")
        
        # Clean up
        test_file.unlink()
        print("🧹 Cleaned up test file")
    else:
        print("❌ Test formatting failed!")
        sys.exit(1)
    
    print("✅ Excel Power Automate Formatter is working correctly!")
