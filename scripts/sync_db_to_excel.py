#!/usr/bin/env python3
"""
Sync Database to Excel Files
Fetches notification data from database tables and updates Excel files with proper column mapping
for Power Automate compatibility.
"""

import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from datetime import datetime
import logging
from sqlalchemy import create_engine, text
import os
import sys

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DatabaseToExcelSync:
    """
    Syncs notification data from database tables to Excel files.
    Maps database columns to Power Automate required format.
    """
    
    def __init__(self, db_path='vendor_timesheet.db', excel_folder='notification_configs'):
        """
        Initialize the sync handler.
        
        Args:
            db_path: Path to SQLite database
            excel_folder: Folder containing Excel files
        """
        self.db_path = db_path
        self.excel_folder = Path(excel_folder)
        self.network_folder = Path('G:/Test')
        
        # Create database connection
        self.engine = create_engine(f'sqlite:///{db_path}')
        
        # Column mapping for Power Automate format
        self.required_columns = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 'Priority']
        
        # Database table to Excel file mapping
        self.table_mappings = {
            'daily_status_reminders': {
                'file': '01_daily_status_reminders.xlsx',
                'query': """
                    SELECT 
                        Vendor_ID as EmployeeID,
                        Contact_Email as ContactEmail,
                        Notification_Message as Message,
                        'Daily Reminder' as NotificationType,
                        CASE 
                            WHEN Priority IS NOT NULL THEN Priority
                            ELSE 'Medium'
                        END as Priority
                    FROM daily_status_reminders
                    WHERE Active = 1
                """,
                'notification_type': 'Daily Reminder'
            },
            'manager_summary_notifications': {
                'file': '02_manager_summary_notifications.xlsx',
                'query': """
                    SELECT 
                        Manager_ID as EmployeeID,
                        Contact_Email as ContactEmail,
                        COALESCE(Custom_Message, 'Team timesheet summary for review') as Message,
                        'Manager Summary' as NotificationType,
                        COALESCE(Priority, 'Medium') as Priority
                    FROM manager_summary_notifications
                    WHERE Active = 1
                """,
                'notification_type': 'Manager Summary'
            },
            'mismatch_notifications': {
                'file': '04_mismatch_notifications.xlsx',
                'query': """
                    SELECT 
                        Vendor_ID as EmployeeID,
                        Contact_Email as ContactEmail,
                        Notification_Message as Message,
                        'Mismatch Alert' as NotificationType,
                        COALESCE(Priority, 'High') as Priority
                    FROM mismatch_notifications
                    WHERE Active = 1
                """,
                'notification_type': 'Mismatch Alert'
            },
            'late_submission_alerts': {
                'file': '09_late_submission_alerts.xlsx',
                'query': """
                    SELECT 
                        Vendor_ID as EmployeeID,
                        Contact_Email as ContactEmail,
                        Alert_Message as Message,
                        'Late Submission' as NotificationType,
                        COALESCE(Priority, 'High') as Priority
                    FROM late_submission_alerts
                    WHERE Active = 1
                """,
                'notification_type': 'Late Submission'
            },
            'holiday_reminder_notifications': {
                'file': '08_holiday_reminder_notifications.xlsx',
                'query': """
                    SELECT 
                        Vendor_ID as EmployeeID,
                        Vendor_Email as ContactEmail,
                        Reminder_Message as Message,
                        'Holiday Reminder' as NotificationType,
                        COALESCE(Priority, 'Medium') as Priority
                    FROM holiday_reminder_notifications
                    WHERE Active = 1
                """,
                'notification_type': 'Holiday Reminder'
            }
        }
    
    def fetch_data_from_db(self, query: str) -> pd.DataFrame:
        """
        Fetch data from database using provided query.
        
        Args:
            query: SQL query to execute
            
        Returns:
            DataFrame with query results
        """
        try:
            with self.engine.connect() as conn:
                df = pd.read_sql_query(query, conn)
                return df
        except Exception as e:
            logger.error(f"Database query failed: {e}")
            return pd.DataFrame()
    
    def create_excel_with_data(self, df: pd.DataFrame, file_path: Path, notification_type: str) -> bool:
        """
        Create or update Excel file with data from DataFrame.
        
        Args:
            df: DataFrame containing notification data
            file_path: Path to Excel file
            notification_type: Type of notification for this file
            
        Returns:
            bool: Success status
        """
        try:
            # Ensure DataFrame has all required columns in correct order
            for col in self.required_columns:
                if col not in df.columns:
                    if col == 'EmployeeID':
                        df[col] = 'N/A'
                    elif col == 'ContactEmail':
                        df[col] = 'admin@company.com'
                    elif col == 'Message':
                        df[col] = f'Notification from {notification_type}'
                    elif col == 'NotificationType':
                        df[col] = notification_type
                    elif col == 'Priority':
                        df[col] = 'Medium'
            
            # Reorder columns to match required format
            df = df[self.required_columns]
            
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "NotificationData"
            
            # Add headers
            for col_idx, header in enumerate(self.required_columns, 1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Add data
            for row_idx, row in df.iterrows():
                for col_idx, col_name in enumerate(self.required_columns, 1):
                    value = row[col_name]
                    # Handle None/NaN values
                    if pd.isna(value):
                        value = ''
                    ws.cell(row=row_idx + 2, column=col_idx, value=value)
            
            # Create table if there's data
            if len(df) > 0:
                table_range = f"A1:E{len(df) + 1}"
                table_name = "NotificationTable"
                
                # Remove existing tables
                for existing_table in list(ws.tables.keys()):
                    del ws.tables[existing_table]
                
                # Create new table
                table = Table(displayName=table_name, ref=table_range)
                
                # Apply table style
                style = TableStyleInfo(
                    name="TableStyleLight9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            wb.save(file_path)
            wb.close()
            
            logger.info(f"✅ Created Excel file with {len(df)} rows: {file_path.name}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to create Excel file {file_path.name}: {e}")
            return False
    
    def sync_table_to_excel(self, table_name: str) -> bool:
        """
        Sync a specific database table to its corresponding Excel file.
        
        Args:
            table_name: Name of the database table
            
        Returns:
            bool: Success status
        """
        if table_name not in self.table_mappings:
            logger.error(f"Unknown table: {table_name}")
            return False
        
        mapping = self.table_mappings[table_name]
        
        # Fetch data from database
        logger.info(f"📊 Fetching data from {table_name}...")
        df = self.fetch_data_from_db(mapping['query'])
        
        if df.empty:
            logger.warning(f"No active records found in {table_name}")
            # Create file with sample data
            df = pd.DataFrame([{
                'EmployeeID': 'SAMPLE001',
                'ContactEmail': 'sample@company.com',
                'Message': f"Sample {mapping['notification_type']} notification",
                'NotificationType': mapping['notification_type'],
                'Priority': 'Medium'
            }])
        
        # Save to local folder
        local_file = self.excel_folder / mapping['file']
        success_local = self.create_excel_with_data(df, local_file, mapping['notification_type'])
        
        # Save to network folder
        network_file = self.network_folder / mapping['file']
        success_network = False
        if self.network_folder.exists():
            success_network = self.create_excel_with_data(df, network_file, mapping['notification_type'])
        
        return success_local and success_network
    
    def sync_all_tables(self) -> dict:
        """
        Sync all configured database tables to Excel files.
        
        Returns:
            dict: Summary of sync operation
        """
        results = {
            'success': True,
            'synced_tables': [],
            'failed_tables': [],
            'total_records': 0
        }
        
        logger.info("🔄 Starting database to Excel sync...")
        
        for table_name in self.table_mappings:
            try:
                # Get record count
                count_query = f"SELECT COUNT(*) as count FROM {table_name} WHERE Active = 1"
                with self.engine.connect() as conn:
                    count_result = conn.execute(text(count_query)).fetchone()
                    record_count = count_result[0] if count_result else 0
                
                # Sync table
                if self.sync_table_to_excel(table_name):
                    results['synced_tables'].append(table_name)
                    results['total_records'] += record_count
                    logger.info(f"✅ Synced {table_name} ({record_count} records)")
                else:
                    results['failed_tables'].append(table_name)
                    results['success'] = False
                    logger.error(f"❌ Failed to sync {table_name}")
                    
            except Exception as e:
                logger.error(f"❌ Error syncing {table_name}: {e}")
                results['failed_tables'].append(table_name)
                results['success'] = False
        
        return results
    
    def sync_specific_table(self, table_name: str, custom_query: str = None) -> bool:
        """
        Sync a specific table with optional custom query.
        
        Args:
            table_name: Name of the table to sync
            custom_query: Optional custom SQL query
            
        Returns:
            bool: Success status
        """
        if custom_query:
            # Use custom query
            df = self.fetch_data_from_db(custom_query)
            if not df.empty:
                file_name = self.table_mappings.get(table_name, {}).get('file', f"{table_name}.xlsx")
                notification_type = self.table_mappings.get(table_name, {}).get('notification_type', 'Custom')
                
                local_file = self.excel_folder / file_name
                network_file = self.network_folder / file_name
                
                success_local = self.create_excel_with_data(df, local_file, notification_type)
                success_network = self.create_excel_with_data(df, network_file, notification_type)
                
                return success_local and success_network
        else:
            # Use default mapping
            return self.sync_table_to_excel(table_name)
        
        return False

def main():
    """Main function to run the sync."""
    # Initialize sync handler
    sync = DatabaseToExcelSync()
    
    # Ensure folders exist
    sync.excel_folder.mkdir(exist_ok=True)
    if not sync.network_folder.exists():
        logger.warning(f"Network folder does not exist: {sync.network_folder}")
        sync.network_folder.mkdir(parents=True, exist_ok=True)
    
    # Sync all tables
    results = sync.sync_all_tables()
    
    # Print summary
    print("\n📊 Database to Excel Sync Summary:")
    print(f"   Success: {results['success']}")
    print(f"   Tables synced: {len(results['synced_tables'])}")
    print(f"   Total records: {results['total_records']}")
    
    if results['synced_tables']:
        print("\n   ✅ Synced tables:")
        for table in results['synced_tables']:
            print(f"      - {table}")
    
    if results['failed_tables']:
        print("\n   ❌ Failed tables:")
        for table in results['failed_tables']:
            print(f"      - {table}")
    
    return 0 if results['success'] else 1

if __name__ == "__main__":
    sys.exit(main())
