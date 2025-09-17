"""
Notification Database Sync System
Keeps Excel notification configuration tables synchronized with database changes.
Supports real-time updates when users are added/removed, roles change, or contact info updates.
"""

import os
import sys
import json
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Any
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from sqlalchemy import create_engine, event, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.engine import Engine

# Add the app directory to the path to import models
current_dir = Path(__file__).parent
sys.path.append(str(current_dir))

try:
    from models import User, Vendor, Manager, UserRole
    from app import db
except ImportError as e:
    print(f"Warning: Could not import models: {e}")
    print("This module should be run in the context of the Flask application")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('notification_sync.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class NotificationDatabaseSync:
    """
    Manages synchronization between database changes and Excel notification configuration tables.
    """
    
    def __init__(self, app=None, excel_folder_path="notification_configs"):
        self.app = app
        self.excel_folder_path = Path(excel_folder_path)
        self.excel_folder_path.mkdir(exist_ok=True)
        
        # Excel file configuration mapping
        self.excel_configs = {
            "01_daily_status_reminders.xlsx": {
                "table_name": "DailyStatusReminders",
                "roles": ["VENDOR"],
                "notification_type": "daily_reminder"
            },
            "02_manager_summary_notifications.xlsx": {
                "table_name": "ManagerSummaryNotifications", 
                "roles": ["MANAGER"],
                "notification_type": "manager_summary"
            },
            "03_manager_all_complete_notifications.xlsx": {
                "table_name": "AllCompleteNotifications",
                "roles": ["MANAGER"], 
                "notification_type": "all_complete"
            },
            "04_mismatch_notifications.xlsx": {
                "table_name": "MismatchNotifications",
                "roles": ["MANAGER", "VENDOR"],
                "notification_type": "mismatch_alert"
            },
            "05_manager_feedback_notifications.xlsx": {
                "table_name": "ManagerFeedbackNotifications",
                "roles": ["VENDOR"],
                "notification_type": "feedback"
            },
            "06_monthly_report_notifications.xlsx": {
                "table_name": "MonthlyReportNotifications",
                "roles": ["MANAGER", "ADMIN"],
                "notification_type": "monthly_report"
            },
            "07_admin_system_alerts.xlsx": {
                "table_name": "AdminSystemAlerts",
                "roles": ["ADMIN"],
                "notification_type": "system_alert"
            },
            "08_holiday_reminder_notifications.xlsx": {
                "table_name": "HolidayReminderNotifications",
                "roles": ["VENDOR", "MANAGER", "ADMIN"],
                "notification_type": "holiday_reminder"
            },
            "09_late_submission_alerts.xlsx": {
                "table_name": "LateSubmissionAlerts",
                "roles": ["MANAGER", "ADMIN"],
                "notification_type": "late_submission"
            },
            "10_billing_correction_notifications.xlsx": {
                "table_name": "BillingCorrectionNotifications",
                "roles": ["MANAGER", "ADMIN"],
                "notification_type": "billing_correction"
            }
        }
        
        # Database event listeners setup
        self.setup_database_listeners()
        
    def init_app(self, app):
        """Initialize with Flask app context"""
        self.app = app
        
    def setup_database_listeners(self):
        """Setup SQLAlchemy event listeners for database changes"""
        logger.info("Setting up database event listeners...")
        
        # Listen for database changes that affect notifications
        @event.listens_for(User, 'after_insert')
        def user_inserted(mapper, connection, target):
            logger.info(f"New user created: {target.username} ({target.role.value})")
            self.sync_user_to_excel_tables(target, action='INSERT')
            
        @event.listens_for(User, 'after_update')
        def user_updated(mapper, connection, target):
            logger.info(f"User updated: {target.username} ({target.role.value})")
            self.sync_user_to_excel_tables(target, action='UPDATE')
            
        @event.listens_for(User, 'after_delete')
        def user_deleted(mapper, connection, target):
            logger.info(f"User deleted: {target.username}")
            self.sync_user_to_excel_tables(target, action='DELETE')
            
        @event.listens_for(Vendor, 'after_insert')
        def vendor_inserted(mapper, connection, target):
            logger.info(f"New vendor profile created: {target.vendor_id}")
            self.sync_vendor_to_excel_tables(target, action='INSERT')
            
        @event.listens_for(Vendor, 'after_update')
        def vendor_updated(mapper, connection, target):
            logger.info(f"Vendor profile updated: {target.vendor_id}")
            self.sync_vendor_to_excel_tables(target, action='UPDATE')
            
        @event.listens_for(Manager, 'after_insert')
        def manager_inserted(mapper, connection, target):
            logger.info(f"New manager profile created: {target.manager_id}")
            self.sync_manager_to_excel_tables(target, action='INSERT')
            
        @event.listens_for(Manager, 'after_update')
        def manager_updated(mapper, connection, target):
            logger.info(f"Manager profile updated: {target.manager_id}")
            self.sync_manager_to_excel_tables(target, action='UPDATE')
            
    def get_database_users(self) -> List[Dict]:
        """Get all users from database with their profiles"""
        try:
            with self.app.app_context():
                users_data = []
                users = db.session.query(User).all()
                
                for user in users:
                    user_dict = {
                        'user_id': user.id,
                        'username': user.username,
                        'email': user.email,
                        'role': user.role.value,
                        'is_active': user.is_active,
                        'created_at': user.created_at,
                        'last_login': user.last_login
                    }
                    
                    # Add profile-specific information
                    if user.role == UserRole.VENDOR and user.vendor_profile:
                        vendor = user.vendor_profile
                        user_dict.update({
                            'vendor_id': vendor.vendor_id,
                            'full_name': vendor.full_name,
                            'department': vendor.department,
                            'company': vendor.company,
                            'band': vendor.band,
                            'location': vendor.location,
                            'manager_id': vendor.manager_id,
                            'phone_number': None  # Not in vendor model
                        })
                    elif user.role == UserRole.MANAGER and user.manager_profile:
                        manager = user.manager_profile
                        user_dict.update({
                            'manager_id': manager.manager_id,
                            'full_name': manager.full_name,
                            'department': manager.department,
                            'team_name': manager.team_name,
                            'phone_number': manager.phone
                        })
                    elif user.role == UserRole.ADMIN:
                        user_dict.update({
                            'full_name': user.username,
                            'department': 'IT Administration',
                            'phone_number': None
                        })
                    
                    users_data.append(user_dict)
                    
                logger.info(f"Retrieved {len(users_data)} users from database")
                return users_data
                
        except Exception as e:
            logger.error(f"Error getting database users: {str(e)}")
            return []
    
    def sync_all_excel_tables(self):
        """Full synchronization of all Excel tables with current database state"""
        logger.info("Starting full synchronization of all Excel tables...")
        
        # Get current database state
        users_data = self.get_database_users()
        if not users_data:
            logger.warning("No users found in database")
            return
        
        # Update each Excel file
        for filename, config in self.excel_configs.items():
            try:
                file_path = self.excel_folder_path / filename
                self.update_excel_table(file_path, config, users_data)
                logger.info(f"Updated {filename}")
            except Exception as e:
                logger.error(f"Error updating {filename}: {str(e)}")
                
        logger.info("Full synchronization completed")
        
    def update_excel_table(self, file_path: Path, config: Dict, users_data: List[Dict]):
        """Update a specific Excel table with current user data"""
        target_roles = config["roles"]
        table_name = config["table_name"]
        notification_type = config["notification_type"]
        
        # Filter users for this notification type
        filtered_users = [user for user in users_data if user['role'] in target_roles and user['is_active']]
        
        if not filtered_users:
            logger.warning(f"No active users found for {file_path.name}")
            return
            
        # Generate Excel table data
        table_data = []
        primary_key_counter = 1
        
        for user in filtered_users:
            row = self.generate_excel_row(user, notification_type, primary_key_counter)
            table_data.append(row)
            primary_key_counter += 1
            
        # Create DataFrame
        df = pd.DataFrame(table_data)
        
        # Write to Excel with table formatting
        self.write_excel_with_table(file_path, df, table_name)
        
    def generate_excel_row(self, user: Dict, notification_type: str, primary_key: int) -> Dict:
        """Generate a row of data for Excel table based on user and notification type"""
        base_row = {
            'Primary_Key': f"NK_{notification_type.upper()}_{primary_key:04d}",
            'Contact_Email': user['email'],
            'Contact_Name': user.get('full_name', user['username']),
            'Role': user['role'],
            'Send_Notification': 'YES',
            'Active': 'YES' if user['is_active'] else 'NO',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'MEDIUM',
            'Created_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Add notification-specific fields
        if notification_type == 'daily_reminder':
            base_row.update({
                'Start_Time': '09:00',
                'End_Time': '18:00',
                'Reminder_Interval_Hours': '3',
                'Max_Reminders_Per_Day': '3',
                'Weekdays_Only': 'YES',
                'Exclude_Holidays': 'YES',
                'Custom_Message': f'Hi {user.get("full_name", user["username"])}, please submit your daily attendance status.',
                'Teams_Channel': user.get('department', 'General'),
                'Phone_Number': user.get('phone_number', ''),
                'Include_Manager_Name': 'YES',
                'Include_Submission_Link': 'YES',
                'Auto_Escalate_Days': '2'
            })
        elif notification_type == 'manager_summary':
            base_row.update({
                'Notification_Times': '12:00,14:00',
                'Weekdays_Only': 'YES',
                'Exclude_Holidays': 'YES',
                'Include_Team_Stats': 'YES',
                'Include_Pending_Count': 'YES',
                'Include_Mismatch_Alerts': 'YES',
                'Teams_Channel': user.get('team_name', user.get('department', 'Management')),
                'Custom_Message': f'Daily team status summary for {user.get("full_name", user["username"])}',
                'Phone_Number': user.get('phone_number', ''),
                'Summary_Format': 'DETAILED',
                'Include_Charts': 'YES'
            })
        elif notification_type == 'mismatch_alert':
            base_row.update({
                'Mismatch_Types': 'ATTENDANCE,SWIPE,LEAVE',
                'Alert_Immediately': 'YES',
                'Include_Resolution_Steps': 'YES',
                'Teams_Channel': user.get('department', 'General'),
                'Custom_Message': 'Attendance mismatch detected - please review and resolve.',
                'Phone_Number': user.get('phone_number', ''),
                'Auto_Escalate_Days': '1',
                'Require_Acknowledgment': 'YES'
            })
        elif notification_type == 'monthly_report':
            base_row.update({
                'Report_Day': '1',
                'Include_Charts': 'YES',
                'Include_Trends': 'YES',
                'Report_Format': 'PDF',
                'Teams_Channel': user.get('department', 'Reports'),
                'Custom_Message': 'Your monthly attendance report is ready.',
                'Phone_Number': user.get('phone_number', ''),
                'Include_Comparisons': 'YES',
                'Auto_Email_Report': 'YES'
            })
        elif notification_type == 'system_alert':
            base_row.update({
                'Alert_Types': 'ERROR,WARNING,INFO',
                'Alert_Immediately': 'YES',
                'Include_System_Status': 'YES',
                'Teams_Channel': 'IT-Alerts',
                'Custom_Message': 'System alert requires your attention.',
                'Phone_Number': user.get('phone_number', ''),
                'Escalation_Levels': '3',
                'Include_Logs': 'YES'
            })
        elif notification_type == 'holiday_reminder':
            base_row.update({
                'Days_Before_Holiday': '3,1',
                'Include_Holiday_Details': 'YES',
                'Teams_Channel': 'General',
                'Custom_Message': 'Upcoming holiday reminder',
                'Phone_Number': user.get('phone_number', ''),
                'Include_Policy_Links': 'YES'
            })
        elif notification_type == 'late_submission':
            base_row.update({
                'Grace_Period_Hours': '2',
                'Alert_Frequency': 'HOURLY',
                'Include_Vendor_List': 'YES',
                'Teams_Channel': user.get('department', 'Management'),
                'Custom_Message': 'Late submission alert',
                'Phone_Number': user.get('phone_number', ''),
                'Auto_Follow_Up': 'YES'
            })
        elif notification_type == 'billing_correction':
            base_row.update({
                'Correction_Types': 'HOURS,OVERTIME,LEAVE',
                'Include_Original_Values': 'YES',
                'Require_Approval': 'YES',
                'Teams_Channel': 'Finance',
                'Custom_Message': 'Billing correction notification',
                'Phone_Number': user.get('phone_number', ''),
                'Include_Audit_Trail': 'YES'
            })
        elif notification_type == 'feedback':
            base_row.update({
                'Feedback_Types': 'APPROVAL,REJECTION,CORRECTION',
                'Include_Manager_Comments': 'YES',
                'Include_Next_Steps': 'YES',
                'Teams_Channel': user.get('department', 'General'),
                'Custom_Message': 'Manager feedback on your submission',
                'Phone_Number': user.get('phone_number', ''),
                'Require_Acknowledgment': 'YES'
            })
        elif notification_type == 'all_complete':
            base_row.update({
                'Team_Completion_Trigger': 'YES',
                'Include_Summary_Stats': 'YES',
                'Celebration_Message': 'YES',
                'Teams_Channel': user.get('team_name', 'Management'),
                'Custom_Message': 'All team members have submitted their status!',
                'Phone_Number': user.get('phone_number', ''),
                'Include_Recognition': 'YES'
            })
            
        return base_row
    
    def write_excel_with_table(self, file_path: Path, df: pd.DataFrame, table_name: str):
        """Write DataFrame to Excel with proper table formatting"""
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = table_name
            
            # Write data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format as Excel table
            if len(df) > 0:
                table_range = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
                table = openpyxl.worksheet.table.Table(displayName=table_name, ref=table_range)
                
                # Add table style
                style = openpyxl.worksheet.table.TableStyleInfo(
                    name="TableStyleMedium9", 
                    showFirstColumn=False,
                    showLastColumn=False, 
                    showRowStripes=True, 
                    showColumnStripes=True
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            
            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            wb.save(str(file_path))
            logger.info(f"Successfully updated {file_path} with {len(df)} records")
            
        except Exception as e:
            logger.error(f"Error writing Excel file {file_path}: {str(e)}")
            raise
    
    def sync_user_to_excel_tables(self, user, action='UPDATE'):
        """Sync a specific user change to relevant Excel tables"""
        if not self.app:
            return
            
        try:
            with self.app.app_context():
                # Get updated user data
                users_data = self.get_database_users()
                
                # Update relevant Excel files based on user role
                for filename, config in self.excel_configs.items():
                    if user.role.value in config["roles"]:
                        file_path = self.excel_folder_path / filename
                        self.update_excel_table(file_path, config, users_data)
                        
                logger.info(f"Synced user {user.username} to Excel tables")
                
        except Exception as e:
            logger.error(f"Error syncing user {user.username}: {str(e)}")
    
    def sync_vendor_to_excel_tables(self, vendor, action='UPDATE'):
        """Sync vendor profile changes to Excel tables"""
        if not self.app:
            return
            
        try:
            with self.app.app_context():
                # Get updated user data
                users_data = self.get_database_users()
                
                # Update vendor-related Excel files
                vendor_files = [filename for filename, config in self.excel_configs.items() 
                              if 'VENDOR' in config["roles"]]
                
                for filename in vendor_files:
                    file_path = self.excel_folder_path / filename
                    config = self.excel_configs[filename]
                    self.update_excel_table(file_path, config, users_data)
                    
                logger.info(f"Synced vendor {vendor.vendor_id} to Excel tables")
                
        except Exception as e:
            logger.error(f"Error syncing vendor {vendor.vendor_id}: {str(e)}")
    
    def sync_manager_to_excel_tables(self, manager, action='UPDATE'):
        """Sync manager profile changes to Excel tables"""
        if not self.app:
            return
            
        try:
            with self.app.app_context():
                # Get updated user data
                users_data = self.get_database_users()
                
                # Update manager-related Excel files
                manager_files = [filename for filename, config in self.excel_configs.items() 
                               if 'MANAGER' in config["roles"]]
                
                for filename in manager_files:
                    file_path = self.excel_folder_path / filename
                    config = self.excel_configs[filename]
                    self.update_excel_table(file_path, config, users_data)
                    
                logger.info(f"Synced manager {manager.manager_id} to Excel tables")
                
        except Exception as e:
            logger.error(f"Error syncing manager {manager.manager_id}: {str(e)}")
    
    def validate_sync_accuracy(self) -> Dict[str, Any]:
        """Validate that Excel tables are in sync with database"""
        validation_results = {
            'total_files': len(self.excel_configs),
            'files_checked': 0,
            'sync_issues': [],
            'last_validation': datetime.now().isoformat(),
            'database_user_count': 0,
            'excel_record_counts': {}
        }
        
        try:
            with self.app.app_context():
                # Get database user count
                db_users = db.session.query(User).filter_by(is_active=True).count()
                validation_results['database_user_count'] = db_users
                
                # Check each Excel file
                for filename, config in self.excel_configs.items():
                    file_path = self.excel_folder_path / filename
                    
                    try:
                        if file_path.exists():
                            df = pd.read_excel(file_path, sheet_name=config["table_name"])
                            excel_count = len(df[df['Active'] == 'YES'])
                            validation_results['excel_record_counts'][filename] = excel_count
                            
                            # Check if counts match expected for roles
                            expected_count = db.session.query(User).filter(
                                User.role.in_([UserRole(role) for role in config["roles"]]),
                                User.is_active == True
                            ).count()
                            
                            if excel_count != expected_count:
                                validation_results['sync_issues'].append({
                                    'file': filename,
                                    'issue': 'count_mismatch',
                                    'expected': expected_count,
                                    'actual': excel_count,
                                    'roles': config["roles"]
                                })
                            
                            validation_results['files_checked'] += 1
                        else:
                            validation_results['sync_issues'].append({
                                'file': filename,
                                'issue': 'file_missing'
                            })
                    except Exception as e:
                        validation_results['sync_issues'].append({
                            'file': filename,
                            'issue': 'read_error',
                            'error': str(e)
                        })
                        
                logger.info(f"Validation completed: {len(validation_results['sync_issues'])} issues found")
                return validation_results
                
        except Exception as e:
            logger.error(f"Validation error: {str(e)}")
            validation_results['sync_issues'].append({
                'file': 'validation',
                'issue': 'validation_error',
                'error': str(e)
            })
            return validation_results

def setup_notification_database_sync(app):
    """Setup function to initialize database sync with Flask app"""
    sync_manager = NotificationDatabaseSync(app)
    
    # Perform initial full sync
    sync_manager.sync_all_excel_tables()
    
    return sync_manager

# CLI interface for manual operations
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Notification Database Sync Manager')
    parser.add_argument('--action', choices=['sync-all', 'validate', 'test-connection'], 
                       default='sync-all', help='Action to perform')
    parser.add_argument('--excel-folder', default='notification_configs', 
                       help='Excel files folder path')
    
    args = parser.parse_args()
    
    # Create sync manager (without app context for CLI)
    sync_manager = NotificationDatabaseSync(excel_folder_path=args.excel_folder)
    
    if args.action == 'sync-all':
        print("This action requires Flask app context. Use setup_notification_database_sync(app) instead.")
    elif args.action == 'validate':
        print("This action requires Flask app context. Use setup_notification_database_sync(app) instead.")
    elif args.action == 'test-connection':
        print(f"Excel folder path: {sync_manager.excel_folder_path}")
        print(f"Excel configs loaded: {len(sync_manager.excel_configs)}")
        print("Configuration test passed!")
