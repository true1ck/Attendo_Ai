#!/usr/bin/env python3
"""
Unified Notification Processor
Reads from all 9 notification Excel files in network drive and creates a unified queue
for Power Automate to process.
"""

import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from datetime import datetime, timedelta
import logging
import json
import sqlite3

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class UnifiedNotificationProcessor:
    """
    Processes all notification types and creates a unified queue for Power Automate.
    """
    
    def __init__(self, network_folder='G:/Test', output_folder='network_folder_simplified'):
        self.network_folder = Path(network_folder)
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        
        self.queue_file = self.output_folder / 'sent_noti_now.xlsx'
        self.tracking_file = self.output_folder / 'notification_tracking.json'
        
        # Required columns for output
        self.output_columns = [
            'EmployeeID', 
            'ContactEmail', 
            'Message', 
            'NotificationType', 
            'Priority', 
            'NTStatusSent'
        ]
        
        # Notification timing rules (in hours)
        self.notification_intervals = {
            'Daily Reminder': 3,  # Every 3 hours
            'Manager Summary': 24,  # Once per day
            'Mismatch Alert': 1,  # Every hour for urgent
            'Late Submission': 2,  # Every 2 hours
            'Holiday Reminder': 48,  # Every 2 days
            'System Alert': 0.5,  # Every 30 minutes for critical
        }
        
        # Max notifications per day
        self.max_daily_notifications = {
            'Daily Reminder': 3,  # Max 3 times per day (every 3 hours)
            'Manager Summary': 1,
            'Mismatch Alert': 6,
            'Late Submission': 4,
            'Holiday Reminder': 1,
            'System Alert': 10,
        }
        
        # Load tracking data
        self.tracking_data = self.load_tracking_data()
    
    def load_tracking_data(self):
        """Load notification tracking data."""
        if self.tracking_file.exists():
            try:
                with open(self.tracking_file, 'r') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def save_tracking_data(self):
        """Save notification tracking data."""
        with open(self.tracking_file, 'w') as f:
            json.dump(self.tracking_data, f, indent=2, default=str)
    
    def should_send_notification(self, employee_id, notification_type):
        """
        Check if a notification should be sent based on timing rules.
        
        Returns:
            bool: True if notification should be sent
        """
        key = f"{employee_id}_{notification_type}"
        current_time = datetime.now()
        
        if key not in self.tracking_data:
            # First time sending
            self.tracking_data[key] = {
                'last_sent': current_time.isoformat(),
                'count_today': 1,
                'date': current_time.date().isoformat()
            }
            return True
        
        tracking = self.tracking_data[key]
        last_sent = datetime.fromisoformat(tracking['last_sent'])
        tracking_date = datetime.fromisoformat(tracking['date']).date()
        
        # Reset daily count if it's a new day
        if current_time.date() > tracking_date:
            tracking['count_today'] = 0
            tracking['date'] = current_time.date().isoformat()
        
        # Check interval
        interval_hours = self.notification_intervals.get(notification_type, 1)
        time_since_last = (current_time - last_sent).total_seconds() / 3600
        
        # Check daily limit
        max_daily = self.max_daily_notifications.get(notification_type, 10)
        
        if time_since_last >= interval_hours and tracking['count_today'] < max_daily:
            tracking['last_sent'] = current_time.isoformat()
            tracking['count_today'] += 1
            return True
        
        return False
    
    def read_notification_files(self):
        """Read all 9 notification Excel files from network folder."""
        notifications = []
        
        # Map of files to their notification types
        file_mappings = {
            '01_daily_status_reminders.xlsx': 'Daily Reminder',
            '02_manager_summary_notifications.xlsx': 'Manager Summary',
            '03_manager_all_complete_notifications.xlsx': 'Manager Summary',
            '04_mismatch_notifications.xlsx': 'Mismatch Alert',
            '05_manager_feedback_notifications.xlsx': 'Manager Summary',
            '06_monthly_report_notifications.xlsx': 'Manager Summary',
            '07_admin_system_alerts.xlsx': 'System Alert',
            '08_holiday_reminder_notifications.xlsx': 'Holiday Reminder',
            '09_late_submission_alerts.xlsx': 'Late Submission'
        }
        
        for file_name, default_type in file_mappings.items():
            file_path = self.network_folder / file_name
            
            if not file_path.exists():
                logger.warning(f"File not found: {file_path}")
                continue
            
            try:
                # Read Excel file
                df = pd.read_excel(file_path)
                
                # Ensure required columns exist
                if not all(col in df.columns for col in ['EmployeeID', 'ContactEmail', 'Message']):
                    logger.warning(f"Missing required columns in {file_name}")
                    continue
                
                # Add notification type if not present
                if 'NotificationType' not in df.columns:
                    df['NotificationType'] = default_type
                
                # Add priority if not present
                if 'Priority' not in df.columns:
                    df['Priority'] = 'Medium'
                
                # Filter based on timing rules
                for _, row in df.iterrows():
                    if pd.notna(row['EmployeeID']) and pd.notna(row['ContactEmail']):
                        notification_type = row.get('NotificationType', default_type)
                        
                        if self.should_send_notification(row['EmployeeID'], notification_type):
                            # Generate appropriate message based on type
                            message = self.generate_message(notification_type, row)
                            priority = self.get_priority(notification_type)
                            
                            notifications.append({
                                'EmployeeID': row['EmployeeID'],
                                'ContactEmail': row['ContactEmail'],
                                'Message': message,
                                'NotificationType': notification_type,
                                'Priority': priority,
                                'NTStatusSent': False
                            })
                
                logger.info(f"✅ Processed {file_name}: {len(df)} rows")
                
            except Exception as e:
                logger.error(f"❌ Error reading {file_name}: {e}")
        
        return notifications
    
    def generate_message(self, notification_type, row):
        """Generate appropriate message based on notification type."""
        employee_id = row.get('EmployeeID', 'Employee')
        
        messages = {
            'Daily Reminder': [
                f"Hi {employee_id}, Please submit your daily timesheet before end of day.",
                f"Reminder: Your timesheet for today is pending submission.",
                f"Action Required: Please log your work hours for today.",
                f"Daily Timesheet Alert: Don't forget to update your hours."
            ],
            'Manager Summary': [
                f"Team Timesheet Summary: Review pending submissions from your team.",
                f"Manager Alert: {employee_id} - Weekly timesheet report ready for review.",
                f"Action Required: Please review and approve team timesheets.",
                f"Weekly Summary: Vendor timesheet status report available."
            ],
            'Mismatch Alert': [
                f"URGENT: Timesheet hours mismatch detected for {employee_id}.",
                f"Discrepancy Alert: Your logged hours don't match system records.",
                f"Action Required: Please review and correct timesheet entries.",
                f"Mismatch Found: Immediate attention needed for timesheet correction."
            ],
            'Late Submission': [
                f"OVERDUE: Timesheet submission deadline has passed for {employee_id}.",
                f"Urgent Reminder: Your timesheet is {employee_id} days overdue.",
                f"Final Notice: Please submit pending timesheet immediately.",
                f"Escalation Alert: Multiple timesheet submissions pending."
            ],
            'System Alert': [
                f"System Notification: Timesheet system maintenance scheduled.",
                f"Important: System update affecting timesheet submissions.",
                f"Alert: New timesheet policy effective immediately.",
                f"System Message: Please update your timesheet profile."
            ],
            'Holiday Reminder': [
                f"Holiday Notice: Office closed on upcoming holiday - submit timesheet early.",
                f"Reminder: Adjusted timesheet deadline due to holiday.",
                f"Holiday Alert: Please submit timesheet before the long weekend.",
                f"Advance Notice: Holiday schedule affecting timesheet submissions."
            ]
        }
        
        # Get message list for type, use default if not found
        msg_list = messages.get(notification_type, [f"{notification_type}: Action required for {employee_id}"])
        
        # Use row message if available and not generic
        if 'Message' in row and pd.notna(row['Message']) and 'notification' not in str(row['Message']).lower():
            return row['Message']
        
        # Otherwise, rotate through message templates
        import random
        return random.choice(msg_list)
    
    def get_priority(self, notification_type):
        """Get priority based on notification type."""
        priority_map = {
            'Daily Reminder': 'Medium',
            'Manager Summary': 'Medium',
            'Mismatch Alert': 'High',
            'Late Submission': 'High',
            'System Alert': 'High',
            'Holiday Reminder': 'Low'
        }
        return priority_map.get(notification_type, 'Medium')
    
    def create_queue_file(self, notifications):
        """Create the output Excel file with notification queue."""
        # Create DataFrame
        df = pd.DataFrame(notifications)
        
        # Ensure all required columns
        for col in self.output_columns:
            if col not in df.columns:
                df[col] = False if col == 'NTStatusSent' else ''
        
        # Reorder columns
        df = df[self.output_columns]
        
        # Create Excel file with table
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NotificationQueue"
        
        # Add headers
        for col_idx, header in enumerate(self.output_columns, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add data
        for row_idx, row in df.iterrows():
            for col_idx, col_name in enumerate(self.output_columns, 1):
                value = row[col_name]
                if pd.isna(value):
                    value = ''
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # Skip table creation to avoid corruption - just format as data
        # Power Automate can still read the data without a formal table
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        wb.save(self.queue_file)
        wb.close()
        
        logger.info(f"✅ Created queue file with {len(df)} notifications: {self.queue_file}")
        return len(df)
    
    def remove_sent_notifications(self):
        """Remove notifications marked as sent from the queue."""
        if not self.queue_file.exists():
            return 0
        
        try:
            # Read current queue
            df = pd.read_excel(self.queue_file)
            
            if 'NTStatusSent' not in df.columns:
                return 0
            
            # Count sent notifications
            sent_count = df[df['NTStatusSent'] == True].shape[0]
            
            # Keep only unsent notifications
            df_unsent = df[df['NTStatusSent'] != True]
            
            # Recreate the file with unsent notifications
            if len(df_unsent) != len(df):
                self.create_queue_file(df_unsent.to_dict('records'))
                logger.info(f"🗑️ Removed {sent_count} sent notifications from queue")
            
            return sent_count
            
        except Exception as e:
            logger.error(f"❌ Error removing sent notifications: {e}")
            return 0
    
    def process_notifications(self):
        """Main process to create notification queue."""
        logger.info("🔄 Starting unified notification processing...")
        
        # Remove sent notifications first
        removed = self.remove_sent_notifications()
        if removed > 0:
            logger.info(f"Cleaned {removed} sent notifications")
        
        # Read all notification files
        notifications = self.read_notification_files()
        
        # Create queue file
        count = self.create_queue_file(notifications)
        
        # Save tracking data
        self.save_tracking_data()
        
        logger.info(f"✅ Processing complete: {count} notifications queued")
        return count

def main():
    """Main function to run the processor."""
    processor = UnifiedNotificationProcessor()
    
    # Process notifications
    count = processor.process_notifications()
    
    print(f"\n📊 Notification Queue Summary:")
    print(f"   Total queued: {count}")
    print(f"   Output file: {processor.queue_file}")
    print(f"   Status: Ready for Power Automate")
    
    return count

if __name__ == "__main__":
    main()
