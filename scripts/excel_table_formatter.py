#!/usr/bin/env python3
"""
Excel Table Formatter Module
Ensures all Excel sheet updates are saved in proper table format for Power Automate compatibility.
This module provides functions to create, update, and maintain Excel files with proper table structures.
"""

import pandas as pd
import os
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelTableFormatter:
    """
    Manages Excel files with proper table formatting for Power Automate integration.
    Ensures all data is structured as named tables that can be easily referenced.
    """
    
    def __init__(self, base_folder="notification_configs"):
        self.base_folder = Path(base_folder)
        self.base_folder.mkdir(exist_ok=True)
        
        # Default table configurations for different notification types
        self.table_configs = {
            'daily_reminders': {
                'table_name': 'DailyStatusReminders',
                'sheet_name': 'Daily_Reminders',
                'style': 'TableStyleMedium9'
            },
            'manager_summary': {
                'table_name': 'ManagerSummaryNotifications', 
                'sheet_name': 'Manager_Summary',
                'style': 'TableStyleMedium2'
            },
            'manager_complete': {
                'table_name': 'AllCompleteNotifications',
                'sheet_name': 'All_Complete',
                'style': 'TableStyleMedium3'
            },
            'mismatch_alerts': {
                'table_name': 'MismatchNotifications',
                'sheet_name': 'Mismatch_Alerts', 
                'style': 'TableStyleMedium4'
            },
            'manager_feedback': {
                'table_name': 'ManagerFeedbackNotifications',
                'sheet_name': 'Manager_Feedback',
                'style': 'TableStyleMedium5'
            },
            'monthly_reports': {
                'table_name': 'MonthlyReportNotifications',
                'sheet_name': 'Monthly_Reports',
                'style': 'TableStyleMedium6'
            },
            'admin_alerts': {
                'table_name': 'AdminSystemAlerts',
                'sheet_name': 'System_Alerts',
                'style': 'TableStyleMedium7'
            },
            'holiday_reminders': {
                'table_name': 'HolidayReminderNotifications',
                'sheet_name': 'Holiday_Reminders',
                'style': 'TableStyleMedium8'
            },
            'late_submissions': {
                'table_name': 'LateSubmissionAlerts',
                'sheet_name': 'Late_Submissions',
                'style': 'TableStyleMedium10'
            },
            'billing_corrections': {
                'table_name': 'BillingCorrectionNotifications',
                'sheet_name': 'Billing_Corrections',
                'style': 'TableStyleMedium11'
            }
        }
    
    def create_excel_table_from_dataframe(self, df: pd.DataFrame, filepath: str, 
                                        config_type: str = None, custom_config: dict = None):
        """
        Create an Excel file with proper table formatting from a DataFrame.
        
        Args:
            df: pandas DataFrame with the data
            filepath: Full path to the Excel file to create
            config_type: Type of notification config (daily_reminders, manager_summary, etc.)
            custom_config: Custom table configuration if config_type is not found
        """
        try:
            # Get table configuration
            if config_type and config_type in self.table_configs:
                config = self.table_configs[config_type]
            elif custom_config:
                config = custom_config
            else:
                config = {
                    'table_name': 'DataTable',
                    'sheet_name': 'Data',
                    'style': 'TableStyleMedium9'
                }
            
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = config['sheet_name']
            
            # Add DataFrame data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Create table if we have data
            if len(df) > 0:
                table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
                
                table = Table(displayName=config['table_name'], ref=table_range)
                
                # Apply table style
                style = TableStyleInfo(
                    name=config['style'], 
                    showFirstColumn=False,
                    showLastColumn=False, 
                    showRowStripes=True, 
                    showColumnStripes=True
                )
                table.tableStyleInfo = style
                
                # Add table to worksheet
                ws.add_table(table)
                
                logger.info(f"✅ Created table '{config['table_name']}' with {len(df)} rows")
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws)
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"✅ Created Excel file with table format: {filepath}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error creating Excel table: {str(e)}")
            return False
    
    def update_excel_table_from_dataframe(self, df: pd.DataFrame, filepath: str, 
                                        config_type: str = None, backup: bool = True):
        """
        Update an existing Excel file with table formatting, preserving the table structure.
        
        Args:
            df: pandas DataFrame with the updated data
            filepath: Path to the existing Excel file
            config_type: Type of notification config
            backup: Whether to create a backup of the existing file
        """
        try:
            filepath = Path(filepath)
            
            # Create backup if requested
            if backup and filepath.exists():
                backup_path = filepath.parent / f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filepath.name}"
                import shutil
                shutil.copy2(filepath, backup_path)
                logger.info(f"📋 Created backup: {backup_path}")
            
            # If file doesn't exist, create new one
            if not filepath.exists():
                return self.create_excel_table_from_dataframe(df, str(filepath), config_type)
            
            # Load existing workbook
            wb = load_workbook(str(filepath))
            
            # Get table configuration
            if config_type and config_type in self.table_configs:
                config = self.table_configs[config_type]
            else:
                config = {
                    'table_name': 'DataTable',
                    'sheet_name': 'Data',
                    'style': 'TableStyleMedium9'
                }
            
            # Find the worksheet (try configured name first, then active sheet)
            if config['sheet_name'] in wb.sheetnames:
                ws = wb[config['sheet_name']]
            else:
                ws = wb.active
                ws.title = config['sheet_name']
            
            # Remove existing table if it exists
            existing_tables = list(ws.tables.keys())
            for table_name in existing_tables:
                del ws.tables[table_name]
            
            # Clear existing data
            ws.delete_rows(1, ws.max_row)
            
            # Add new DataFrame data
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Create new table if we have data
            if len(df) > 0:
                table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
                
                table = Table(displayName=config['table_name'], ref=table_range)
                
                # Apply table style
                style = TableStyleInfo(
                    name=config['style'], 
                    showFirstColumn=False,
                    showLastColumn=False, 
                    showRowStripes=True, 
                    showColumnStripes=True
                )
                table.tableStyleInfo = style
                
                # Add table to worksheet
                ws.add_table(table)
                
                logger.info(f"✅ Updated table '{config['table_name']}' with {len(df)} rows")
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws)
            
            # Save workbook
            wb.save(str(filepath))
            logger.info(f"✅ Updated Excel file with table format: {filepath}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error updating Excel table: {str(e)}")
            return False
    
    def pandas_to_excel_table(self, df: pd.DataFrame, filepath: str, 
                             table_name: str = None, sheet_name: str = None, 
                             table_style: str = 'TableStyleMedium9'):
        """
        Simple function to save pandas DataFrame as Excel table with custom parameters.
        
        Args:
            df: pandas DataFrame
            filepath: Output file path
            table_name: Name for the Excel table (default: auto-generated)
            sheet_name: Name for the worksheet (default: 'Data')
            table_style: Excel table style (default: TableStyleMedium9)
        """
        if table_name is None:
            table_name = f"Table_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if sheet_name is None:
            sheet_name = 'Data'
        
        custom_config = {
            'table_name': table_name,
            'sheet_name': sheet_name,
            'style': table_style
        }
        
        return self.create_excel_table_from_dataframe(df, filepath, custom_config=custom_config)
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths for better readability."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with some padding, but cap at reasonable maximum
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def validate_table_structure(self, filepath: str):
        """
        Validate that an Excel file has proper table structure.
        
        Returns:
            dict: Validation results with table info
        """
        try:
            filepath = Path(filepath)
            if not filepath.exists():
                return {'valid': False, 'error': 'File does not exist'}
            
            wb = load_workbook(str(filepath))
            ws = wb.active
            
            tables = list(ws.tables.keys())
            
            if not tables:
                return {'valid': False, 'error': 'No tables found in worksheet'}
            
            table_info = []
            for table_name in tables:
                table = ws.tables[table_name]
                table_info.append({
                    'name': table_name,
                    'range': table.ref,
                    'style': table.tableStyleInfo.name if table.tableStyleInfo else None
                })
            
            return {
                'valid': True,
                'tables': table_info,
                'worksheet': ws.title,
                'total_tables': len(tables)
            }
            
        except Exception as e:
            return {'valid': False, 'error': str(e)}
    
    def get_notification_file_path(self, notification_type: str):
        """Get the standard file path for a notification type."""
        file_mappings = {
            'daily_reminders': '01_daily_status_reminders.xlsx',
            'manager_summary': '02_manager_summary_notifications.xlsx',
            'manager_complete': '03_manager_all_complete_notifications.xlsx',
            'mismatch_alerts': '04_mismatch_notifications.xlsx',
            'manager_feedback': '05_manager_feedback_notifications.xlsx',
            'monthly_reports': '06_monthly_report_notifications.xlsx',
            'admin_alerts': '07_admin_system_alerts.xlsx',
            'holiday_reminders': '08_holiday_reminder_notifications.xlsx',
            'late_submissions': '09_late_submission_alerts.xlsx',
            'billing_corrections': '10_billing_correction_notifications.xlsx'
        }
        
        if notification_type in file_mappings:
            return self.base_folder / file_mappings[notification_type]
        else:
            return self.base_folder / f"{notification_type}.xlsx"

# Global instance for easy import
excel_table_formatter = ExcelTableFormatter()

# Convenience functions for easy use
def save_dataframe_as_table(df: pd.DataFrame, filepath: str, table_name: str = None, 
                           sheet_name: str = None, table_style: str = 'TableStyleMedium9'):
    """
    Save a pandas DataFrame as an Excel file with proper table formatting.
    
    Args:
        df: pandas DataFrame to save
        filepath: Output Excel file path
        table_name: Name for the Excel table
        sheet_name: Name for the worksheet
        table_style: Excel table style
    """
    return excel_table_formatter.pandas_to_excel_table(
        df, filepath, table_name, sheet_name, table_style
    )

def update_notification_table(df: pd.DataFrame, notification_type: str, backup: bool = True):
    """
    Update a notification Excel file with proper table formatting.
    
    Args:
        df: pandas DataFrame with updated data
        notification_type: Type of notification (daily_reminders, manager_summary, etc.)
        backup: Whether to create backup before updating
    """
    filepath = excel_table_formatter.get_notification_file_path(notification_type)
    return excel_table_formatter.update_excel_table_from_dataframe(
        df, str(filepath), notification_type, backup
    )

def create_notification_table(df: pd.DataFrame, notification_type: str):
    """
    Create a new notification Excel file with proper table formatting.
    
    Args:
        df: pandas DataFrame with data
        notification_type: Type of notification
    """
    filepath = excel_table_formatter.get_notification_file_path(notification_type)
    return excel_table_formatter.create_excel_table_from_dataframe(
        df, str(filepath), notification_type
    )

def validate_notification_table(notification_type: str):
    """
    Validate that a notification Excel file has proper table structure.
    
    Args:
        notification_type: Type of notification to validate
    """
    filepath = excel_table_formatter.get_notification_file_path(notification_type)
    return excel_table_formatter.validate_table_structure(str(filepath))

if __name__ == "__main__":
    # Test the module
    import sys
    
    print("🧪 Testing Excel Table Formatter...")
    
    # Create test data
    test_data = pd.DataFrame({
        'ID': [1, 2, 3],
        'Name': ['John Doe', 'Jane Smith', 'Bob Johnson'],
        'Email': ['john@test.com', 'jane@test.com', 'bob@test.com'],
        'Status': ['Active', 'Active', 'Pending'],
        'Date': [datetime.now().strftime('%Y-%m-%d')] * 3
    })
    
    # Test creating table
    test_file = "test_table_output.xlsx"
    success = save_dataframe_as_table(
        test_data, 
        test_file, 
        table_name="TestTable",
        sheet_name="Test_Data"
    )
    
    if success:
        print(f"✅ Successfully created test Excel file: {test_file}")
        
        # Validate the table structure
        validation = excel_table_formatter.validate_table_structure(test_file)
        if validation['valid']:
            print(f"✅ Table validation passed: {validation}")
        else:
            print(f"❌ Table validation failed: {validation}")
        
        # Clean up test file
        try:
            os.remove(test_file)
            print("🧹 Cleaned up test file")
        except:
            pass
    else:
        print("❌ Failed to create test Excel file")
        sys.exit(1)
    
    print("✅ Excel Table Formatter module is working correctly!")
