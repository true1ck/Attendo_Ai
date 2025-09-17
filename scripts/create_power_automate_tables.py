#!/usr/bin/env python3
"""
Script to create Excel files with proper table formatting for Microsoft Power Automate
Each Excel file will have data formatted as a table that Power Automate can easily reference
"""

import pandas as pd
import os
from datetime import datetime, time
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def create_power_automate_excel_files():
    """Create Excel files with proper table formatting for Power Automate integration"""
    
    # Ensure notification_configs directory exists
    config_dir = "notification_configs"
    os.makedirs(config_dir, exist_ok=True)
    
    # 1. Daily Status Reminders Table
    vendor_reminder_data = [
        {
            'Primary_Key': 'VENDOR_V001',
            'Contact_Email': 'vendor1@company.com',
            'Contact_Name': 'John Doe',
            'Vendor_ID': 'V001',
            'Department': 'MTB_WCS_MSE7_MS1',
            'Company': 'ABC Solutions',
            'Manager_ID': 'M001',
            'Send_Notification': 'YES',
            'Reminder_Interval_Hours': 3,
            'Start_Time': '09:00',
            'End_Time': '18:00',
            'Max_Reminders_Per_Day': 4,
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'YES',
            'Custom_Message': 'Please submit your daily attendance status',
            'Teams_Channel': 'vendor_notifications',
            'Phone_Number': '+1234567890',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'MEDIUM',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'VENDOR_V002',
            'Contact_Email': 'vendor2@company.com',
            'Contact_Name': 'Alice Smith',
            'Vendor_ID': 'V002',
            'Department': 'MTB_WCS_MSE7_MS2',
            'Company': 'XYZ Solutions',
            'Manager_ID': 'M001',
            'Send_Notification': 'YES',
            'Reminder_Interval_Hours': 3,
            'Start_Time': '09:00',
            'End_Time': '18:00',
            'Max_Reminders_Per_Day': 4,
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'YES',
            'Custom_Message': 'Daily attendance reminder',
            'Teams_Channel': 'vendor_notifications',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'EMAIL,SMS',
            'Priority': 'MEDIUM',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=vendor_reminder_data,
        filename=f"{config_dir}/01_daily_status_reminders.xlsx",
        table_name="DailyStatusReminders",
        worksheet_name="Daily Status Reminders"
    )
    
    # 2. Manager Summary Notifications Table
    manager_summary_data = [
        {
            'Primary_Key': 'MGR_SUMMARY_M001',
            'Contact_Email': 'manager1@company.com',
            'Contact_Name': 'Jane Smith',
            'Manager_ID': 'M001',
            'Department': 'Engineering',
            'Team_Name': 'Team Alpha',
            'Send_Notification': 'YES',
            'Notification_Times': '12:00,14:00',
            'Include_Team_Stats': 'YES',
            'Include_Individual_Status': 'YES',
            'Include_Pending_Count': 'YES',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'YES',
            'Custom_Message': 'Team attendance summary',
            'Teams_Channel': 'team_m001',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'MGR_SUMMARY_M002',
            'Contact_Email': 'manager2@company.com',
            'Contact_Name': 'Bob Johnson',
            'Manager_ID': 'M002',
            'Department': 'Operations',
            'Team_Name': 'Team Beta',
            'Send_Notification': 'YES',
            'Notification_Times': '12:00,14:00',
            'Include_Team_Stats': 'YES',
            'Include_Individual_Status': 'NO',
            'Include_Pending_Count': 'YES',
            'Weekdays_Only': 'YES',
            'Exclude_Holidays': 'YES',
            'Custom_Message': 'Operations team summary',
            'Teams_Channel': 'team_m002',
            'Phone_Number': '+1234567892',
            'Notification_Method': 'EMAIL',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=manager_summary_data,
        filename=f"{config_dir}/02_manager_summary_notifications.xlsx",
        table_name="ManagerSummaryNotifications",
        worksheet_name="Manager Summary"
    )
    
    # 3. Manager All-Complete Notifications Table
    manager_complete_data = [
        {
            'Primary_Key': 'MGR_COMPLETE_M001',
            'Contact_Email': 'manager1@company.com',
            'Contact_Name': 'Jane Smith',
            'Manager_ID': 'M001',
            'Department': 'Engineering',
            'Team_Name': 'Team Alpha',
            'Send_Notification': 'YES',
            'Trigger_Condition': 'ALL_TEAM_SUBMITTED',
            'Include_Completion_Stats': 'YES',
            'Include_Next_Actions': 'YES',
            'Auto_Approve_Option': 'NO',
            'Custom_Message': 'All team members have submitted their status',
            'Teams_Channel': 'team_m001',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=manager_complete_data,
        filename=f"{config_dir}/03_manager_all_complete_notifications.xlsx",
        table_name="AllCompleteNotifications",
        worksheet_name="All Complete"
    )
    
    # 4. Mismatch Notifications Table
    mismatch_notification_data = [
        {
            'Primary_Key': 'MISMATCH_MGR_M001',
            'Contact_Email': 'manager1@company.com',
            'Contact_Name': 'Jane Smith',
            'Role': 'MANAGER',
            'Manager_ID': 'M001',
            'Vendor_ID': '',
            'Department': 'Engineering',
            'Send_Notification': 'YES',
            'Mismatch_Types': 'SWIPE_WEB,WFH_SWIPE,LEAVE_SWIPE',
            'Include_Vendor_Details': 'YES',
            'Include_Explanation_Required': 'YES',
            'Auto_Escalate_Days': 3,
            'Custom_Message': 'New attendance mismatches found for your team',
            'Teams_Channel': 'team_m001',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'MISMATCH_VND_V001',
            'Contact_Email': 'vendor1@company.com',
            'Contact_Name': 'John Doe',
            'Role': 'VENDOR',
            'Manager_ID': '',
            'Vendor_ID': 'V001',
            'Department': 'MTB_WCS_MSE7_MS1',
            'Send_Notification': 'YES',
            'Mismatch_Types': 'ALL',
            'Include_Vendor_Details': 'YES',
            'Include_Explanation_Required': 'YES',
            'Auto_Escalate_Days': 2,
            'Custom_Message': 'Please provide explanation for attendance mismatch',
            'Teams_Channel': '',
            'Phone_Number': '+1234567890',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'MEDIUM',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=mismatch_notification_data,
        filename=f"{config_dir}/04_mismatch_notifications.xlsx",
        table_name="MismatchNotifications",
        worksheet_name="Mismatch Alerts"
    )
    
    # 5. Manager Feedback Notifications Table
    feedback_notification_data = [
        {
            'Primary_Key': 'FEEDBACK_V001',
            'Contact_Email': 'vendor1@company.com',
            'Contact_Name': 'John Doe',
            'Vendor_ID': 'V001',
            'Department': 'MTB_WCS_MSE7_MS1',
            'Manager_ID': 'M001',
            'Send_Notification': 'YES',
            'Feedback_Types': 'REJECTION,CORRECTION_REQUEST,APPROVAL',
            'Include_Manager_Comments': 'YES',
            'Include_Corrective_Actions': 'YES',
            'Include_Deadline': 'YES',
            'Response_Required_Hours': 24,
            'Custom_Message': 'Manager feedback on your attendance submission',
            'Teams_Channel': '',
            'Phone_Number': '+1234567890',
            'Notification_Method': 'TEAMS,EMAIL,SMS',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=feedback_notification_data,
        filename=f"{config_dir}/05_manager_feedback_notifications.xlsx",
        table_name="ManagerFeedbackNotifications",
        worksheet_name="Manager Feedback"
    )
    
    # 6. Monthly Report Notifications Table
    report_notification_data = [
        {
            'Primary_Key': 'REPORT_MGR_M001',
            'Contact_Email': 'manager1@company.com',
            'Contact_Name': 'Jane Smith',
            'Role': 'MANAGER',
            'Manager_ID': 'M001',
            'Department': 'Engineering',
            'Send_Notification': 'YES',
            'Report_Types': 'MONTHLY_TEAM,MONTHLY_INDIVIDUAL,MISMATCH_SUMMARY',
            'Auto_Generate_Monthly': 'YES',
            'Generation_Day': 1,
            'Include_Excel_Attachment': 'YES',
            'Include_PDF_Summary': 'YES',
            'Include_Charts': 'YES',
            'Custom_Message': 'Monthly attendance report for your team',
            'Teams_Channel': 'team_m001',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'MEDIUM',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'REPORT_ADMIN_1',
            'Contact_Email': 'admin@company.com',
            'Contact_Name': 'System Administrator',
            'Role': 'ADMIN',
            'Manager_ID': '',
            'Department': 'ADMINISTRATION',
            'Send_Notification': 'YES',
            'Report_Types': 'MONTHLY_ALL,SYSTEM_SUMMARY,AUDIT_REPORT,BILLING_REPORT',
            'Auto_Generate_Monthly': 'YES',
            'Generation_Day': 1,
            'Include_Excel_Attachment': 'YES',
            'Include_PDF_Summary': 'YES',
            'Include_Charts': 'YES',
            'Custom_Message': 'Monthly system-wide attendance report',
            'Teams_Channel': 'admin_channel',
            'Phone_Number': '+1234567892',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=report_notification_data,
        filename=f"{config_dir}/06_monthly_report_notifications.xlsx",
        table_name="MonthlyReportNotifications",
        worksheet_name="Monthly Reports"
    )
    
    # 7. Admin System Alerts Table
    admin_alert_data = [
        {
            'Primary_Key': 'ADMIN_ALERT_1',
            'Contact_Email': 'admin@company.com',
            'Contact_Name': 'System Administrator',
            'Role': 'ADMIN',
            'Send_Notification': 'YES',
            'Alert_Types': 'SYSTEM_ERROR,DATABASE_ISSUE,IMPORT_FAILURE,HIGH_MISMATCHES,SECURITY_ALERT',
            'Severity_Levels': 'HIGH,CRITICAL',
            'Include_System_Stats': 'YES',
            'Include_Error_Details': 'YES',
            'Include_Recommended_Actions': 'YES',
            'Auto_Escalate_Minutes': 30,
            'Custom_Message': 'System alert requiring immediate attention',
            'Teams_Channel': 'admin_alerts',
            'Phone_Number': '+1234567892',
            'Notification_Method': 'TEAMS,EMAIL,SMS',
            'Priority': 'CRITICAL',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=admin_alert_data,
        filename=f"{config_dir}/07_admin_system_alerts.xlsx",
        table_name="AdminSystemAlerts",
        worksheet_name="System Alerts"
    )
    
    # 8. Holiday Reminder Notifications Table
    holiday_reminder_data = [
        {
            'Primary_Key': 'HOLIDAY_ALL_USERS',
            'Contact_Email': 'all@company.com',
            'Contact_Name': 'All Users',
            'Role': 'ALL',
            'User_ID': 'ALL',
            'Department': 'ALL',
            'Send_Notification': 'YES',
            'Advance_Notice_Days': 7,
            'Include_Holiday_Calendar': 'YES',
            'Include_Working_Day_Info': 'YES',
            'Notification_Times': '09:00',
            'Weekend_Notifications': 'NO',
            'Custom_Message': 'Upcoming holiday notification',
            'Teams_Channel': 'general',
            'Phone_Number': '',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'LOW',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'HOLIDAY_MANAGERS_ONLY',
            'Contact_Email': 'managers@company.com',
            'Contact_Name': 'All Managers',
            'Role': 'MANAGER',
            'User_ID': 'MANAGERS',
            'Department': 'ALL',
            'Send_Notification': 'YES',
            'Advance_Notice_Days': 14,
            'Include_Holiday_Calendar': 'YES',
            'Include_Working_Day_Info': 'YES',
            'Notification_Times': '09:00',
            'Weekend_Notifications': 'NO',
            'Custom_Message': 'Holiday planning notification for managers',
            'Teams_Channel': 'managers',
            'Phone_Number': '',
            'Notification_Method': 'EMAIL',
            'Priority': 'MEDIUM',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=holiday_reminder_data,
        filename=f"{config_dir}/08_holiday_reminder_notifications.xlsx",
        table_name="HolidayReminderNotifications",
        worksheet_name="Holiday Reminders"
    )
    
    # 9. Late Submission Alert Notifications Table
    late_submission_data = [
        {
            'Primary_Key': 'LATE_MGR_M001',
            'Contact_Email': 'manager1@company.com',
            'Contact_Name': 'Jane Smith',
            'Role': 'MANAGER',
            'Manager_ID': 'M001',
            'Department': 'Engineering',
            'Send_Notification': 'YES',
            'Alert_After_Hours': 24,
            'Include_Late_History': 'YES',
            'Include_Team_Statistics': 'YES',
            'Include_Individual_Details': 'YES',
            'Escalate_After_Days': 3,
            'Custom_Message': 'Late submission alert for team members',
            'Teams_Channel': 'team_m001',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'MEDIUM',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'LATE_ADMIN_1',
            'Contact_Email': 'admin@company.com',
            'Contact_Name': 'System Administrator',
            'Role': 'ADMIN',
            'Manager_ID': '',
            'Department': 'ADMINISTRATION',
            'Send_Notification': 'YES',
            'Alert_After_Hours': 48,
            'Include_Late_History': 'YES',
            'Include_Team_Statistics': 'YES',
            'Include_Individual_Details': 'NO',
            'Escalate_After_Days': 7,
            'Custom_Message': 'System-wide late submission report',
            'Teams_Channel': 'admin_channel',
            'Phone_Number': '+1234567892',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'LOW',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=late_submission_data,
        filename=f"{config_dir}/09_late_submission_alerts.xlsx",
        table_name="LateSubmissionAlerts",
        worksheet_name="Late Submissions"
    )
    
    # 10. Billing Correction Notifications Table
    billing_correction_data = [
        {
            'Primary_Key': 'BILLING_MGR_M001',
            'Contact_Email': 'manager1@company.com',
            'Contact_Name': 'Jane Smith',
            'Role': 'MANAGER',
            'Manager_ID': 'M001',
            'Department': 'Engineering',
            'Send_Notification': 'YES',
            'Correction_Types': 'OFFSET_ADDITION,OFFSET_DEDUCTION,BILLING_ADJUSTMENT',
            'Include_Financial_Details': 'YES',
            'Include_Justification': 'YES',
            'Include_Approval_History': 'YES',
            'Require_Acknowledgment': 'YES',
            'Custom_Message': 'Billing correction notification for your team',
            'Teams_Channel': 'team_m001',
            'Phone_Number': '+1234567891',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'HIGH',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        {
            'Primary_Key': 'BILLING_ADMIN_1',
            'Contact_Email': 'admin@company.com',
            'Contact_Name': 'System Administrator',
            'Role': 'ADMIN',
            'Manager_ID': '',
            'Department': 'ADMINISTRATION',
            'Send_Notification': 'YES',
            'Correction_Types': 'ALL',
            'Include_Financial_Details': 'YES',
            'Include_Justification': 'YES',
            'Include_Approval_History': 'YES',
            'Require_Acknowledgment': 'YES',
            'Custom_Message': 'System billing correction notification',
            'Teams_Channel': 'admin_billing',
            'Phone_Number': '+1234567892',
            'Notification_Method': 'TEAMS,EMAIL',
            'Priority': 'CRITICAL',
            'Active': 'YES',
            'Created_Date': datetime.now().strftime('%Y-%m-%d'),
            'Last_Modified': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    ]
    
    create_excel_with_table(
        data=billing_correction_data,
        filename=f"{config_dir}/10_billing_correction_notifications.xlsx",
        table_name="BillingCorrectionNotifications",
        worksheet_name="Billing Corrections"
    )
    
    print("✅ Created 10 Excel files with Power Automate table formatting:")
    print("1. 01_daily_status_reminders.xlsx -> Table: DailyStatusReminders")
    print("2. 02_manager_summary_notifications.xlsx -> Table: ManagerSummaryNotifications")
    print("3. 03_manager_all_complete_notifications.xlsx -> Table: AllCompleteNotifications")
    print("4. 04_mismatch_notifications.xlsx -> Table: MismatchNotifications")
    print("5. 05_manager_feedback_notifications.xlsx -> Table: ManagerFeedbackNotifications")
    print("6. 06_monthly_report_notifications.xlsx -> Table: MonthlyReportNotifications")
    print("7. 07_admin_system_alerts.xlsx -> Table: AdminSystemAlerts")
    print("8. 08_holiday_reminder_notifications.xlsx -> Table: HolidayReminderNotifications")
    print("9. 09_late_submission_alerts.xlsx -> Table: LateSubmissionAlerts")
    print("10. 10_billing_correction_notifications.xlsx -> Table: BillingCorrectionNotifications")
    
    print("\n🔗 These tables are now ready for Microsoft Power Automate integration!")
    print("📋 Each Excel file contains a properly formatted table that can be referenced in Power Automate flows.")
    
    return True

def create_excel_with_table(data, filename, table_name, worksheet_name):
    """Create an Excel file with properly formatted table for Power Automate"""
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create workbook and add data
    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_name
    
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    # Create table
    table_range = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    
    table = Table(displayName=table_name, ref=table_range)
    
    # Add table style
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    
    # Add table to worksheet
    ws.add_table(table)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    print(f"✅ Created {filename} with table '{table_name}'")

if __name__ == "__main__":
    create_power_automate_excel_files()
