#!/usr/bin/env python3
"""
Populate Excel files with vendor data from the database
Uses the actual vendors table structure to create notification data
"""

import sqlite3
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_vendor_data():
    """Fetch vendor data from database."""
    conn = sqlite3.connect('vendor_timesheet.db')
    cursor = conn.cursor()
    
    # Get vendors with their information
    query = """
        SELECT 
            vendor_id,
            full_name,
            department,
            company,
            manager_id
        FROM vendors
        WHERE vendor_id IS NOT NULL
        LIMIT 20
    """
    
    cursor.execute(query)
    vendors = cursor.fetchall()
    
    # Also get users to find email addresses
    cursor.execute("""
        SELECT 
            v.vendor_id,
            u.email
        FROM vendors v
        LEFT JOIN users u ON v.user_id = u.id
        WHERE v.vendor_id IS NOT NULL
    """)
    
    vendor_emails = {row[0]: row[1] for row in cursor.fetchall() if row[1]}
    
    conn.close()
    
    return vendors, vendor_emails

def create_notification_excel(file_path: Path, notification_type: str, vendors_data: list, vendor_emails: dict):
    """
    Create Excel file with notification data.
    
    Args:
        file_path: Path to save the Excel file
        notification_type: Type of notification
        vendors_data: List of vendor tuples
        vendor_emails: Dictionary of vendor_id to email
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NotificationData"
    
    # Required columns for Power Automate
    headers = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 'Priority']
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Generate notification messages based on type
    messages = {
        'Daily Reminder': [
            'Please submit your daily timesheet',
            'Remember to log your work hours for today',
            'Daily timesheet submission deadline approaching',
            'Complete your timesheet entry before end of day'
        ],
        'Manager Summary': [
            'Team timesheet summary for review',
            'Weekly vendor timesheet report ready',
            'Please review team submissions'
        ],
        'Mismatch Alert': [
            'Timesheet hours mismatch detected',
            'Discrepancy found in logged hours',
            'Please review and correct timesheet entries'
        ],
        'Late Submission': [
            'Timesheet submission is overdue',
            'Urgent: Submit pending timesheet',
            'Final reminder for timesheet submission'
        ]
    }
    
    # Add vendor data
    row_idx = 2
    for vendor in vendors_data:
        vendor_id = vendor[0] if vendor[0] else f'VENDOR{row_idx-1:03d}'
        full_name = vendor[1] if vendor[1] else 'Vendor User'
        department = vendor[2] if vendor[2] else 'General'
        
        # Get email or generate one
        email = vendor_emails.get(vendor_id, f'{vendor_id.lower()}@company.com')
        
        # Select appropriate message
        msg_list = messages.get(notification_type, ['General notification'])
        message = msg_list[(row_idx - 2) % len(msg_list)]
        
        # Determine priority based on notification type
        if notification_type in ['Mismatch Alert', 'Late Submission']:
            priority = 'High'
        elif notification_type == 'Daily Reminder':
            priority = 'Medium' if (row_idx % 2) == 0 else 'High'
        else:
            priority = 'Medium'
        
        # Write row data
        ws.cell(row=row_idx, column=1, value=vendor_id)
        ws.cell(row=row_idx, column=2, value=email)
        ws.cell(row=row_idx, column=3, value=message)
        ws.cell(row=row_idx, column=4, value=notification_type)
        ws.cell(row=row_idx, column=5, value=priority)
        
        row_idx += 1
    
    # Create table if we have data
    if row_idx > 2:
        table_range = f"A1:E{row_idx - 1}"
        table_name = "NotificationTable"
        
        # Create table
        table = Table(displayName=table_name, ref=table_range)
        
        # Apply table style
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(file_path)
    wb.close()
    
    logger.info(f"✅ Created {file_path.name} with {row_idx - 2} rows")

def main(network_folder_path=None):
    """Main function to populate Excel files in network folder only.
    
    Args:
        network_folder_path: Optional path to network folder. If not provided, uses default.
    """
    
    # Get vendor data from database
    vendors_data, vendor_emails = get_vendor_data()
    
    if not vendors_data:
        logger.error("No vendor data found in database")
        # Create sample data
        vendors_data = [
            ('VENDOR001', 'Jane Vendor', 'IT', 'Company A', 'MGR001'),
            ('VENDOR002', 'Mike Vendor', 'IT', 'Company A', 'MGR001'),
            ('VENDOR003', 'Sarah Vendor', 'Finance', 'Company B', 'MGR002'),
            ('VENDOR004', 'David Vendor', 'Finance', 'Company B', 'MGR002'),
            ('TEST001', 'Test User', 'Testing', 'Company C', 'MGR003')
        ]
        vendor_emails = {
            'VENDOR001': 'jane.vendor@company.com',
            'VENDOR002': 'mike.vendor@company.com',
            'VENDOR003': 'sarah.vendor@company.com',
            'VENDOR004': 'david.vendor@company.com',
            'TEST001': 'test001@company.com'
        }
    
    logger.info(f"📊 Found {len(vendors_data)} vendors in database")
    
    # Define files to create
    files_to_create = [
        ('01_daily_status_reminders.xlsx', 'Daily Reminder'),
        ('02_manager_summary_notifications.xlsx', 'Manager Summary'),
        ('03_manager_all_complete_notifications.xlsx', 'Manager Summary'),
        ('04_mismatch_notifications.xlsx', 'Mismatch Alert'),
        ('05_manager_feedback_notifications.xlsx', 'Manager Summary'),
        ('06_monthly_report_notifications.xlsx', 'Manager Summary'),
        ('07_admin_system_alerts.xlsx', 'System Alert'),
        ('08_holiday_reminder_notifications.xlsx', 'Holiday Reminder'),
        ('09_late_submission_alerts.xlsx', 'Late Submission')
    ]
    
    # Only update network folder (not local)
    network_folder = Path(network_folder_path) if network_folder_path else Path('G:/Test')
    
    if not network_folder.exists():
        network_folder.mkdir(exist_ok=True)
        logger.info(f"Created network folder: {network_folder}")
    
    # Create files ONLY in network folder
    for file_name, notification_type in files_to_create:
        # Create in network folder only
        network_file = network_folder / file_name
        create_notification_excel(network_file, notification_type, vendors_data, vendor_emails)
    
    print("\n✅ Successfully populated Excel files with vendor data!")
    print(f"   Network folder: {network_folder}")
    print(f"   Total vendors: {len(vendors_data)}")
    print("   Note: Local folder files remain unchanged")

if __name__ == "__main__":
    main()
