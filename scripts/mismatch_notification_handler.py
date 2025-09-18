#!/usr/bin/env python3
"""
Mismatch Notifications Handler with Real-time Sync
==================================================
This module handles updating 04_mismatch_notifications.xlsx in real-time when 
attendance mismatches are detected. It includes:

1. Real-time mismatch detection
2. Local Excel file updates with table formatting
3. Network drive synchronization with sync features intact
4. Integration with existing notification system

Features:
- Automatic mismatch detection
- Real-time Excel updates
- Table formatting for Power Automate
- Stop notifications flag support
- Force update capabilities
- Comprehensive logging and error handling
- Network drive sync with conflict resolution
"""

import os
import pandas as pd
import shutil
import json
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
import sqlite3
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class MismatchData:
    """Data structure for mismatch information"""
    employee_id: str
    contact_email: str
    employee_name: str
    mismatch_date: date
    web_status: str
    swipe_status: str
    mismatch_type: str
    severity: str
    explanation_required: bool
    manager_id: str = None
    manager_email: str = None
    notification_required: bool = True
    priority: str = "HIGH"

class MismatchNotificationHandler:
    """Handles mismatch notifications with real-time Excel sync"""
    
    def __init__(self, 
                 db_path: str = "vendor_timesheet.db",
                 local_excel_folder: str = "notification_configs",
                 network_folder: str = "network_folder_simplified"):
        
        self.db_path = Path(db_path)
        self.local_excel_folder = Path(local_excel_folder)
        self.network_folder = Path(network_folder)
        
        # Excel file paths
        self.local_excel_file = self.local_excel_folder / "04_mismatch_notifications.xlsx"
        self.network_excel_file = self.network_folder / "04_mismatch_notifications.xlsx"
        
        # Sync control files
        self.sync_control_file = self.network_folder / "mismatch_sync_control.json"
        self.notification_context_file = self.network_folder / "notification_context.json"
        
        # Ensure folders exist
        self.local_excel_folder.mkdir(exist_ok=True, parents=True)
        self.network_folder.mkdir(exist_ok=True, parents=True)
        
        # Load sync settings
        self.sync_settings = self._load_sync_settings()
        
    def _load_sync_settings(self) -> Dict[str, Any]:
        """Load sync control settings from network folder"""
        try:
            if self.sync_control_file.exists():
                with open(self.sync_control_file, 'r') as f:
                    settings = json.load(f)
                logger.info(f"✅ Loaded mismatch sync settings: {len(settings)} configurations")
                return settings
            else:
                # Create default sync settings
                default_settings = {
                    "global_sync_enabled": True,
                    "stop_notifications": False,
                    "force_update_mode": False,
                    "auto_retry_enabled": True,
                    "max_retry_count": 3,
                    "sync_interval_minutes": 5,  # More frequent for mismatches
                    "mismatch_detection_enabled": True,
                    "severity_levels": ["HIGH", "MEDIUM", "LOW"],
                    "employee_settings": {},
                    "last_sync_time": datetime.now().isoformat()
                }
                self._save_sync_settings(default_settings)
                return default_settings
        except Exception as e:
            logger.error(f"❌ Error loading sync settings: {str(e)}")
            return {"global_sync_enabled": True, "stop_notifications": False}
    
    def _save_sync_settings(self, settings: Dict[str, Any]):
        """Save sync control settings to network folder"""
        try:
            settings["last_updated"] = datetime.now().isoformat()
            with open(self.sync_control_file, 'w') as f:
                json.dump(settings, f, indent=2)
            logger.info("✅ Mismatch sync settings saved")
        except Exception as e:
            logger.error(f"❌ Error saving sync settings: {str(e)}")
    
    def detect_real_time_mismatches(self, check_date: date = None) -> List[MismatchData]:
        """
        Detect attendance mismatches in real-time
        
        Args:
            check_date: Date to check mismatches for (defaults to today)
            
        Returns:
            List of MismatchData objects for detected mismatches
        """
        if check_date is None:
            check_date = date.today()
        
        logger.info(f"🔍 Detecting real-time mismatches for {check_date}")
        
        try:
            conn = sqlite3.connect(self.db_path)
            
            # Query to detect mismatches between web status and swipe records
            query = """
            SELECT 
                v.vendor_id as employee_id,
                v.full_name as employee_name,
                u.email as contact_email,
                ds.status as web_status,
                ds.status_date,
                sr.login_time,
                sr.logout_time,
                sr.attendance_status,
                v.manager_id,
                mu.email as manager_email,
                m.full_name as manager_name
            FROM vendors v
            JOIN users u ON v.user_id = u.id
            LEFT JOIN daily_statuses ds ON v.id = ds.vendor_id AND ds.status_date = ?
            LEFT JOIN swipe_records sr ON v.id = sr.vendor_id AND sr.attendance_date = ?
            LEFT JOIN managers m ON v.manager_id = m.manager_id
            LEFT JOIN users mu ON m.user_id = mu.id
            WHERE u.is_active = 1
            ORDER BY v.vendor_id
            """
            
            df = pd.read_sql_query(query, conn, params=[check_date.isoformat(), check_date.isoformat()])
            conn.close()
            
            mismatches = []
            
            for _, row in df.iterrows():
                mismatch = self._analyze_mismatch(row, check_date)
                if mismatch:
                    # Check if not already notified
                    if not self._already_notified_mismatch(mismatch.employee_id, check_date):
                        mismatches.append(mismatch)
                        logger.info(f"🚨 Detected mismatch: {mismatch.employee_name} - {mismatch.mismatch_type}")
            
            logger.info(f"📊 Found {len(mismatches)} new mismatches requiring notifications")
            return mismatches
            
        except Exception as e:
            logger.error(f"❌ Error detecting mismatches: {str(e)}")
            return []
    
    def _analyze_mismatch(self, row, check_date: date) -> Optional[MismatchData]:
        """Analyze a row to determine if it's a mismatch"""
        try:
            web_status = row.get('web_status', '').upper() if row.get('web_status') else 'NOT_SUBMITTED'
            login_time = row.get('login_time')
            logout_time = row.get('logout_time')
            attendance_status = row.get('attendance_status', '')
            
            # Determine swipe status based on actual data
            if login_time and logout_time:
                swipe_status = 'FULL_DAY_OFFICE'
            elif login_time or logout_time:
                swipe_status = 'PARTIAL_SWIPE'
            elif attendance_status in ['AA', 'AB']:  # Absent codes
                swipe_status = 'ABSENT_MARKED'
            else:
                swipe_status = 'NO_SWIPE'
            
            # Detect mismatch patterns
            mismatch_type = None
            severity = "MEDIUM"
            
            if web_status == 'NOT_SUBMITTED' and swipe_status != 'NO_SWIPE':
                mismatch_type = 'NO_WEB_STATUS_BUT_SWIPED'
                severity = "HIGH"
            elif web_status == 'WORK_FROM_HOME' and swipe_status != 'NO_SWIPE':
                mismatch_type = 'WFH_BUT_SWIPED'
                severity = "HIGH"
            elif web_status == 'ON_LEAVE' and swipe_status != 'NO_SWIPE':
                mismatch_type = 'ON_LEAVE_BUT_SWIPED'
                severity = "HIGH"
            elif web_status in ['IN_OFFICE_FULL', 'IN_OFFICE_HALF'] and swipe_status == 'NO_SWIPE':
                mismatch_type = 'OFFICE_BUT_NO_SWIPE'
                severity = "HIGH"
            elif web_status in ['IN_OFFICE_FULL'] and swipe_status == 'PARTIAL_SWIPE':
                mismatch_type = 'FULL_DAY_BUT_PARTIAL_SWIPE'
                severity = "MEDIUM"
            
            if mismatch_type:
                return MismatchData(
                    employee_id=row['employee_id'],
                    contact_email=row['contact_email'],
                    employee_name=row['employee_name'],
                    mismatch_date=check_date,
                    web_status=web_status,
                    swipe_status=swipe_status,
                    mismatch_type=mismatch_type,
                    severity=severity,
                    explanation_required=True,
                    manager_id=row.get('manager_id'),
                    manager_email=row.get('manager_email'),
                    priority=severity
                )
            
            return None
            
        except Exception as e:
            logger.error(f"❌ Error analyzing mismatch for {row.get('employee_id', 'unknown')}: {str(e)}")
            return None
    
    def _already_notified_mismatch(self, employee_id: str, check_date: date) -> bool:
        """Check if employee was already notified for mismatch on this date"""
        try:
            if not self.local_excel_file.exists():
                return False
            
            df = pd.read_excel(self.local_excel_file)
            
            # Check for existing notification for same employee and date
            if 'MismatchDate' in df.columns:
                existing = df[
                    (df['EmployeeID'] == employee_id) & 
                    (pd.to_datetime(df['MismatchDate']).dt.date == check_date) &
                    (df.get('NotificationSent', False) == True)
                ]
                return not existing.empty
            
            return False
            
        except Exception as e:
            logger.error(f"❌ Error checking notification history: {str(e)}")
            return False
    
    def update_local_excel(self, mismatches: List[MismatchData]) -> bool:
        """
        Update the local Excel file with mismatch data
        
        Args:
            mismatches: List of mismatch data to update
            
        Returns:
            True if update successful, False otherwise
        """
        if not mismatches:
            logger.info("ℹ️ No mismatches to update")
            return True
        
        logger.info(f"📝 Updating local Excel file with {len(mismatches)} mismatches")
        
        try:
            # Load existing Excel file or create new one
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                logger.info(f"📖 Loaded existing Excel with {len(df)} records")
            else:
                # Create new DataFrame with proper structure
                df = pd.DataFrame(columns=[
                    'RecordID', 'EmployeeID', 'ContactEmail', 'EmployeeName',
                    'MismatchDate', 'WebStatus', 'SwipeStatus', 'MismatchType',
                    'Message', 'NotificationType', 'Priority', 'Severity',
                    'CreatedTime', 'NotificationSent', 'NotificationSentTime',
                    'ExplanationRequired', 'ManagerID', 'ManagerEmail', 'RetryCount'
                ])
                logger.info("📝 Created new Excel file structure")
            
            # Add new mismatch records
            new_records = []
            current_time = datetime.now()
            
            for mismatch in mismatches:
                # Generate unique RecordID
                record_id = f"MISMATCH_{mismatch.employee_id}_{mismatch.mismatch_date.strftime('%Y%m%d')}_{current_time.strftime('%H%M%S')}"
                
                # Create notification message
                message = self._generate_mismatch_message(mismatch)
                
                # Create new record
                new_record = {
                    'RecordID': record_id,
                    'EmployeeID': mismatch.employee_id,
                    'ContactEmail': mismatch.contact_email,
                    'EmployeeName': mismatch.employee_name,
                    'MismatchDate': mismatch.mismatch_date,
                    'WebStatus': mismatch.web_status,
                    'SwipeStatus': mismatch.swipe_status,
                    'MismatchType': mismatch.mismatch_type,
                    'Message': message,
                    'NotificationType': 'MISMATCH_ALERT',
                    'Priority': mismatch.priority,
                    'Severity': mismatch.severity,
                    'CreatedTime': current_time,
                    'NotificationSent': False,
                    'NotificationSentTime': pd.NaT,
                    'ExplanationRequired': mismatch.explanation_required,
                    'ManagerID': mismatch.manager_id or '',
                    'ManagerEmail': mismatch.manager_email or '',
                    'RetryCount': 0
                }
                
                new_records.append(new_record)
                logger.info(f"➕ Added mismatch record for {mismatch.employee_name}")
            
            # Append new records to DataFrame
            if new_records:
                new_df = pd.DataFrame(new_records)
                df = pd.concat([df, new_df], ignore_index=True)
                
                # Sort by CreatedTime descending (newest first)
                df = df.sort_values('CreatedTime', ascending=False).reset_index(drop=True)
            
            # Save updated DataFrame
            df.to_excel(self.local_excel_file, index=False, engine='openpyxl')
            logger.info(f"✅ Local Excel updated successfully with {len(new_records)} new records")
            
            # Format as table for Power Automate compatibility
            table_success = self._format_excel_as_table(
                self.local_excel_file, 
                "MismatchNotifications"
            )
            
            if table_success:
                logger.info("✅ Excel file formatted as table")
            else:
                logger.warning("⚠️ Failed to format Excel as table, but data was saved")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error updating local Excel: {str(e)}")
            return False
    
    def _generate_mismatch_message(self, mismatch: MismatchData) -> str:
        """Generate notification message for mismatch"""
        mismatch_descriptions = {
            'NO_WEB_STATUS_BUT_SWIPED': 'You swiped but did not submit web status',
            'WFH_BUT_SWIPED': 'You marked Work From Home but have swipe records',
            'ON_LEAVE_BUT_SWIPED': 'You marked On Leave but have swipe records',
            'OFFICE_BUT_NO_SWIPE': 'You marked In Office but have no swipe records',
            'FULL_DAY_BUT_PARTIAL_SWIPE': 'You marked Full Day but have partial swipe records'
        }
        
        description = mismatch_descriptions.get(mismatch.mismatch_type, 'Attendance mismatch detected')
        
        return (f"🚨 Attendance Mismatch Alert - {description} for {mismatch.mismatch_date.strftime('%B %d, %Y')}. "
                f"Web Status: {mismatch.web_status.replace('_', ' ').title()}, "
                f"Swipe Status: {mismatch.swipe_status.replace('_', ' ').title()}. "
                f"{'Please provide an explanation.' if mismatch.explanation_required else ''}")
    
    def _format_excel_as_table(self, excel_file_path: Path, table_name: str = "MismatchNotifications") -> bool:
        """Format Excel file as a proper table for Power Automate compatibility"""
        try:
            logger.info(f"🎨 Formatting Excel file as table: {excel_file_path}")
            
            # Load the workbook
            wb = load_workbook(excel_file_path)
            ws = wb.active
            
            # Check if data exists
            if ws.max_row < 2:  # Only header row or empty
                logger.info("ℹ️ No data to format as table")
                return True
            
            # Remove existing tables if any
            tables_to_remove = []
            for table in ws.tables.values():
                tables_to_remove.append(table.displayName)
            
            for table_name_to_remove in tables_to_remove:
                del ws.tables[table_name_to_remove]
                logger.info(f"🗑️ Removed existing table: {table_name_to_remove}")
            
            # Define the data range
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Create table reference
            table_range = f"A1:{get_column_letter(max_col)}{max_row}"
            
            # Create table
            table = Table(displayName=table_name, ref=table_range)
            
            # Add table style
            style = TableStyleInfo(
                name="TableStyleMedium12",  # Red/orange table style for alerts
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            # Add the table to the worksheet
            ws.add_table(table)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                ws.column_dimensions[column].width = adjusted_width
            
            # Save the workbook
            wb.save(excel_file_path)
            wb.close()
            
            logger.info(f"✅ Excel file formatted as table: {table_name}")
            logger.info(f"📊 Table range: {table_range} ({max_row} rows, {max_col} columns)")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error formatting Excel as table: {str(e)}")
            return False
    
    def sync_to_network_drive(self, preserve_settings: bool = True) -> bool:
        """
        Sync local Excel file to network drive with sync features intact
        
        Args:
            preserve_settings: Whether to preserve existing sync settings
            
        Returns:
            True if sync successful, False otherwise
        """
        logger.info("🔄 Starting mismatch notifications network drive sync...")
        
        # Check global sync settings
        if not self.sync_settings.get("global_sync_enabled", True):
            logger.info("⏭️ Sync disabled globally - skipping")
            return True
        
        if self.sync_settings.get("stop_notifications", False):
            logger.info("🛑 Stop notifications flag set - skipping sync")
            return True
        
        try:
            # Ensure network folder exists
            self.network_folder.mkdir(exist_ok=True, parents=True)
            
            # Check for conflicts if network file exists
            conflict_resolved = True
            if self.network_excel_file.exists() and preserve_settings:
                conflict_resolved = self._resolve_sync_conflicts()
            
            if not conflict_resolved:
                logger.warning("⚠️ Sync conflict not resolved - aborting")
                return False
            
            # Copy local file to network drive
            if self.local_excel_file.exists():
                shutil.copy2(self.local_excel_file, self.network_excel_file)
                logger.info("✅ Excel file synced to network drive")
                
                # Format network file as table for Power Automate
                table_success = self._format_excel_as_table(
                    self.network_excel_file, 
                    "MismatchNotifications"
                )
                
                if table_success:
                    logger.info("✅ Network Excel file formatted as table")
                else:
                    logger.warning("⚠️ Failed to format network Excel as table")
                
                # Update sync metadata
                self._update_sync_metadata()
                
                # Update notification context for Power Automate
                self._update_notification_context()
                
                return True
            else:
                logger.warning("⚠️ Local Excel file not found - nothing to sync")
                return False
                
        except Exception as e:
            logger.error(f"❌ Error syncing to network drive: {str(e)}")
            return False
    
    def _resolve_sync_conflicts(self) -> bool:
        """Resolve conflicts between local and network files"""
        try:
            # Check modification times
            local_mtime = self.local_excel_file.stat().st_mtime if self.local_excel_file.exists() else 0
            network_mtime = self.network_excel_file.stat().st_mtime if self.network_excel_file.exists() else 0
            
            # If force update is enabled, always prefer local
            if self.sync_settings.get("force_update_mode", False):
                logger.info("🔄 Force update mode - using local file")
                return True
            
            # If network file is newer, merge records
            if network_mtime > local_mtime:
                logger.info("🔄 Network file is newer - merging records")
                return self._merge_excel_files()
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error resolving sync conflicts: {str(e)}")
            return False
    
    def _merge_excel_files(self) -> bool:
        """Merge local and network Excel files to avoid data loss"""
        try:
            # Load both files
            local_df = pd.read_excel(self.local_excel_file) if self.local_excel_file.exists() else pd.DataFrame()
            network_df = pd.read_excel(self.network_excel_file) if self.network_excel_file.exists() else pd.DataFrame()
            
            if local_df.empty and network_df.empty:
                return True
            
            # Combine and remove duplicates based on RecordID
            combined_df = pd.concat([network_df, local_df], ignore_index=True)
            merged_df = combined_df.drop_duplicates(subset=['RecordID'], keep='last')
            
            # Sort by CreatedTime descending
            merged_df = merged_df.sort_values('CreatedTime', ascending=False).reset_index(drop=True)
            
            # Save merged result to local file
            merged_df.to_excel(self.local_excel_file, index=False, engine='openpyxl')
            
            logger.info(f"✅ Merged Excel files: {len(merged_df)} total records")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error merging Excel files: {str(e)}")
            return False
    
    def _update_sync_metadata(self):
        """Update sync metadata in network folder"""
        try:
            metadata = {
                "last_sync_time": datetime.now().isoformat(),
                "source": "mismatch_notification_handler",
                "sync_status": "success",
                "local_file_path": str(self.local_excel_file),
                "network_file_path": str(self.network_excel_file),
                "records_synced": self._get_record_count()
            }
            
            metadata_file = self.network_folder / "mismatch_sync_metadata.json"
            with open(metadata_file, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            # Update global sync settings
            self.sync_settings["last_sync_time"] = metadata["last_sync_time"]
            self._save_sync_settings(self.sync_settings)
            
        except Exception as e:
            logger.error(f"❌ Error updating sync metadata: {str(e)}")
    
    def _update_notification_context(self):
        """Update notification context for Power Automate integration"""
        try:
            context = {
                "mismatch_notifications": {
                    "last_updated": datetime.now().isoformat(),
                    "file_path": str(self.network_excel_file),
                    "records_available": self._get_record_count(),
                    "pending_notifications": self._get_pending_notification_count(),
                    "sync_enabled": not self.sync_settings.get("stop_notifications", False)
                }
            }
            
            # Load existing context if it exists
            if self.notification_context_file.exists():
                with open(self.notification_context_file, 'r') as f:
                    existing_context = json.load(f)
                existing_context.update(context)
                context = existing_context
            
            # Save updated context
            with open(self.notification_context_file, 'w') as f:
                json.dump(context, f, indent=2)
            
            logger.info("✅ Notification context updated for Power Automate")
            
        except Exception as e:
            logger.error(f"❌ Error updating notification context: {str(e)}")
    
    def _get_record_count(self) -> int:
        """Get total number of records in local Excel file"""
        try:
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                return len(df)
            return 0
        except:
            return 0
    
    def _get_pending_notification_count(self) -> int:
        """Get number of pending notifications in local Excel file"""
        try:
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                pending = df[df.get('NotificationSent', True) == False]
                return len(pending)
            return 0
        except:
            return 0
    
    def mark_notifications_sent(self, record_ids: List[str]) -> bool:
        """
        Mark notifications as sent in both local and network Excel files
        
        Args:
            record_ids: List of RecordIDs to mark as sent
            
        Returns:
            True if update successful, False otherwise
        """
        logger.info(f"✅ Marking {len(record_ids)} mismatch notifications as sent")
        
        try:
            # Update local file
            if self.local_excel_file.exists():
                df = pd.read_excel(self.local_excel_file)
                
                # Update matching records
                mask = df['RecordID'].isin(record_ids)
                if mask.any():
                    df.loc[mask, 'NotificationSent'] = True
                    df.loc[mask, 'NotificationSentTime'] = datetime.now()
                    
                    # Save updated file
                    df.to_excel(self.local_excel_file, index=False, engine='openpyxl')
                    
                    # Format as table
                    self._format_excel_as_table(self.local_excel_file, "MismatchNotifications")
                    
                    # Sync to network drive
                    self.sync_to_network_drive()
                    
                    logger.info(f"✅ Updated {mask.sum()} mismatch notification records")
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"❌ Error marking notifications as sent: {str(e)}")
            return False
    
    def process_real_time_mismatches(self, check_date: date = None) -> Dict[str, Any]:
        """
        Complete workflow: detect mismatches, update Excel, sync to network
        
        Args:
            check_date: Date to process mismatches for (defaults to today)
            
        Returns:
            Dictionary with processing results
        """
        if check_date is None:
            check_date = date.today()
        
        logger.info("🚀 Starting real-time mismatch processing workflow")
        
        results = {
            "date": check_date.isoformat(),
            "mismatches_detected": 0,
            "local_update_success": False,
            "network_sync_success": False,
            "notifications_required": 0,
            "errors": []
        }
        
        try:
            # Step 1: Detect real-time mismatches
            mismatches = self.detect_real_time_mismatches(check_date)
            results["mismatches_detected"] = len(mismatches)
            results["notifications_required"] = len([m for m in mismatches if m.notification_required])
            
            if not mismatches:
                logger.info("ℹ️ No new mismatches detected")
                results["local_update_success"] = True
                results["network_sync_success"] = True
                return results
            
            # Step 2: Update local Excel file
            local_success = self.update_local_excel(mismatches)
            results["local_update_success"] = local_success
            
            if not local_success:
                results["errors"].append("Failed to update local Excel file")
                return results
            
            # Step 3: Sync to network drive
            network_success = self.sync_to_network_drive()
            results["network_sync_success"] = network_success
            
            if not network_success:
                results["errors"].append("Failed to sync to network drive")
            
            # Step 4: Log summary
            logger.info(f"📊 Mismatch processing completed:")
            logger.info(f"  - Mismatches detected: {results['mismatches_detected']}")
            logger.info(f"  - Notifications required: {results['notifications_required']}")
            logger.info(f"  - Local update: {'✅' if local_success else '❌'}")
            logger.info(f"  - Network sync: {'✅' if network_success else '❌'}")
            
            return results
            
        except Exception as e:
            error_msg = f"Fatal error in mismatch processing: {str(e)}"
            logger.error(f"❌ {error_msg}")
            results["errors"].append(error_msg)
            return results
    
    def get_pending_notifications(self) -> List[Dict[str, Any]]:
        """Get list of pending mismatch notifications that need to be sent"""
        try:
            if not self.local_excel_file.exists():
                return []
            
            df = pd.read_excel(self.local_excel_file)
            pending = df[df.get('NotificationSent', True) == False].to_dict('records')
            
            logger.info(f"📋 Found {len(pending)} pending mismatch notifications")
            return pending
            
        except Exception as e:
            logger.error(f"❌ Error getting pending notifications: {str(e)}")
            return []
    
    def configure_employee_sync_settings(self, employee_id: str, settings: Dict[str, Any]):
        """Configure sync settings for a specific employee"""
        if "employee_settings" not in self.sync_settings:
            self.sync_settings["employee_settings"] = {}
        
        self.sync_settings["employee_settings"][employee_id] = settings
        self._save_sync_settings(self.sync_settings)
        
        logger.info(f"✅ Updated sync settings for employee {employee_id}")

# Global instance for easy access
mismatch_notification_handler = MismatchNotificationHandler()

def process_real_time_mismatches(check_date: date = None) -> Dict[str, Any]:
    """Public function to process real-time mismatches"""
    return mismatch_notification_handler.process_real_time_mismatches(check_date)

def mark_mismatch_notification_sent(record_ids: List[str]) -> bool:
    """Public function to mark mismatch notifications as sent"""
    return mismatch_notification_handler.mark_notifications_sent(record_ids)

def get_pending_mismatch_notifications() -> List[Dict[str, Any]]:
    """Public function to get pending mismatch notifications"""
    return mismatch_notification_handler.get_pending_notifications()

def configure_mismatch_sync_settings(employee_id: str = None, **settings):
    """Public function to configure sync settings"""
    if employee_id:
        mismatch_notification_handler.configure_employee_sync_settings(employee_id, settings)
    else:
        # Update global settings
        mismatch_notification_handler.sync_settings.update(settings)
        mismatch_notification_handler._save_sync_settings(mismatch_notification_handler.sync_settings)

def convert_mismatch_to_table_format(excel_file_path: str = None, table_name: str = "MismatchNotifications") -> bool:
    """Public function to convert existing mismatch Excel file to table format"""
    if excel_file_path is None:
        excel_file_path = "notification_configs/04_mismatch_notifications.xlsx"
    
    try:
        file_path = Path(excel_file_path)
        if not file_path.exists():
            print(f"❌ Excel file not found: {excel_file_path}")
            return False
        
        print(f"🎨 Converting mismatch Excel file to table format: {excel_file_path}")
        
        # Use the handler's method to format as table
        success = mismatch_notification_handler._format_excel_as_table(file_path, table_name)
        
        if success:
            print(f"✅ Successfully converted to table format: {table_name}")
            print(f"📊 File ready for Power Automate integration")
        else:
            print(f"❌ Failed to convert to table format")
        
        return success
        
    except Exception as e:
        print(f"❌ Error converting to table format: {str(e)}")
        return False

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Mismatch Notification Handler')
    parser.add_argument('--date', type=str, help='Date to process (YYYY-MM-DD format)')
    parser.add_argument('--action', choices=['process', 'pending', 'sync', 'status', 'format-table'], 
                        default='process', help='Action to perform')
    parser.add_argument('--employee-id', type=str, help='Specific employee ID to configure')
    parser.add_argument('--stop-notifications', action='store_true', 
                        help='Stop notifications for employee')
    parser.add_argument('--force-update', action='store_true', 
                        help='Force update for employee')
    parser.add_argument('--excel-file', type=str, 
                        help='Excel file path (for format-table action)')
    parser.add_argument('--table-name', type=str, default='MismatchNotifications',
                        help='Table name for formatting')
    
    args = parser.parse_args()
    
    target_date = None
    if args.date:
        try:
            target_date = datetime.strptime(args.date, '%Y-%m-%d').date()
        except ValueError:
            print("❌ Invalid date format. Use YYYY-MM-DD")
            exit(1)
    
    if args.action == 'process':
        results = process_real_time_mismatches(target_date)
        print("📊 Processing Results:")
        for key, value in results.items():
            print(f"  {key}: {value}")
    
    elif args.action == 'pending':
        pending = get_pending_mismatch_notifications()
        print(f"📋 Found {len(pending)} pending mismatch notifications:")
        for notification in pending:
            print(f"  - {notification.get('EmployeeName', notification.get('EmployeeID'))}: {notification.get('MismatchType', 'Unknown')}")
    
    elif args.action == 'sync':
        success = mismatch_notification_handler.sync_to_network_drive()
        print(f"🔄 Network sync: {'✅ Success' if success else '❌ Failed'}")
    
    elif args.action == 'status':
        record_count = mismatch_notification_handler._get_record_count()
        pending_count = mismatch_notification_handler._get_pending_notification_count()
        print(f"📊 Status:")
        print(f"  Total records: {record_count}")
        print(f"  Pending notifications: {pending_count}")
        print(f"  Sync enabled: {not mismatch_notification_handler.sync_settings.get('stop_notifications', False)}")
    
    elif args.action == 'format-table':
        success = convert_mismatch_to_table_format(args.excel_file, args.table_name)
        print(f"🎨 Table formatting: {'✅ Success' if success else '❌ Failed'}")
    
    # Handle employee-specific configuration
    if args.employee_id:
        settings = {}
        if args.stop_notifications:
            settings['stop_notifications'] = True
        if args.force_update:
            settings['force_update'] = True
        
        if settings:
            configure_mismatch_sync_settings(args.employee_id, **settings)
            print(f"✅ Updated settings for employee {args.employee_id}")
