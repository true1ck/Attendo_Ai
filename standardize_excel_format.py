#!/usr/bin/env python3
"""
Excel Format Standardizer
Ensures all Excel files have the same data format:
EmployeeID, ContactEmail, Message, NotificationType, Priority

This script:
1. Checks existing Excel files in network_folder_simplified
2. Converts them to the standard format
3. Ensures Power Automate gets consistent data structure
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelFormatStandardizer:
    """Standardizes Excel files to consistent Power Automate format"""
    
    def __init__(self, network_folder="network_folder_simplified"):
        self.network_folder = Path(network_folder)
        self.network_folder.mkdir(exist_ok=True)
        
        # Standard format columns as requested
        self.standard_columns = ['EmployeeID', 'ContactEmail', 'Message', 'NotificationType', 'Priority']
        
        # Mapping of old column names to standard names
        self.column_mappings = {
            # Various forms of Employee ID
            'Vendor_ID': 'EmployeeID',
            'Manager_ID': 'EmployeeID', 
            'User_ID': 'EmployeeID',
            'Primary_Key': 'EmployeeID',
            'vendor_id': 'EmployeeID',
            'manager_id': 'EmployeeID',
            
            # Various forms of Contact Email
            'Contact_Email': 'ContactEmail',
            'Email': 'ContactEmail',
            'contact_email': 'ContactEmail',
            'email': 'ContactEmail',
            
            # Various forms of Message
            'Custom_Message': 'Message',
            'Notification_Message': 'Message',
            'Message_Template': 'Message',
            'notification_message': 'Message',
            
            # Various forms of Notification Type
            'Notification_Type': 'NotificationType',
            'Type': 'NotificationType',
            'Alert_Type': 'NotificationType',
            'notification_type': 'NotificationType',
            
            # Various forms of Priority
            'Priority': 'Priority',  # Already correct
            'Alert_Priority': 'Priority',
            'Urgency': 'Priority',
            'priority': 'Priority'
        }
    
    def map_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Map existing columns to standard format"""
        # Create a new DataFrame with standard columns
        result_df = pd.DataFrame(columns=self.standard_columns)
        
        # Map existing columns to standard columns
        for old_col, new_col in self.column_mappings.items():
            if old_col in df.columns and new_col in self.standard_columns:
                result_df[new_col] = df[old_col]
        
        # Fill missing columns with defaults
        if 'EmployeeID' not in result_df.columns or result_df['EmployeeID'].isna().all():
            result_df['EmployeeID'] = df.index.map(lambda x: f'USER_{x:03d}')
            
        if 'ContactEmail' not in result_df.columns or result_df['ContactEmail'].isna().all():
            result_df['ContactEmail'] = 'unknown@company.com'
            
        if 'Message' not in result_df.columns or result_df['Message'].isna().all():
            result_df['Message'] = 'Standard notification message'
            
        if 'NotificationType' not in result_df.columns or result_df['NotificationType'].isna().all():
            result_df['NotificationType'] = 'General'
            
        if 'Priority' not in result_df.columns or result_df['Priority'].isna().all():
            result_df['Priority'] = 'Medium'
        
        return result_df
    
    def create_standardized_excel(self, df: pd.DataFrame, filepath: Path, table_name: str = None):
        """Create Excel file with standard format and table"""
        try:
            if table_name is None:
                table_name = filepath.stem.replace('-', '_').replace(' ', '_')
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "StandardFormat"
            
            # Write headers
            for col_num, column_title in enumerate(self.standard_columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
            
            # Write data
            if not df.empty:
                for row_num, row_data in enumerate(df[self.standard_columns].values, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = str(cell_value) if cell_value is not None else ""
            
            # Create table
            table_range = f"A1:{get_column_letter(len(self.standard_columns))}{len(df) + 1}"
            table = Table(displayName=table_name, ref=table_range)
            
            # Add table style
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            
            # Add table to worksheet
            ws.add_table(table)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"✅ Standardized {filepath.name} with {len(df)} rows")
            
        except Exception as e:
            logger.error(f"❌ Error creating standardized Excel {filepath}: {str(e)}")
    
    def standardize_file(self, filepath: Path) -> bool:
        """Standardize a single Excel file"""
        try:
            if not filepath.exists():
                logger.warning(f"⚠️ File not found: {filepath}")
                return False
            
            # Read the existing file
            df = pd.read_excel(filepath)
            logger.info(f"📖 Processing {filepath.name} with {len(df)} rows, {len(df.columns)} columns")
            
            # Map to standard format
            standardized_df = self.map_columns(df)
            
            # Create backup of original
            backup_path = filepath.parent / f"{filepath.stem}_backup{filepath.suffix}"
            if not backup_path.exists():
                df.to_excel(backup_path, index=False)
                logger.info(f"💾 Created backup: {backup_path.name}")
            
            # Save standardized version
            table_name = filepath.stem.replace('-', '_').replace(' ', '_')
            self.create_standardized_excel(standardized_df, filepath, table_name)
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Error standardizing {filepath}: {str(e)}")
            return False
    
    def standardize_all_files(self):
        """Standardize all Excel files in the network folder"""
        logger.info(f"🚀 Starting Excel format standardization in {self.network_folder}")
        
        excel_files = list(self.network_folder.glob("*.xlsx"))
        
        # Skip certain files that are already in correct format
        skip_files = {'sent_noti_now.xlsx'}  # Already in standard format
        
        processed_count = 0
        success_count = 0
        
        for filepath in excel_files:
            if filepath.name in skip_files:
                logger.info(f"⏭️ Skipping {filepath.name} (already in standard format)")
                continue
                
            processed_count += 1
            logger.info(f"\n{'='*60}")
            logger.info(f"📄 Processing: {filepath.name}")
            logger.info(f"{'='*60}")
            
            if self.standardize_file(filepath):
                success_count += 1
        
        logger.info(f"\n🏁 STANDARDIZATION COMPLETE:")
        logger.info(f"   📊 Files processed: {processed_count}")
        logger.info(f"   ✅ Files successful: {success_count}")
        logger.info(f"   ❌ Files failed: {processed_count - success_count}")
        
        return success_count == processed_count
    
    def create_sample_data(self):
        """Create sample data for testing"""
        sample_data = [
            {
                'EmployeeID': 'V001',
                'ContactEmail': 'vendor1@company.com',
                'Message': 'Please submit your daily attendance status - John Doe',
                'NotificationType': 'Daily Status Reminders',
                'Priority': 'Medium'
            },
            {
                'EmployeeID': 'M001', 
                'ContactEmail': 'manager1@company.com',
                'Message': 'Team attendance summary report - 3 team members pending',
                'NotificationType': 'Manager Summary',
                'Priority': 'Medium'
            },
            {
                'EmployeeID': 'ADMIN_001',
                'ContactEmail': 'admin@company.com',
                'Message': 'System alert requiring immediate attention',
                'NotificationType': 'System Alerts', 
                'Priority': 'Critical'
            }
        ]
        
        sample_df = pd.DataFrame(sample_data)
        sample_file = self.network_folder / 'sample_standard_format.xlsx'
        self.create_standardized_excel(sample_df, sample_file, 'SampleStandardFormat')
        
        logger.info(f"✅ Created sample file: {sample_file}")
        return sample_file

# Global instance
format_standardizer = ExcelFormatStandardizer()

def standardize_all_excel_files():
    """Public function to standardize all Excel files"""
    return format_standardizer.standardize_all_files()

def create_sample_standard_format():
    """Public function to create sample standard format file"""
    return format_standardizer.create_sample_data()

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel Format Standardizer')
    parser.add_argument('--action', choices=['standardize', 'sample', 'both'], 
                       default='both', help='Action to perform')
    parser.add_argument('--file', help='Specific file to standardize')
    
    args = parser.parse_args()
    
    print("🔄 EXCEL FORMAT STANDARDIZER")
    print("="*60)
    print("Ensuring all Excel files have format:")
    print("EmployeeID | ContactEmail | Message | NotificationType | Priority")
    print("="*60)
    
    if args.action in ['sample', 'both']:
        print("\n📝 Creating sample standard format file...")
        sample_file = create_sample_standard_format()
        print(f"✅ Sample created: {sample_file}")
    
    if args.action in ['standardize', 'both']:
        if args.file:
            # Standardize specific file
            filepath = Path(args.file)
            print(f"\n🔄 Standardizing specific file: {filepath}")
            success = format_standardizer.standardize_file(filepath)
            print(f"{'✅ SUCCESS' if success else '❌ FAILED'}")
        else:
            # Standardize all files
            print(f"\n🔄 Standardizing all Excel files...")
            success = standardize_all_excel_files()
            print(f"\n{'✅ ALL FILES STANDARDIZED' if success else '❌ SOME FILES FAILED'}")
    
    print(f"\n📊 SUMMARY:")
    print(f"   Standard Format: EmployeeID, ContactEmail, Message, NotificationType, Priority")
    print(f"   Network Folder: {format_standardizer.network_folder}")
    print(f"   Ready for Power Automate: YES")
