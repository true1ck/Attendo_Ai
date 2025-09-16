"""
ATTENDO Application - Simple Working Version
This combines everything in one file to avoid circular imports
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date
import os
import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
import pytz
import threading
import shutil
import time
from pathlib import Path
from sqlalchemy.orm import joinedload

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'hackathon-attendo-vendor-timesheet-2025'

# Add JSON filter for templates
import json
@app.template_filter('fromjson')
def fromjson_filter(value):
    """Parse JSON string to Python object"""
    if not value:
        return {}
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return {}
# Use absolute path for SQLite database to ensure it's created in the correct location
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.abspath("vendor_timesheet.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Notification Configuration
app.config['BASE_URL'] = os.getenv('BASE_URL', 'http://localhost:5000')
app.config['SMTP_SERVER'] = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
app.config['SMTP_PORT'] = int(os.getenv('SMTP_PORT', 587))
app.config['SMTP_USER'] = os.getenv('SMTP_USER')  # Set this in environment
app.config['SMTP_PASSWORD'] = os.getenv('SMTP_PASSWORD')  # Set this in environment
app.config['SMS_API_URL'] = os.getenv('SMS_API_URL')  # Optional SMS API endpoint
app.config['SMS_API_KEY'] = os.getenv('SMS_API_KEY')  # Optional SMS API key

# Initialize models first
import models

# Initialize extensions
db = models.db
db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Import other modules
from import_routes import import_bp
from notifications import start_notification_scheduler
from swagger_ui import register_swagger_ui
from notification_service import notification_service, setup_notification_scheduler

# Register blueprints
app.register_blueprint(import_bp)

# Register Swagger UI
register_swagger_ui(app)

# Register Power Automate API
try:
    from scripts.power_automate_api import register_power_automate_api
    register_power_automate_api(app)
except ImportError as e:
    print(f"âš ï¸ Power Automate API not available: {e}")

# Initialize notification service
notification_service.init_app(app)

# Real-time sync configuration
app.config['SYNC_INTERVAL_MINUTES'] = 5
app.config['WEBHOOK_TIMEOUT_SECONDS'] = 30
app.config['ENABLE_POWER_AUTOMATE_WEBHOOKS'] = True
app.config['POWER_AUTOMATE_WEBHOOK_URLS'] = []
app.config['SYNC_MONITORING_INTERVAL_MINUTES'] = 30
app.config['SYNC_VALIDATION_INTERVAL_HOURS'] = 6
app.config['ADMIN_EMAILS'] = ['admin@attendo.com']
app.config['SYNC_ALERT_FROM'] = 'attendo-sync@attendo.com'

# Excel Sync Configuration
app.config['EXCEL_SYNC_ENABLED'] = True
app.config['EXCEL_LOCAL_FOLDER'] = "notification_configs"
app.config['EXCEL_NETWORK_FOLDER'] = None  # Will be set via admin interface
app.config['EXCEL_SYNC_INTERVAL'] = 600  # 10 minutes in seconds

# Excel Sync Global Variables
excel_sync_running = False
excel_sync_paused = False
excel_sync_thread = None
excel_sync_status = {
    'last_sync': None,
    'status': 'Stopped',
    'files_synced': 0,
    'errors': []
}

# Import notification scheduler service
try:
    from notification_scheduler_service import notification_scheduler_service
    notification_scheduler_available = True
except ImportError as e:
    print(f"âš ï¸ Notification scheduler service not available: {e}")
    notification_scheduler_available = False
    notification_scheduler_service = None

@login_manager.user_loader
def load_user(user_id):
    from models import User
    return User.query.get(int(user_id))

def create_tables():
    """Create database tables and ensure default Admin user exists"""
    with app.app_context():
        from models import User, UserRole
        db.create_all()
        
        # Run database migrations if needed
        try:
            from migrate_database import migrate_database
            print("ðŸ”„ Checking for database migrations...")
            migrate_database()
        except Exception as e:
            print(f"âš ï¸ Migration check failed: {e}")
        
        # Ensure default Admin user exists (Admin/admin123)
        from sqlalchemy import or_
        admin_user = User.query.filter(or_(
            User.username == 'Admin',
            User.username == 'admin',
            User.email == 'admin@attendo.com'
        )).first()
        if admin_user:
            # Normalize credentials to requested ones
            admin_user.username = 'Admin'
            admin_user.email = 'admin@attendo.com'
            admin_user.role = UserRole.ADMIN
            admin_user.is_active = True
            admin_user.set_password('admin123')
            db.session.commit()
        else:
            admin_user = User(
                username='Admin',
                email='admin@attendo.com',
                role=UserRole.ADMIN,
                is_active=True
            )
            admin_user.set_password('admin123')
            db.session.add(admin_user)
            db.session.commit()
        
        # Setup notification scheduler after tables are created
        setup_notification_scheduler(app, notification_service)
        
        # Initialize real-time notification sync system
        try:
            from scripts.setup_realtime_notification_sync import setup_complete_notification_sync_system
            print("\nðŸš€ Initializing Real-Time Notification Sync System...")
            sync_components = setup_complete_notification_sync_system(app)
            
            # Store components in app context for access
            app.sync_manager = sync_components['sync_manager']
            app.realtime_sync = sync_components['realtime_sync']
            app.monitor = sync_components['monitor']
            
            print("âœ… Real-Time Notification Sync System initialized successfully!")
        except Exception as e:
            print(f"âš ï¸ Real-Time Sync System initialization failed: {str(e)}")
            print("   App will continue without real-time sync features")

def initialize_demo_data():
    """Initialize the database with demo data for hackathon presentation"""
    from demo_data import create_demo_data
    create_demo_data()

# Import all the route functions directly here
from models import User, Vendor, Manager, DailyStatus, SwipeRecord, Holiday, MismatchRecord, NotificationLog, AuditLog, SystemConfiguration, LeaveRecord, WFHRecord, UserRole, AttendanceStatus, ApprovalStatus
from utils import create_audit_log, generate_monthly_report, import_swipe_data, detect_mismatches, set_system_config, get_system_config, generate_ai_insights

@app.route('/')
def index():
    """Home page - redirect to appropriate dashboard based on user role"""
    if current_user.is_authenticated:
        if current_user.role == UserRole.VENDOR:
            return redirect(url_for('vendor_dashboard'))
        elif current_user.role == UserRole.MANAGER:
            return redirect(url_for('manager_dashboard'))
        elif current_user.role == UserRole.ADMIN:
            return redirect(url_for('admin_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login with clear feedback for common errors"""
    prefill_username = ''
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()
        prefill_username = username

        # Basic validation
        if not username and not password:
            flash('Please enter your username and password.', 'warning')
            return render_template('login.html', prefill_username=prefill_username)
        if not username:
            flash('Please enter your username.', 'warning')
            return render_template('login.html', prefill_username=prefill_username)
        if not password:
            flash('Please enter your password.', 'warning')
            return render_template('login.html', prefill_username=prefill_username)

        # Find user by username (or email fallback if needed later)
        user = User.query.filter_by(username=username).first()

        if not user:
            flash('User not found. Please check your username.', 'error')
            return render_template('login.html', prefill_username=prefill_username)

        if not user.is_active:
            flash('Your account is disabled. Please contact the administrator.', 'error')
            return render_template('login.html', prefill_username=prefill_username)
        
        if not user.check_password(password):
            flash('Incorrect password. Please try again.', 'error')
            return render_template('login.html', prefill_username=prefill_username)

        # Success
        login_user(user)
        user.last_login = datetime.utcnow()
        db.session.commit()
        create_audit_log(user.id, 'LOGIN', 'users', user.id, {}, {'last_login': str(datetime.utcnow())})
        flash(f'Welcome {user.username}!', 'success')
        return redirect(url_for('index'))
    
    return render_template('login.html', prefill_username=prefill_username)

@app.route('/logout')
@login_required
def logout():
    """User logout"""
    create_audit_log(current_user.id, 'LOGOUT', 'users', current_user.id, {}, {})
    logout_user()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

@app.route('/vendor/dashboard')
@login_required
def vendor_dashboard():
    """Vendor dashboard showing status submission and history"""
    if current_user.role != UserRole.VENDOR:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    vendor = current_user.vendor_profile
    if not vendor:
        flash('Vendor profile not found', 'error')
        return redirect(url_for('login'))
    
    # Get manager information
    manager = None
    if vendor.manager_id:
        manager = Manager.query.filter_by(manager_id=vendor.manager_id).first()
    
    # Get today's status
    today = date.today()
    today_status = DailyStatus.query.filter_by(
        vendor_id=vendor.id, 
        status_date=today
    ).first()
    
    # Get recent status history
    recent_statuses = DailyStatus.query.filter_by(
        vendor_id=vendor.id
    ).order_by(DailyStatus.status_date.desc()).limit(10).all()
    
    # Get pending mismatches
    pending_mismatches = MismatchRecord.query.filter_by(
        vendor_id=vendor.id,
        manager_approval=ApprovalStatus.PENDING
    ).all()
    
    # Check if today is weekend or holiday
    from utils import is_non_working_day, get_non_working_day_reason
    is_non_working_day = is_non_working_day(today)
    non_working_reason = get_non_working_day_reason(today) if is_non_working_day else None
    
    # Legacy support - keep existing variables
    is_weekend = today.weekday() >= 5  # Saturday=5, Sunday=6
    is_holiday = Holiday.query.filter_by(holiday_date=today).first() is not None
    
    return render_template('vendor_dashboard.html', 
                         vendor=vendor,
                         manager=manager,
                         today_status=today_status,
                         recent_statuses=recent_statuses,
                         pending_mismatches=pending_mismatches,
                         is_weekend=is_weekend,
                         is_holiday=is_holiday,
                         is_non_working_day=is_non_working_day,
                         non_working_reason=non_working_reason,
                         today=today)

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    """Admin dashboard"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    # Get statistics
    total_vendors = Vendor.query.count()
    total_managers = Manager.query.count()
    total_statuses_today = DailyStatus.query.filter_by(status_date=date.today()).count()
    
    # Upcoming holidays (next 5)
    upcoming_holidays = Holiday.query.filter(
        Holiday.holiday_date >= date.today()
    ).order_by(Holiday.holiday_date.asc()).limit(5).all()
    
    # Get persistent system issues from database
    try:
        from system_issues import SystemIssueManager, report_excel_sync_error, report_database_error, report_service_down
        
        # Check for new issues and report them
        
        # 1. Check Excel sync errors
        if excel_sync_status.get('errors') and len(excel_sync_status['errors']) > 0:
            latest_error = excel_sync_status['errors'][-1]
            report_excel_sync_error(latest_error)
        
        # 2. Check if Excel sync service should be running but isn't
        if app.config.get('EXCEL_NETWORK_FOLDER') and not excel_sync_running:
            report_service_down("Excel Sync Service")
        
        # 3. Test database connectivity
        try:
            test_count = db.session.query(User).count()
            if test_count < 0:  # Invalid result
                report_database_error("Database query returned invalid results")
        except Exception as e:
            report_database_error(str(e))
        
        # Get current active issues
        system_issues = SystemIssueManager.get_active_issues_count()
        active_issues = SystemIssueManager.get_active_issues()
        
    except ImportError:
        # Fallback if system_issues module not available
        system_issues = 0
        active_issues = []
    
    # Get business workflow metrics (separate from system issues)
    pending_mismatches = MismatchRecord.query.filter_by(manager_approval=ApprovalStatus.PENDING).count()
    
    # Create system stats object for template
    system_stats = {
        'total_vendors': total_vendors,
        'total_managers': total_managers,
        'todays_submissions': total_statuses_today,
        'system_issues': system_issues,
        'active_issues': [issue.to_dict() for issue in active_issues] if 'active_issues' in locals() else [],
        'pending_mismatches': pending_mismatches  # Business metric, not system issue
    }
    
    return render_template('admin_dashboard.html',
                         system_stats=system_stats,
                         total_vendors=total_vendors,
                         total_managers=total_managers,
                         total_statuses_today=total_statuses_today,
                         upcoming_holidays=upcoming_holidays)

# Admin: Holidays management page
@app.route('/admin/holidays')
@login_required
def admin_holidays():
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    holidays = Holiday.query.order_by(Holiday.holiday_date.asc()).all()
    return render_template('admin_holidays.html', holidays=holidays)

# Admin: System Settings page (bulk import, templates)
@app.route('/admin/system-settings')
@login_required
def admin_system_settings():
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    return render_template('admin_system_settings.html')

# Admin: Audit logs page
@app.route('/admin/audit-logs')
@login_required
def admin_audit_logs():
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
    return render_template('admin_audit_logs.html', logs=logs)

# Shared: Holiday Calendar view (Vendor, Manager, Admin)
@app.route('/holiday-calendar')
@login_required
def holiday_calendar_view():
    """Render the shared holiday calendar page for all roles"""
    # Default to current month on load; client will fetch via API
    today_str = date.today().strftime('%Y-%m')
    return render_template('holiday_calendar.html', current_month=today_str)

# Shared: Holidays API
@app.route('/api/holidays')
@login_required
def api_list_holidays():
    """Return holidays as JSON. Optional start/end date (inclusive) in YYYY-MM-DD.
    If not provided, returns all holidays.
    """
    try:
        start_str = request.args.get('start')
        end_str = request.args.get('end')
        query = Holiday.query
        if start_str and end_str:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            query = query.filter(Holiday.holiday_date >= start_date, Holiday.holiday_date <= end_date)
        elif start_str:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            query = query.filter(Holiday.holiday_date >= start_date)
        elif end_str:
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            query = query.filter(Holiday.holiday_date <= end_date)
        holidays = query.order_by(Holiday.holiday_date.asc()).all()
        return jsonify({
            'status': 'success',
            'holidays': [
                {
                    'id': h.id,
                    'date': h.holiday_date.isoformat(),
                    'name': h.name,
                    'description': h.description or ''
                } for h in holidays
            ]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Admin: Teams & Groups
@app.route('/admin/teams')
@login_required
def admin_teams():
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    managers = Manager.query.all()
    # Build summary by manager and department
    team_data = []
    for m in managers:
        members = m.team_vendors.all() if m.team_vendors else []
        dept_counts = {}
        for v in members:
            dept_counts[v.department] = dept_counts.get(v.department, 0) + 1
        team_data.append({'manager': m, 'members': members, 'dept_counts': dept_counts})
    return render_template('admin_teams.html', team_data=team_data)

# Admin: Reconciliation report page
@app.route('/admin/reconciliation')
@login_required
def admin_reconciliation():
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    # Load latest mismatches (limit for performance)
    mismatches = MismatchRecord.query.order_by(MismatchRecord.mismatch_date.desc()).limit(500).all()
    total = len(mismatches)
    pending = len([m for m in mismatches if m.manager_approval == ApprovalStatus.PENDING])
    approved = len([m for m in mismatches if m.manager_approval == ApprovalStatus.APPROVED])
    rejected = len([m for m in mismatches if m.manager_approval == ApprovalStatus.REJECTED])
    summary = {
        'total': total,
        'pending': pending,
        'approved': approved,
        'rejected': rejected
    }
    return render_template('admin_reconciliation.html', mismatches=mismatches, summary=summary)

# Admin: Billing corrections (offsets)
@app.route('/admin/billing-corrections', methods=['GET', 'POST'])
@login_required
def admin_billing_corrections():
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    if request.method == 'POST':
        try:
            vendor_id = request.form['vendor_id']
            date_str = request.form['date']
            corrected_hours = float(request.form['corrected_hours'])
            reason = request.form.get('reason', '')
            vendor = Vendor.query.filter_by(vendor_id=vendor_id).first()
            if not vendor:
                flash('Vendor not found', 'error')
                return redirect(url_for('admin_billing_corrections'))
            
            # Try to get existing hours for this vendor on this date
            correction_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            existing_status = DailyStatus.query.filter_by(
                vendor_id=vendor.id, 
                status_date=correction_date
            ).first()
            
            # Capture old values for comparison
            old_values = {}
            if existing_status and existing_status.total_hours:
                old_values['previous_hours'] = existing_status.total_hours
                old_values['vendor_id'] = vendor_id
                old_values['date'] = date_str
                old_values['status'] = existing_status.status.value if existing_status.status else None
            else:
                # Check swipe records for historical data
                swipe_record = SwipeRecord.query.filter_by(
                    vendor_id=vendor.id,
                    attendance_date=correction_date
                ).first()
                if swipe_record and swipe_record.total_hours:
                    old_values['previous_hours'] = swipe_record.total_hours
                    old_values['source'] = 'swipe_record'
                old_values['vendor_id'] = vendor_id
                old_values['date'] = date_str
            
            # Log correction in audit log with proper old/new values comparison
            new_values = {
                'vendor_id': vendor_id,
                'date': date_str,
                'corrected_hours': corrected_hours,
                'reason': reason,
                'correction_type': 'manual_override'
            }
            
            create_audit_log(current_user.id,
                             'BILLING_CORRECTION',
                             'billing',
                             vendor.id,
                             old_values,
                             new_values)
            
            # Create informative success message
            if old_values.get('previous_hours'):
                success_msg = f'Billing correction recorded: {vendor_id} on {date_str} updated from {old_values["previous_hours"]} hrs to {corrected_hours} hrs'
            else:
                success_msg = f'Billing correction recorded: {vendor_id} on {date_str} set to {corrected_hours} hrs (new entry)'
            
            flash(success_msg, 'success')
        except Exception as e:
            flash(f'Error recording correction: {str(e)}', 'error')
    # Show last 50 corrections from audit logs
    corrections = AuditLog.query.filter(AuditLog.action == 'BILLING_CORRECTION').order_by(AuditLog.created_at.desc()).limit(50).all()
    return render_template('admin_billing_corrections.html', corrections=corrections)

# Admin: Vendors JSON for dashboard table
@app.route('/admin/vendors.json')
@login_required
def admin_vendors_json():
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    vendors = Vendor.query.join(User).order_by(Vendor.vendor_id.asc()).all()
    data = []
    for v in vendors:
        data.append({
            'vendor_id': v.vendor_id,
            'name': v.full_name,
            'email': v.user_account.email if v.user_account else '',
            'manager': v.manager_id or '-',
            'status': 'Active' if (v.user_account and v.user_account.is_active) else 'Inactive',
            'last_login': v.user_account.last_login.strftime('%Y-%m-%d %H:%M') if (v.user_account and v.user_account.last_login) else '-'
        })
    return jsonify({'vendors': data})

# Admin: Managers JSON for dropdowns
@app.route('/admin/managers.json')
@login_required
def admin_managers_json():
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    managers = Manager.query.join(User, Manager.user_id == User.id).order_by(Manager.full_name.asc()).all()
    items = []
    for m in managers:
        items.append({
            'manager_id': m.manager_id,
            'full_name': m.full_name,
            'email': m.user_account.email if m.user_account else '',
            'department': m.department,
            'team_name': m.team_name or ''
        })
    return jsonify({'managers': items})

# Admin: Add a vendor (JSON or form)
@app.route('/admin/vendor', methods=['POST'])
@login_required
def admin_add_vendor():
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    try:
        payload = request.get_json(silent=True) or request.form
        employee_id = (payload.get('vendor_id') or payload.get('employee_id') or '').strip()
        full_name = (payload.get('full_name') or payload.get('employee_name') or '').strip()
        email = (payload.get('email') or (f"{employee_id}@company.com" if employee_id else '')).strip()
        department = (payload.get('department') or '').strip()
        company = (payload.get('company') or payload.get('business_unit') or '').strip()
        band = (payload.get('band') or 'B1').strip()
        location = (payload.get('location') or payload.get('floor_unit') or '').strip()
        manager_id = (payload.get('manager_id') or '').strip() or None
        password = (payload.get('password') or 'vendor123')

        # Basic validation
        if not employee_id or not full_name:
            return jsonify({'error': 'vendor_id and full_name are required'}), 400

        # Uniqueness checks
        if Vendor.query.filter_by(vendor_id=employee_id).first() or User.query.filter_by(username=employee_id).first():
            return jsonify({'error': 'Vendor ID already exists'}), 400
        if email and User.query.filter_by(email=email).first():
            return jsonify({'error': 'Email already in use'}), 400

        # Require existing manager if provided
        if manager_id:
            mgr = Manager.query.filter_by(manager_id=manager_id).first()
            if not mgr:
                return jsonify({'error': 'Selected manager does not exist'}), 400

        # Create user
        user = User(username=employee_id, email=email or f"{employee_id}@company.com", role=UserRole.VENDOR, is_active=True)
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        # Create vendor
        vendor = Vendor(
            user_id=user.id,
            vendor_id=employee_id,
            full_name=full_name,
            department=department,
            company=company,
            band=band,
            location=location,
            manager_id=manager_id
        )
        db.session.add(vendor)
        db.session.commit()
        create_audit_log(current_user.id, 'CREATE', 'vendors', vendor.id, {}, {
            'vendor_id': employee_id,
            'full_name': full_name,
            'manager_id': manager_id
        })
        return jsonify({'success': True, 'message': 'Vendor created'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Admin: Add a manager (JSON or form)
@app.route('/admin/manager', methods=['POST'])
@login_required
def admin_add_manager():
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    try:
        payload = request.get_json(silent=True) or request.form
        manager_id = (payload.get('manager_id') or '').strip()
        full_name = (payload.get('full_name') or '').strip()
        email = (payload.get('email') or (f"{manager_id}@company.com" if manager_id else '')).strip()
        department = (payload.get('department') or '').strip()
        team_name = (payload.get('team_name') or '').strip()
        phone = (payload.get('phone') or '').strip()
        password = (payload.get('password') or 'manager123')

        if not manager_id or not full_name:
            return jsonify({'error': 'manager_id and full_name are required'}), 400

        # Uniqueness checks
        if Manager.query.filter_by(manager_id=manager_id).first() or User.query.filter_by(username=manager_id).first():
            return jsonify({'error': 'Manager ID already exists'}), 400
        if email and User.query.filter_by(email=email).first():
            return jsonify({'error': 'Email already in use'}), 400

        # Create manager user
        user = User(username=manager_id, email=email or f"{manager_id}@company.com", role=UserRole.MANAGER, is_active=True)
        user.set_password(password)
        db.session.add(user)
        db.session.flush()

        # Create manager profile
        manager = Manager(
            manager_id=manager_id,
            user_id=user.id,
            full_name=full_name,
            department=department,
            team_name=team_name,
            email=email or None,
            phone=phone or None
        )
        db.session.add(manager)
        db.session.commit()

        create_audit_log(current_user.id, 'CREATE', 'managers', manager.id, {}, {
            'manager_id': manager_id,
            'full_name': full_name,
            'department': department
        })
        return jsonify({'success': True, 'message': 'Manager created'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Admin: Update vendor
@app.route('/admin/vendor/<vendor_id>/update', methods=['POST'])
@login_required
def admin_update_vendor(vendor_id):
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    try:
        v = Vendor.query.filter_by(vendor_id=vendor_id).first_or_404()
        payload = request.get_json(silent=True) or request.form
        old = {'full_name': v.full_name, 'department': v.department, 'company': v.company, 'band': v.band, 'location': v.location, 'manager_id': v.manager_id}
        v.full_name = payload.get('full_name', v.full_name)
        v.department = payload.get('department', v.department)
        v.company = payload.get('company', v.company)
        v.band = payload.get('band', v.band)
        v.location = payload.get('location', v.location)
        v.manager_id = payload.get('manager_id', v.manager_id)
        if v.user_account:
            v.user_account.email = payload.get('email', v.user_account.email)
        db.session.commit()
        create_audit_log(current_user.id, 'UPDATE', 'vendors', v.id, old, {'full_name': v.full_name})
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Admin: Activate/Deactivate vendor
@app.route('/admin/vendor/<vendor_id>/activate', methods=['POST'])
@login_required
def admin_activate_vendor(vendor_id):
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    v = Vendor.query.filter_by(vendor_id=vendor_id).first_or_404()
    if v.user_account:
        v.user_account.is_active = True
        db.session.commit()
        create_audit_log(current_user.id, 'UPDATE', 'users', v.user_account.id, {'is_active': False}, {'is_active': True})
    return jsonify({'success': True})

@app.route('/admin/vendor/<vendor_id>/deactivate', methods=['POST'])
@login_required
def admin_deactivate_vendor(vendor_id):
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    v = Vendor.query.filter_by(vendor_id=vendor_id).first_or_404()
    if v.user_account:
        v.user_account.is_active = False
        db.session.commit()
        create_audit_log(current_user.id, 'UPDATE', 'users', v.user_account.id, {'is_active': True}, {'is_active': False})
    return jsonify({'success': True})

# Admin: System settings API (notification toggles, weekends)
@app.route('/admin/system-settings/config', methods=['GET', 'POST'])
@login_required
def admin_system_settings_config():
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    if request.method == 'GET':
        return jsonify({
            'dailyReminders': get_system_config('dailyReminders', 'true') == 'true',
            'managerAlerts': get_system_config('managerAlerts', 'true') == 'true',
            'mismatchAlerts': get_system_config('mismatchAlerts', 'true') == 'true',
            'weekend_days': get_system_config('weekend_days', '[5,6]')
        })
    # POST save
    daily = 'true' if (request.form.get('dailyReminders') in ['true', 'on', '1']) else 'false'
    mgr = 'true' if (request.form.get('managerAlerts') in ['true', 'on', '1']) else 'false'
    mm = 'true' if (request.form.get('mismatchAlerts') in ['true', 'on', '1']) else 'false'
    set_system_config('dailyReminders', daily, 'Daily reminder notifications', current_user.id)
    set_system_config('managerAlerts', mgr, 'Manager alerts', current_user.id)
    set_system_config('mismatchAlerts', mm, 'Mismatch alerts', current_user.id)
    if request.form.get('weekend_days'):
        set_system_config('weekend_days', request.form.get('weekend_days'), 'Weekend days indices', current_user.id)
    return jsonify({'success': True})

@app.route('/manager/dashboard')
@login_required
def manager_dashboard():
    """Manager dashboard with real team data"""
    if current_user.role != UserRole.MANAGER:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    manager = current_user.manager_profile
    if not manager:
        flash('Manager profile not found', 'error')
        return redirect(url_for('login'))
    
    # Get team statistics
    today = date.today()
    team_vendors = manager.team_vendors.all() if manager.team_vendors else []
    vendor_ids = [v.id for v in team_vendors]
    
    # Get today's statuses for team
    today_statuses = []
    present_count = 0
    pending_approvals = 0
    mismatches_count = 0
    
    for vendor in team_vendors:
        # Get today's status
        today_status = DailyStatus.query.filter_by(
            vendor_id=vendor.id, 
            status_date=today
        ).first()
        
        if today_status:
            if today_status.status in [AttendanceStatus.IN_OFFICE_FULL, AttendanceStatus.IN_OFFICE_HALF, 
                                     AttendanceStatus.WFH_FULL, AttendanceStatus.WFH_HALF]:
                present_count += 1
            if today_status.approval_status == ApprovalStatus.PENDING:
                pending_approvals += 1
                
        # Get pending mismatches for this vendor
        vendor_mismatches = MismatchRecord.query.filter_by(
            vendor_id=vendor.id,
            manager_approval=ApprovalStatus.PENDING
        ).count()
        mismatches_count += vendor_mismatches
        
        # Add to today's status list
        today_statuses.append({
            'vendor': vendor,
            'status': today_status,
            'has_mismatch': vendor_mismatches > 0
        })
    
    # Pending approvals for the team (recent)
    pending_statuses = []
    if vendor_ids:
        pending_statuses = DailyStatus.query.filter(
            DailyStatus.vendor_id.in_(vendor_ids),
            DailyStatus.approval_status == ApprovalStatus.PENDING
        ).order_by(DailyStatus.status_date.desc(), DailyStatus.submitted_at.desc()).limit(50).all()
    
    team_stats = {
        'total_members': len(team_vendors),
        'present_today': present_count,
        'pending_approvals': pending_approvals,
        'mismatches': mismatches_count
    }
    
    return render_template('manager_dashboard.html', 
                         manager=manager,
                         team_stats=team_stats,
                         today_statuses=today_statuses,
                         pending_statuses=pending_statuses,
                         today_date=today.strftime('%B %d, %Y'))


@app.route('/manager/ai-insights')
@login_required
def manager_ai_insights():
    """AI Insights page backed by heuristic predictions."""
    if current_user.role != UserRole.MANAGER:
        flash('Access denied', 'error')
        return redirect(url_for('index'))

    manager = current_user.manager_profile
    if not manager:
        flash('Manager profile not found', 'error')
        return redirect(url_for('login'))

    try:
        # Optional query param to adjust window (days)
        window = request.args.get('window', default='7')
        try:
            prediction_window = max(3, min(30, int(window)))
        except Exception:
            prediction_window = 7

        predictions, ai_stats, risk_distribution = generate_ai_insights(manager.id, prediction_window_days=prediction_window)

        return render_template(
            'ai_insights.html',
            ai_stats=ai_stats,
            predictions=predictions,
            risk_distribution=risk_distribution
        )
    except Exception as e:
        flash(f'Error loading AI insights: {str(e)}', 'error')
        # Fallback to template with no data (will show demo placeholders)
        return render_template('ai_insights.html', ai_stats={}, predictions=[], risk_distribution={})

# AI quick action APIs (report, schedule, logs, override)
@app.route('/api/ai/report')
@login_required
def api_ai_report():
    """Generate an AI insights report for the manager's team (Excel or JSON)."""
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied'}), 403
    manager = current_user.manager_profile
    if not manager:
        return jsonify({'error': 'Manager profile not found'}), 404

    window = request.args.get('window', default='7')
    fmt = request.args.get('format', default='excel')
    try:
        prediction_window = max(3, min(30, int(window)))
    except Exception:
        prediction_window = 7

    predictions, ai_stats, _ = generate_ai_insights(manager.id, prediction_window_days=prediction_window)

    # Prepare tabular rows
    rows = []
    for p in predictions:
        rows.append({
            'Vendor Name': p['vendor_name'],
            'Vendor ID': p['vendor_id'],
            'Predicted Date': p['predicted_date'],
            'Likelihood %': p['likelihood'],
            'Risk Level': p['risk_level'],
            'Recommendation': p['recommendation'],
            'Reasons': '; '.join(p['reasons'])
        })

    if fmt == 'excel' and rows:
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        df = pd.DataFrame(rows)
        meta = pd.DataFrame([
            {'Metric': 'Absence Predictions', 'Value': ai_stats.get('absence_predictions')},
            {'Metric': 'WFH Predictions', 'Value': ai_stats.get('wfh_predictions')},
            {'Metric': 'Risk Alerts', 'Value': ai_stats.get('risk_alerts')},
            {'Metric': 'Predictions Made', 'Value': ai_stats.get('predictions_made')},
            {'Metric': 'Last Trained', 'Value': ai_stats.get('last_trained')},
        ])
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            meta.to_excel(writer, sheet_name='Summary', index=False)
            df.to_excel(writer, sheet_name='Predictions', index=False)
        buf.seek(0)
        filename = f"ai_report_{date.today().isoformat()}_w{prediction_window}.xlsx"
        try:
            create_audit_log(current_user.id, 'EXPORT', 'ai_report', None, None, {'window': prediction_window})
        except Exception:
            pass
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Default JSON
    return jsonify({'status': 'success', 'ai_stats': ai_stats, 'predictions': rows})

@app.route('/api/ai/schedule', methods=['POST'])
@login_required
def api_ai_schedule():
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied'}), 403
    schedule = request.json.get('schedule', 'daily') if request.is_json else request.form.get('schedule', 'daily')
    try:
        set_system_config('ai_schedule', schedule, 'AI analysis schedule', current_user.id)
        create_audit_log(current_user.id, 'UPDATE', 'ai_schedule', None, None, {'schedule': schedule})
        return jsonify({'status': 'success', 'message': f'Scheduled AI analysis: {schedule}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/ai/model-logs')
@login_required
def api_ai_model_logs():
    if current_user.role not in [UserRole.MANAGER, UserRole.ADMIN]:
        return jsonify({'error': 'Access denied'}), 403
    # Return recent audit logs related to AI or exports as a simple log source
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(20).all()
    data = []
    for log in logs:
        data.append({
            'timestamp': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'action': log.action,
            'table': log.table_name,
            'user_id': log.user_id
        })
    return jsonify({'status': 'success', 'logs': data})

@app.route('/api/manager/vendor-summary/<string:vendor_id>')
@login_required
def get_vendor_summary(vendor_id):
    """Get detailed summary for a specific vendor"""
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied'}), 403
    
    manager = current_user.manager_profile
    if not manager:
        return jsonify({'error': 'Manager profile not found'}), 404
    
    try:
        # Find the vendor
        vendor = Vendor.query.filter_by(vendor_id=vendor_id).first()
        if not vendor:
            return jsonify({'error': 'Vendor not found'}), 404
        
        # Verify vendor is in manager's team
        if vendor.manager_id != manager.manager_id:
            return jsonify({'error': 'You can only view summary for your team members'}), 403
        
        # Get date range (last 30 days)
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
        
        # Get all statuses in date range
        statuses = DailyStatus.query.filter(
            DailyStatus.vendor_id == vendor.id,
            DailyStatus.status_date >= start_date,
            DailyStatus.status_date <= end_date
        ).order_by(DailyStatus.status_date.desc()).all()
        
        # Calculate statistics
        stats = {
            'in_office_full': 0,
            'in_office_half': 0,
            'wfh_full': 0,
            'wfh_half': 0,
            'leave_full': 0,
            'leave_half': 0,
            'total_submitted': len(statuses),
            'pending_approval': 0,
            'approved': 0,
            'rejected': 0,
            'total_working_hours': 0.0,
            'total_extra_hours': 0.0
        }
        
        recent_activities = []
        
        for status in statuses:
            # Count by status type
            if status.status == AttendanceStatus.IN_OFFICE_FULL:
                stats['in_office_full'] += 1
            elif status.status == AttendanceStatus.IN_OFFICE_HALF:
                stats['in_office_half'] += 1
            elif status.status == AttendanceStatus.WFH_FULL:
                stats['wfh_full'] += 1
            elif status.status == AttendanceStatus.WFH_HALF:
                stats['wfh_half'] += 1
            elif status.status == AttendanceStatus.LEAVE_FULL:
                stats['leave_full'] += 1
            elif status.status == AttendanceStatus.LEAVE_HALF:
                stats['leave_half'] += 1
            
            # Count by approval status
            if status.approval_status == ApprovalStatus.PENDING:
                stats['pending_approval'] += 1
            elif status.approval_status == ApprovalStatus.APPROVED:
                stats['approved'] += 1
            elif status.approval_status == ApprovalStatus.REJECTED:
                stats['rejected'] += 1
            
            # Add working hours if available
            if status.total_hours:
                stats['total_working_hours'] += status.total_hours
            if status.extra_hours:
                stats['total_extra_hours'] += status.extra_hours
            
            # Add to recent activities
            recent_activities.append({
                'date': status.status_date.strftime('%Y-%m-%d'),
                'status': status.status.value.replace('_', ' ').title(),
                'location': status.location or '-',
                'approval': status.approval_status.value.title(),
                'submitted_at': status.submitted_at.strftime('%Y-%m-%d %H:%M') if status.submitted_at else '-'
            })
        
        # Calculate working days in period
        from utils import is_non_working_day
        working_days = 0
        current_date = start_date
        while current_date <= end_date:
            if not is_non_working_day(current_date):
                working_days += 1
            current_date += timedelta(days=1)
        
        # Get pending mismatches
        pending_mismatches = MismatchRecord.query.filter_by(
            vendor_id=vendor.id,
            manager_approval=ApprovalStatus.PENDING
        ).count()
        
        # Calculate rates
        stats['total_office'] = stats['in_office_full'] + stats['in_office_half']
        stats['total_wfh'] = stats['wfh_full'] + stats['wfh_half']
        stats['total_leave'] = stats['leave_full'] + stats['leave_half']
        stats['submission_rate'] = round((stats['total_submitted'] / working_days * 100), 1) if working_days > 0 else 0
        stats['office_rate'] = round((stats['total_office'] / stats['total_submitted'] * 100), 1) if stats['total_submitted'] > 0 else 0
        stats['wfh_rate'] = round((stats['total_wfh'] / stats['total_submitted'] * 100), 1) if stats['total_submitted'] > 0 else 0
        stats['working_days'] = working_days
        stats['pending_mismatches'] = pending_mismatches
        
        # Calculate working hours averages
        stats['total_working_hours'] = round(stats['total_working_hours'], 1)
        stats['total_extra_hours'] = round(stats['total_extra_hours'], 1)
        stats['average_hours_per_day'] = round(stats['total_working_hours'] / max(1, stats['total_submitted']), 1) if stats['total_submitted'] > 0 else 0
        
        vendor_summary = {
            'vendor': {
                'id': vendor.vendor_id,
                'name': vendor.full_name,
                'email': vendor.user_account.email if vendor.user_account else 'N/A',
                'department': vendor.department,
                'company': vendor.company,
                'band': vendor.band,
                'location': vendor.location,
                'created_at': vendor.created_at.strftime('%Y-%m-%d') if vendor.created_at else 'N/A'
            },
            'period': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'working_days': working_days
            },
            'statistics': stats,
            'recent_activities': recent_activities[:10]  # Last 10 activities
        }
        
        return jsonify({
            'status': 'success',
            'data': vendor_summary
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/ai/override', methods=['POST'])
@login_required
def api_ai_override():
    if current_user.role not in [UserRole.MANAGER, UserRole.ADMIN]:
        return jsonify({'error': 'Access denied'}), 403
    enable = None
    if request.is_json:
        enable = request.json.get('enable')
    if enable is None:
        enable = request.form.get('enable')
    # default: toggle off
    enabled = False if (enable is None) else (str(enable).lower() in ['true', '1', 'yes'])
    try:
        set_system_config('ai_enabled', 'true' if enabled else 'false', 'Emergency override toggle', current_user.id)
        create_audit_log(current_user.id, 'UPDATE', 'ai_enabled', None, None, {'enabled': enabled})
        return jsonify({'status': 'success', 'enabled': enabled})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============= API ROUTES FOR SWAGGER =============

@app.route('/api/dashboard/stats')
@login_required
def api_dashboard_stats():
    """Get dashboard statistics based on user role"""
    if current_user.role == UserRole.ADMIN:
        stats = {
            'total_users': User.query.count(),
            'total_vendors': Vendor.query.count(),
            'total_managers': Manager.query.count(),
            'active_today': DailyStatus.query.filter_by(status_date=date.today()).count(),
            'system_health': 'Excellent',
            'notifications': 5
        }
    elif current_user.role == UserRole.MANAGER:
        manager = current_user.manager_profile
        team_vendors = manager.team_vendors.all() if manager and manager.team_vendors else []
        stats = {
            'team_size': len(team_vendors),
            'pending_approvals': DailyStatus.query.filter(
                DailyStatus.vendor_id.in_([v.id for v in team_vendors]),
                DailyStatus.approval_status == ApprovalStatus.PENDING
            ).count() if team_vendors else 0,
            'approved_today': DailyStatus.query.filter(
                DailyStatus.vendor_id.in_([v.id for v in team_vendors]),
                DailyStatus.status_date == date.today(),
                DailyStatus.approval_status == ApprovalStatus.APPROVED
            ).count() if team_vendors else 0
        }
    else:  # Vendor
        vendor = current_user.vendor_profile
        
        # Calculate proper attendance rate for current month
        attendance_rate = 0
        if vendor:
            today = date.today()
            start_of_month = today.replace(day=1)
            
            # Get statuses for current month
            month_statuses = DailyStatus.query.filter(
                DailyStatus.vendor_id == vendor.id,
                DailyStatus.status_date >= start_of_month,
                DailyStatus.status_date <= today
            ).all()
            
            # Calculate working days in current month up to today
            working_days_this_month = 0
            current_date = start_of_month
            while current_date <= today:
                if current_date.weekday() < 5:  # Monday to Friday
                    is_holiday = Holiday.query.filter_by(holiday_date=current_date).first()
                    if not is_holiday:
                        working_days_this_month += 1
                current_date += timedelta(days=1)
            
            # Calculate attendance with half-day weighting
            working_day_submissions = 0
            for status in month_statuses:
                if status.status_date.weekday() < 5:  # Monday to Friday
                    is_holiday = Holiday.query.filter_by(holiday_date=status.status_date).first()
                    if not is_holiday:
                        if status.status in [AttendanceStatus.IN_OFFICE_HALF, AttendanceStatus.WFH_HALF, AttendanceStatus.LEAVE_HALF]:
                            working_day_submissions += 0.5
                        else:
                            working_day_submissions += 1
            
            attendance_rate = min(100.0, (working_day_submissions / working_days_this_month * 100)) if working_days_this_month > 0 else 0
        
        stats = {
            'total_statuses': DailyStatus.query.filter_by(vendor_id=vendor.id).count() if vendor else 0,
            'pending_mismatches': MismatchRecord.query.filter_by(
                vendor_id=vendor.id, 
                manager_approval=ApprovalStatus.PENDING
            ).count() if vendor else 0,
            'attendance_rate': round(attendance_rate, 1),
            'current_month_days': DailyStatus.query.filter(
                DailyStatus.vendor_id == vendor.id,
                DailyStatus.status_date >= date.today().replace(day=1)
            ).count() if vendor else 0
        }
    
    return jsonify(stats)

@app.route('/api/charts/attendance-trends')
@login_required
def api_attendance_trends():
    """Get attendance trend data for charts"""
    # Mock data for demo - replace with actual database queries
    data = {
        'labels': ['Week 1', 'Week 2', 'Week 3', 'Week 4'],
        'office_attendance': [85, 78, 92, 88],
        'wfh_attendance': [12, 18, 5, 8],
        'leave_requests': [3, 4, 3, 4]
    }
    return jsonify(data)

@app.route('/api/notifications')
@login_required
def api_notifications():
    """Get user notifications"""
    notifications = [
        {
            'id': 1,
            'title': 'Status Approval Required',
            'message': '3 vendor statuses pending your approval',
            'type': 'warning',
            'timestamp': datetime.now().isoformat()
        },
        {
            'id': 2,
            'title': 'Monthly Report Ready',
            'message': 'January 2025 attendance report is ready for download',
            'type': 'info',
            'timestamp': datetime.now().isoformat()
        }
    ]
    return jsonify(notifications)

@app.route('/api/test/notification/<notification_type>')
@login_required
def test_notification(notification_type):
    """Test notification sending (for managers only)"""
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied - managers only'}), 403
    
    try:
        if notification_type == 'summary':
            notification_service.send_daily_summary_notifications('manual')
            message = 'Daily summary notifications sent successfully'
        elif notification_type == 'urgent':
            notification_service.send_urgent_reminder_notifications()
            message = 'Urgent reminder notifications sent successfully'
        else:
            return jsonify({'error': 'Invalid notification type. Use "summary" or "urgent"'}), 400
        
        return jsonify({
            'status': 'success',
            'message': message,
            'notification_type': notification_type
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Failed to send notifications: {str(e)}'
        }), 500

@app.route('/api/export/monthly-report')
@login_required
def api_export_monthly_report():
    """Export monthly attendance report"""
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    format_type = request.args.get('format', 'excel')
    
    try:
        if current_user.role == UserRole.VENDOR:
            # For vendors, generate personal report
            vendor = current_user.vendor_profile
            if not vendor:
                return jsonify({'error': 'Vendor profile not found'}), 404
            
            # Parse month string
            year, month_num = map(int, month.split('-'))
            start_date = date(year, month_num, 1)
            if month_num == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month_num + 1, 1) - timedelta(days=1)
            
            # Get vendor's statuses for the month
            statuses = DailyStatus.query.filter(
                DailyStatus.vendor_id == vendor.id,
                DailyStatus.status_date >= start_date,
                DailyStatus.status_date <= end_date
            ).order_by(DailyStatus.status_date.asc()).all()
            
            # Create report data
            report_data = []
            for status in statuses:
                report_data.append({
                    'Date': status.status_date.strftime('%Y-%m-%d'),
                    'Weekday': status.status_date.strftime('%A'),
                    'Status': status.status.value.replace('_', ' ').title(),
                    'Location': status.location or '-',
                    'Comments': status.comments or '-',
                    'Approval Status': status.approval_status.value.title(),
                    'Submitted At': status.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if status.submitted_at else '-'
                })
            
            if format_type == 'excel' and report_data:
                import pandas as pd
                from io import BytesIO
                from flask import send_file
                
                # Create Excel file in memory
                df = pd.DataFrame(report_data)
                excel_buffer = BytesIO()
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=f'{vendor.full_name} - {month}', index=False)
                    
                    # Get workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets[f'{vendor.full_name} - {month}']
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                excel_buffer.seek(0)
                
                filename = f'{vendor.full_name}_{month}_attendance_report.xlsx'
                return send_file(
                    excel_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            elif not report_data:
                return jsonify({
                    'status': 'error',
                    'message': f'No attendance data found for {month}'
                }), 404
            
            else:
                # Return JSON format
                return jsonify({
                    'status': 'success',
                    'data': report_data,
                    'vendor_name': vendor.full_name,
                    'month': month,
                    'total_records': len(report_data)
                })
        
        elif current_user.role == UserRole.MANAGER:
            # For managers, use existing team report logic
            manager = current_user.manager_profile
            if not manager:
                return jsonify({'error': 'Manager profile not found'}), 404
            
            report_data = generate_monthly_report(manager.id, month)
            
            if format_type == 'excel' and report_data:
                import pandas as pd
                from io import BytesIO
                from flask import send_file
                
                df = pd.DataFrame(report_data)
                excel_buffer = BytesIO()
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=f'Team Report - {month}', index=False)
                
                excel_buffer.seek(0)
                filename = f'team_report_{month}.xlsx'
                return send_file(
                    excel_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            return jsonify({
                'status': 'success',
                'data': report_data,
                'month': month
            })
        
        else:  # Admin
            report_data = generate_monthly_report(None, month)  # All vendors
            
            if format_type == 'excel':
                if not report_data:
                    return jsonify({
                        'status': 'error',
                        'message': f'No attendance data found for {month}'
                    }), 404
                    
                # Generate Excel file
                import pandas as pd
                from io import BytesIO
                from flask import send_file
                
                df = pd.DataFrame(report_data)
                excel_buffer = BytesIO()
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=f'System Report - {month}', index=False)
                    
                    # Get workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets[f'System Report - {month}']
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                excel_buffer.seek(0)
                
                filename = f'system_monthly_report_{month}.xlsx'
                return send_file(
                    excel_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            elif not report_data:
                return jsonify({
                    'status': 'error',
                    'message': f'No attendance data found for {month}'
                }), 404
            
            else:
                # Return JSON format
                return jsonify({
                    'status': 'success',
                    'data': report_data,
                    'month': month
                })
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error generating report: {str(e)}'
        }), 500

@app.route('/monthly-report')
@login_required
def monthly_report_view():
    """Display monthly report with charts and graphs"""
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    if current_user.role == UserRole.VENDOR:
        vendor = current_user.vendor_profile
        if not vendor:
            flash('Vendor profile not found', 'error')
            return redirect(url_for('index'))
        
        return render_template('monthly_report.html', 
                             user_type='vendor',
                             profile=vendor,
                             selected_month=month)
    
    elif current_user.role == UserRole.MANAGER:
        manager = current_user.manager_profile
        if not manager:
            flash('Manager profile not found', 'error')
            return redirect(url_for('index'))
        
        return render_template('monthly_report.html',
                             user_type='manager', 
                             profile=manager,
                             selected_month=month)
    
    else:  # Admin
        return render_template('monthly_report.html',
                             user_type='admin',
                             profile=current_user,
                             selected_month=month)

@app.route('/api/monthly-report-data')
@login_required
def api_monthly_report_data():
    """API endpoint for monthly report data (for charts)"""
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    
    try:
        # Common month parsing
        year, month_num = map(int, month.split('-'))
        start_date = date(year, month_num, 1)
        end_date = (date(year + 1, 1, 1) - timedelta(days=1)) if month_num == 12 else (date(year, month_num + 1, 1) - timedelta(days=1))

        if current_user.role == UserRole.VENDOR:
            vendor = current_user.vendor_profile
            if not vendor:
                return jsonify({'error': 'Vendor profile not found'}), 404
            
            # Get vendor's statuses for the month
            statuses = DailyStatus.query.filter(
                DailyStatus.vendor_id == vendor.id,
                DailyStatus.status_date >= start_date,
                DailyStatus.status_date <= end_date
            ).order_by(DailyStatus.status_date.asc()).all()
            
            # Process data for charts
            status_counts = {}
            daily_data = []
            approval_counts = {'pending': 0, 'approved': 0, 'rejected': 0}
            
            for status in statuses:
                # Count status types
                status_type = status.status.value
                status_counts[status_type] = status_counts.get(status_type, 0) + 1
                
                # Count approval status
                approval_status = status.approval_status.value
                approval_counts[approval_status] += 1
                
                # Daily data for timeline
                daily_data.append({
                    'date': status.status_date.strftime('%Y-%m-%d'),
                    'status': status_type,
                    'location': status.location or 'Not specified'
                })
            
            # Calculate working days in month
            total_working_days = 0
            current_date = start_date
            while current_date <= end_date:
                if current_date.weekday() < 5:  # Monday = 0, Friday = 4
                    is_holiday = Holiday.query.filter_by(holiday_date=current_date).first()
                    if not is_holiday:
                        total_working_days += 1
                current_date += timedelta(days=1)
            
            # Calculate attendance rate properly (should not exceed 100%)
            # Count half-days as 0.5 and only count working days
            working_day_submissions = 0
            for status in statuses:
                # Only count submissions on actual working days
                if status.status_date.weekday() < 5:  # Monday to Friday
                    is_holiday = Holiday.query.filter_by(holiday_date=status.status_date).first()
                    if not is_holiday:
                        # Count half-days as 0.5, full days as 1
                        if status.status in [AttendanceStatus.IN_OFFICE_HALF, AttendanceStatus.WFH_HALF, AttendanceStatus.LEAVE_HALF]:
                            working_day_submissions += 0.5
                        else:
                            working_day_submissions += 1
            
            # Cap attendance rate at 100% to prevent over 100% values
            attendance_rate = min(100.0, (working_day_submissions / total_working_days * 100)) if total_working_days > 0 else 0
            submitted_days = len(statuses)
            
            return jsonify({
                'success': True,
                'user_type': 'vendor',
                'vendor_name': vendor.full_name,
                'vendor_id': vendor.vendor_id,
                'month': month,
                'status_breakdown': status_counts,
                'approval_breakdown': approval_counts,
                'daily_data': daily_data,
                'summary': {
                    'total_working_days': total_working_days,
                    'submitted_days': submitted_days,
                    'attendance_rate': round(attendance_rate, 1),
                    'pending_approvals': approval_counts['pending']
                }
            })
        
        elif current_user.role == UserRole.MANAGER:
            manager = current_user.manager_profile
            if not manager:
                return jsonify({'error': 'Manager profile not found'}), 404
            
            team_vendors = manager.team_vendors.all() if manager.team_vendors else []
            vendor_ids = [v.id for v in team_vendors]
            team_size = len(vendor_ids)
            
            # If no team members
            if team_size == 0:
                return jsonify({
                    'success': True,
                    'user_type': 'manager',
                    'month': month,
                    'status_breakdown': {},
                    'approval_breakdown': {'pending': 0, 'approved': 0, 'rejected': 0},
                    'daily_data': [],
                    'summary': {
                        'total_working_days': 0,
                        'submitted_days': 0,
                        'attendance_rate': 0,
                        'pending_approvals': 0
                    }
                })
            
            # Get statuses for team
            statuses = DailyStatus.query.filter(
                DailyStatus.vendor_id.in_(vendor_ids),
                DailyStatus.status_date >= start_date,
                DailyStatus.status_date <= end_date
            ).all()
            
            status_counts = {}
            approval_counts = {'pending': 0, 'approved': 0, 'rejected': 0}
            from collections import defaultdict
            statuses_by_date = defaultdict(list)
            for s in statuses:
                status_counts[s.status.value] = status_counts.get(s.status.value, 0) + 1
                approval_counts[s.approval_status.value] += 1
                statuses_by_date[s.status_date].append(s)
            
            # Working days
            total_working_days = 0
            submitted_working_statuses = 0
            daily_data = []
            d = start_date
            while d <= end_date:
                is_weekend = d.weekday() >= 5
                is_holiday = Holiday.query.filter_by(holiday_date=d).first() is not None
                if not is_weekend and not is_holiday:
                    total_working_days += 1
                    day_statuses = statuses_by_date.get(d, [])
                    # Count with half-day weighting for proper attendance rate
                    for ds in day_statuses:
                        if ds.status in [AttendanceStatus.IN_OFFICE_HALF, AttendanceStatus.WFH_HALF, AttendanceStatus.LEAVE_HALF]:
                            submitted_working_statuses += 0.5
                        else:
                            submitted_working_statuses += 1
                    # Aggregate to categories
                    counts = {'in_office_full': 0, 'wfh_full': 0, 'leave_full': 0, 'absent': 0}
                    present_vendor_ids = set()
                    for ds in day_statuses:
                        present_vendor_ids.add(ds.vendor_id)
                        if ds.status in [AttendanceStatus.IN_OFFICE_FULL, AttendanceStatus.IN_OFFICE_HALF]:
                            counts['in_office_full'] += 1
                        elif ds.status in [AttendanceStatus.WFH_FULL, AttendanceStatus.WFH_HALF]:
                            counts['wfh_full'] += 1
                        elif ds.status in [AttendanceStatus.LEAVE_FULL, AttendanceStatus.LEAVE_HALF]:
                            counts['leave_full'] += 1
                        elif ds.status == AttendanceStatus.ABSENT:
                            counts['absent'] += 1
                    # Count implicit absences (no submission)
                    counts['absent'] += max(0, team_size - len(present_vendor_ids))
                    # Choose dominant category with tie-breaker order
                    order = ['in_office_full', 'wfh_full', 'leave_full', 'absent']
                    dominant = max(order, key=lambda k: (counts[k], -order.index(k)))
                    daily_data.append({'date': d.strftime('%Y-%m-%d'), 'status': dominant, 'location': 'Various'})
                d += timedelta(days=1)
            
            potential_entries = total_working_days * team_size
            attendance_rate = min(100.0, (submitted_working_statuses / potential_entries * 100)) if potential_entries > 0 else 0
            
            return jsonify({
                'success': True,
                'user_type': 'manager',
                'month': month,
                'status_breakdown': status_counts,
                'approval_breakdown': approval_counts,
                'daily_data': daily_data,
                'summary': {
                    'total_working_days': total_working_days,
                    'submitted_days': submitted_working_statuses,
                    'attendance_rate': round(attendance_rate, 1),
                    'pending_approvals': approval_counts['pending']
                }
            })
        
        else:  # Admin
            # All vendors
            all_vendors = Vendor.query.all()
            vendor_ids = [v.id for v in all_vendors]
            total_vendors = len(vendor_ids)
            if total_vendors == 0:
                return jsonify({
                    'success': True,
                    'user_type': 'admin',
                    'month': month,
                    'status_breakdown': {},
                    'approval_breakdown': {'pending': 0, 'approved': 0, 'rejected': 0},
                    'daily_data': [],
                    'summary': {
                        'total_working_days': 0,
                        'submitted_days': 0,
                        'attendance_rate': 0,
                        'pending_approvals': 0
                    }
                })
            
            statuses = DailyStatus.query.filter(
                DailyStatus.vendor_id.in_(vendor_ids),
                DailyStatus.status_date >= start_date,
                DailyStatus.status_date <= end_date
            ).all()
            
            status_counts = {}
            approval_counts = {'pending': 0, 'approved': 0, 'rejected': 0}
            from collections import defaultdict
            statuses_by_date = defaultdict(list)
            for s in statuses:
                status_counts[s.status.value] = status_counts.get(s.status.value, 0) + 1
                approval_counts[s.approval_status.value] += 1
                statuses_by_date[s.status_date].append(s)
            
            total_working_days = 0
            submitted_working_statuses = 0
            daily_data = []
            d = start_date
            while d <= end_date:
                is_weekend = d.weekday() >= 5
                is_holiday = Holiday.query.filter_by(holiday_date=d).first() is not None
                if not is_weekend and not is_holiday:
                    total_working_days += 1
                    day_statuses = statuses_by_date.get(d, [])
                    # Count with half-day weighting for proper attendance rate
                    for ds in day_statuses:
                        if ds.status in [AttendanceStatus.IN_OFFICE_HALF, AttendanceStatus.WFH_HALF, AttendanceStatus.LEAVE_HALF]:
                            submitted_working_statuses += 0.5
                        else:
                            submitted_working_statuses += 1
                    counts = {'in_office_full': 0, 'wfh_full': 0, 'leave_full': 0, 'absent': 0}
                    present_vendor_ids = set()
                    for ds in day_statuses:
                        present_vendor_ids.add(ds.vendor_id)
                        if ds.status in [AttendanceStatus.IN_OFFICE_FULL, AttendanceStatus.IN_OFFICE_HALF]:
                            counts['in_office_full'] += 1
                        elif ds.status in [AttendanceStatus.WFH_FULL, AttendanceStatus.WFH_HALF]:
                            counts['wfh_full'] += 1
                        elif ds.status in [AttendanceStatus.LEAVE_FULL, AttendanceStatus.LEAVE_HALF]:
                            counts['leave_full'] += 1
                        elif ds.status == AttendanceStatus.ABSENT:
                            counts['absent'] += 1
                    counts['absent'] += max(0, total_vendors - len(present_vendor_ids))
                    order = ['in_office_full', 'wfh_full', 'leave_full', 'absent']
                    dominant = max(order, key=lambda k: (counts[k], -order.index(k)))
                    daily_data.append({'date': d.strftime('%Y-%m-%d'), 'status': dominant, 'location': 'Various'})
                d += timedelta(days=1)
            
            potential_entries = total_working_days * total_vendors
            attendance_rate = min(100.0, (submitted_working_statuses / potential_entries * 100)) if potential_entries > 0 else 0
            
            return jsonify({
                'success': True,
                'user_type': 'admin',
                'month': month,
                'status_breakdown': status_counts,
                'approval_breakdown': approval_counts,
                'daily_data': daily_data,
                'summary': {
                    'total_working_days': total_working_days,
                    'submitted_days': submitted_working_statuses,
                    'attendance_rate': round(attendance_rate, 1),
                    'pending_approvals': approval_counts['pending']
                }
            })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/manager/mismatches')
@login_required
def manager_mismatches():
    """Review attendance mismatches between swipe records and submitted status"""
    if current_user.role != UserRole.MANAGER:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    manager = current_user.manager_profile
    if not manager:
        flash('Manager profile not found', 'error')
        return redirect(url_for('login'))
    
    # Get filters
    status_filter = request.args.get('status', 'pending')
    vendor_filter = request.args.get('vendor', 'all')
    
    # Get team vendors
    team_vendors = manager.team_vendors.all()
    vendor_ids = [v.id for v in team_vendors]
    
    # Build query for mismatches
    mismatch_query = MismatchRecord.query.filter(
        MismatchRecord.vendor_id.in_(vendor_ids)
    )
    
    if status_filter != 'all':
        if status_filter == 'pending':
            mismatch_query = mismatch_query.filter(
                MismatchRecord.manager_approval == ApprovalStatus.PENDING
            )
        elif status_filter == 'approved':
            mismatch_query = mismatch_query.filter(
                MismatchRecord.manager_approval == ApprovalStatus.APPROVED
            )
        elif status_filter == 'rejected':
            mismatch_query = mismatch_query.filter(
                MismatchRecord.manager_approval == ApprovalStatus.REJECTED
            )
    
    if vendor_filter != 'all':
        vendor = next((v for v in team_vendors if v.vendor_id == vendor_filter), None)
        if vendor:
            mismatch_query = mismatch_query.filter(MismatchRecord.vendor_id == vendor.id)
    
    # Get mismatches with vendor details
    mismatches = mismatch_query.order_by(MismatchRecord.mismatch_date.desc()).all()
    
    # Group mismatches by vendor
    vendor_mismatches = {}
    for mismatch in mismatches:
        vendor_id = mismatch.vendor.vendor_id
        if vendor_id not in vendor_mismatches:
            vendor_mismatches[vendor_id] = {
                'vendor': mismatch.vendor,
                'mismatches': [],
                'pending_count': 0,
                'total_count': 0
            }
        
        vendor_mismatches[vendor_id]['mismatches'].append(mismatch)
        vendor_mismatches[vendor_id]['total_count'] += 1
        if mismatch.manager_approval == ApprovalStatus.PENDING:
            vendor_mismatches[vendor_id]['pending_count'] += 1
    
    # Calculate summary stats
    total_mismatches = len(mismatches)
    pending_mismatches = len([m for m in mismatches if m.manager_approval == ApprovalStatus.PENDING])
    vendors_with_mismatches = len(vendor_mismatches)
    
    summary_stats = {
        'total_mismatches': total_mismatches,
        'pending_mismatches': pending_mismatches,
        'vendors_affected': vendors_with_mismatches,
        'resolution_rate': round((total_mismatches - pending_mismatches) / total_mismatches * 100, 1) if total_mismatches > 0 else 100
    }
    
    return render_template('manager_mismatches.html',
                         manager=manager,
                         vendor_mismatches=vendor_mismatches,
                         summary_stats=summary_stats,
                         status_filter=status_filter,
                         vendor_filter=vendor_filter,
                         team_vendors=team_vendors,
                         approval_statuses=ApprovalStatus)

@app.route('/manager/mismatches/table')
@login_required
def manager_mismatches_table():
    """Manager's team mismatches in table format (similar to admin reconciliation)"""
    if current_user.role != UserRole.MANAGER:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    manager = current_user.manager_profile
    if not manager:
        flash('Manager profile not found', 'error')
        return redirect(url_for('login'))
    
    # Get filters
    status_filter = request.args.get('status', 'all')
    vendor_filter = request.args.get('vendor', 'all')
    conflict_filter = request.args.get('conflict', 'all')
    
    # Get team vendors
    team_vendors = manager.team_vendors.all()
    vendor_ids = [v.id for v in team_vendors]
    
    # Build query for mismatches
    mismatch_query = MismatchRecord.query.filter(
        MismatchRecord.vendor_id.in_(vendor_ids)
    )
    
    # Apply status filter
    if status_filter != 'all':
        if status_filter == 'pending':
            mismatch_query = mismatch_query.filter(
                MismatchRecord.manager_approval == ApprovalStatus.PENDING
            )
        elif status_filter == 'approved':
            mismatch_query = mismatch_query.filter(
                MismatchRecord.manager_approval == ApprovalStatus.APPROVED
            )
        elif status_filter == 'rejected':
            mismatch_query = mismatch_query.filter(
                MismatchRecord.manager_approval == ApprovalStatus.REJECTED
            )
    
    # Apply vendor filter
    if vendor_filter != 'all':
        vendor = next((v for v in team_vendors if v.vendor_id == vendor_filter), None)
        if vendor:
            mismatch_query = mismatch_query.filter(MismatchRecord.vendor_id == vendor.id)
    
    # Get mismatches ordered by date (latest first)
    mismatches = mismatch_query.order_by(MismatchRecord.mismatch_date.desc()).limit(100).all()
    
    # Apply conflict filter after fetching (since it's based on logic)
    if conflict_filter != 'all':
        filtered_mismatches = []
        for m in mismatches:
            priority = 'low'  # default
            if m.web_status and m.swipe_status:
                if (m.web_status.value in ['wfh_full', 'wfh_half', 'leave_full', 'leave_half'] and m.swipe_status == 'AP'):
                    priority = 'high'
                elif (m.web_status.value in ['in_office_full', 'in_office_half'] and m.swipe_status == 'AA'):
                    priority = 'medium'
            
            if conflict_filter == priority:
                filtered_mismatches.append(m)
        
        mismatches = filtered_mismatches
    
    # Calculate summary stats
    total_mismatches = len(mismatches)
    pending_mismatches = len([m for m in mismatches if m.manager_approval == ApprovalStatus.PENDING])
    approved_mismatches = len([m for m in mismatches if m.manager_approval == ApprovalStatus.APPROVED])
    team_members_count = len(team_vendors)
    
    summary = {
        'total': total_mismatches,
        'pending': pending_mismatches,
        'approved': approved_mismatches,
        'team_members': team_members_count
    }
    
    return render_template('manager_mismatches_table.html',
                         mismatches=mismatches,
                         summary=summary,
                         status_filter=status_filter,
                         vendor_filter=vendor_filter,
                         conflict_filter=conflict_filter,
                         team_vendors=team_vendors)

@app.route('/manager/mismatch/<int:mismatch_id>/approve', methods=['POST'])
@login_required
def approve_mismatch_explanation(mismatch_id):
    """Approve vendor's mismatch explanation"""
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        mismatch = MismatchRecord.query.get_or_404(mismatch_id)
        manager = current_user.manager_profile
        
# Verify vendor is in manager's team
        if mismatch.vendor.manager_id != manager.manager_id:
            return jsonify({'error': 'You can only review mismatches for your team members'}), 403
        
        action = request.form.get('action')
        manager_comments = request.form.get('comments', '')
        
        if action == 'approve':
            mismatch.manager_approval = ApprovalStatus.APPROVED
            message = 'Mismatch explanation approved'
        elif action == 'reject':
            mismatch.manager_approval = ApprovalStatus.REJECTED
            message = 'Mismatch explanation rejected'
        else:
            return jsonify({'error': 'Invalid action'}), 400
        
        mismatch.manager_comments = manager_comments
        mismatch.approved_by = current_user.id
        mismatch.approved_at = datetime.utcnow()
        
        db.session.commit()
        
        # Create audit log
        create_audit_log(current_user.id, action.upper(), 'mismatch_records', mismatch_id,
                       {'manager_approval': 'pending'}, 
                       {'manager_approval': action, 'manager_comments': manager_comments})
        
        return jsonify({
            'status': 'success',
            'message': message,
            'action': action,
            'mismatch_id': mismatch_id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/manager/reports')
@login_required
def manager_reports():
    """Team reports and audit history dashboard"""
    if current_user.role != UserRole.MANAGER:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    manager = current_user.manager_profile
    if not manager:
        flash('Manager profile not found', 'error')
        return redirect(url_for('login'))
    
    # Get date range filters
    start_date_str = request.args.get('start_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()
    
    # Get team vendors
    team_vendors = manager.team_vendors.all()
    vendor_ids = [v.id for v in team_vendors]
    
    # Generate team attendance report
    report_data = []
    for vendor in team_vendors:
        # Get attendance statistics
        statuses = DailyStatus.query.filter(
            DailyStatus.vendor_id == vendor.id,
            DailyStatus.status_date >= start_date,
            DailyStatus.status_date <= end_date
        ).all()
        
        total_days = len(statuses)
        office_days = len([s for s in statuses if s.status in [AttendanceStatus.IN_OFFICE_FULL, AttendanceStatus.IN_OFFICE_HALF]])
        wfh_days = len([s for s in statuses if s.status in [AttendanceStatus.WFH_FULL, AttendanceStatus.WFH_HALF]])
        leave_days = len([s for s in statuses if s.status in [AttendanceStatus.LEAVE_FULL, AttendanceStatus.LEAVE_HALF]])
        pending_days = len([s for s in statuses if s.approval_status == ApprovalStatus.PENDING])
        
        # Calculate working days in period
        working_days = 0
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # Monday to Friday
                is_holiday = Holiday.query.filter_by(holiday_date=current_date).first()
                if not is_holiday:
                    working_days += 1
            current_date += timedelta(days=1)
        
        # Calculate attendance rate properly (should not exceed 100%)
        # Count half-days as 0.5 and only count working days
        working_day_submissions = 0
        for s in statuses:
            # Only count submissions on actual working days
            if s.status_date.weekday() < 5:  # Monday to Friday
                is_holiday = Holiday.query.filter_by(holiday_date=s.status_date).first()
                if not is_holiday:
                    # Count half-days as 0.5, full days as 1
                    if s.status in [AttendanceStatus.IN_OFFICE_HALF, AttendanceStatus.WFH_HALF, AttendanceStatus.LEAVE_HALF]:
                        working_day_submissions += 0.5
                    else:
                        working_day_submissions += 1
        
        # Cap attendance rate at 100% to prevent over 100% values
        attendance_rate = min(100.0, (working_day_submissions / working_days * 100)) if working_days > 0 else 0
        
        report_data.append({
            'vendor': vendor,
            'total_days': total_days,
            'office_days': office_days,
            'wfh_days': wfh_days,
            'leave_days': leave_days,
            'pending_days': pending_days,
            'working_days': working_days,
            'attendance_rate': round(attendance_rate, 1)
        })
    
    # Get audit history for team - show only meaningful manager actions (exclude logins/last_login updates)
    audit_logs = AuditLog.query.filter(
        AuditLog.user_id.in_([manager.user_id] + [v.user_id for v in team_vendors]),
        AuditLog.action.notin_(['LOGIN', 'LOGOUT']),
        AuditLog.table_name != 'users',
        AuditLog.created_at >= datetime.combine(start_date, datetime.min.time()),
        AuditLog.created_at <= datetime.combine(end_date, datetime.max.time())
    ).order_by(AuditLog.created_at.desc()).limit(50).all()
    
    # Calculate summary statistics
    total_submissions = sum([r['total_days'] for r in report_data])
    total_working_days = sum([r['working_days'] for r in report_data])
    avg_attendance_rate = sum([r['attendance_rate'] for r in report_data]) / len(report_data) if report_data else 0
    total_pending = sum([r['pending_days'] for r in report_data])
    
    # Cap percentages to 100% to avoid values like 101.5%
    submission_rate_raw = (total_submissions / total_working_days * 100) if total_working_days > 0 else 0
    submission_rate_capped = round(min(100.0, submission_rate_raw), 1)
    avg_attendance_capped = round(min(100.0, avg_attendance_rate), 1)

    summary_stats = {
        'total_submissions': total_submissions,
        'total_working_days': total_working_days,
        'avg_attendance_rate': avg_attendance_capped,
        'total_pending': total_pending,
        'submission_rate': submission_rate_capped
    }
    
    return render_template('manager_reports.html',
                         manager=manager,
                         report_data=report_data,
                         audit_logs=audit_logs,
                         summary_stats=summary_stats,
                         start_date=start_date,
                         end_date=end_date)

# Export team report for selected date range (Excel or JSON)
@app.route('/api/export/team-report')
@login_required
def api_export_team_report():
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied'}), 403
    
    manager = current_user.manager_profile
    if not manager:
        return jsonify({'error': 'Manager profile not found'}), 404
    
    start_date_str = request.args.get('start_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    format_type = request.args.get('format', 'excel')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    
    team_vendors = manager.team_vendors.all() if manager.team_vendors else []
    rows = []
    
    # Compute working days in period
    working_days_period = 0
    cur = start_date
    while cur <= end_date:
        if cur.weekday() < 5 and (Holiday.query.filter_by(holiday_date=cur).first() is None):
            working_days_period += 1
        cur += timedelta(days=1)
    
    for v in team_vendors:
        statuses = DailyStatus.query.filter(
            DailyStatus.vendor_id == v.id,
            DailyStatus.status_date >= start_date,
            DailyStatus.status_date <= end_date
        ).all()
        total_days = len(statuses)
        office_days = len([s for s in statuses if s.status in [AttendanceStatus.IN_OFFICE_FULL, AttendanceStatus.IN_OFFICE_HALF]])
        wfh_days = len([s for s in statuses if s.status in [AttendanceStatus.WFH_FULL, AttendanceStatus.WFH_HALF]])
        leave_days = len([s for s in statuses if s.status in [AttendanceStatus.LEAVE_FULL, AttendanceStatus.LEAVE_HALF]])
        pending_days = len([s for s in statuses if s.approval_status == ApprovalStatus.PENDING])
        # Calculate attendance rate properly (should not exceed 100%)
        # Count half-days as 0.5 and only count working days
        working_day_submissions = 0
        for s in statuses:
            # Only count submissions on actual working days
            if s.status_date.weekday() < 5:  # Monday to Friday
                is_holiday = Holiday.query.filter_by(holiday_date=s.status_date).first()
                if not is_holiday:
                    # Count half-days as 0.5, full days as 1
                    if s.status in [AttendanceStatus.IN_OFFICE_HALF, AttendanceStatus.WFH_HALF, AttendanceStatus.LEAVE_HALF]:
                        working_day_submissions += 0.5
                    else:
                        working_day_submissions += 1
        
        # Cap attendance rate at 100% to prevent over 100% values
        attendance_rate = min(100.0, round((working_day_submissions / working_days_period * 100), 1)) if working_days_period > 0 else 0
        rows.append({
            'Vendor Name': v.full_name,
            'Vendor ID': v.vendor_id,
            'Department': v.department,
            'Company': v.company,
            'Office Days': office_days,
            'WFH Days': wfh_days,
            'Leave Days': leave_days,
            'Pending Days': pending_days,
            'Working Days': working_days_period,
            'Attendance Rate (%)': attendance_rate
        })
    
    if format_type == 'excel' and rows:
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        df = pd.DataFrame(rows)
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Team Report', index=False)
        excel_buffer.seek(0)
        filename = f"team_report_{start_date_str}_to_{end_date_str}.xlsx"
        return send_file(excel_buffer, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    return jsonify({'status': 'success', 'data': rows, 'start_date': start_date_str, 'end_date': end_date_str})

# Export audit log for selected date range
@app.route('/api/export/audit-log')
@login_required
def api_export_audit_log():
    if current_user.role != UserRole.MANAGER:
        return jsonify({'error': 'Access denied'}), 403
    manager = current_user.manager_profile
    if not manager:
        return jsonify({'error': 'Manager profile not found'}), 404
    start_date_str = request.args.get('start_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    format_type = request.args.get('format', 'excel')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    
    team_vendors = manager.team_vendors.all() if manager.team_vendors else []
    logs = AuditLog.query.filter(
        AuditLog.user_id.in_([manager.user_id] + [v.user_id for v in team_vendors]),
        AuditLog.created_at >= datetime.combine(start_date, datetime.min.time()),
        AuditLog.created_at <= datetime.combine(end_date, datetime.max.time())
    ).order_by(AuditLog.created_at.asc()).all()
    
    rows = []
    for log in logs:
        rows.append({
            'Timestamp': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'Action': log.action,
            'Table': log.table_name,
            'Record ID': log.record_id or '',
            'User ID': log.user_id,
            'Old Values': (log.old_values or '')[:200],
            'New Values': (log.new_values or '')[:200]
        })
    
    if format_type == 'excel' and rows:
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        df = pd.DataFrame(rows)
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Audit Log', index=False)
        excel_buffer.seek(0)
        filename = f"audit_log_{start_date_str}_to_{end_date_str}.xlsx"
        return send_file(excel_buffer, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    return jsonify({'status': 'success', 'data': rows, 'start_date': start_date_str, 'end_date': end_date_str})

# Vendor details for a given date range (manager view)
@app.route('/manager/vendor/<string:vendor_id>/details')
@login_required
def manager_vendor_details(vendor_id):
    if current_user.role != UserRole.MANAGER:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    manager = current_user.manager_profile
    vendor = Vendor.query.filter_by(vendor_id=vendor_id).first()
    if not vendor or vendor.manager_id != manager.manager_id:
        flash('Vendor not found in your team', 'error')
        return redirect(url_for('manager_reports'))
    
    start_date_str = request.args.get('start_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', date.today().strftime('%Y-%m-%d'))
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except Exception:
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()
    
    statuses = DailyStatus.query.filter(
        DailyStatus.vendor_id == vendor.id,
        DailyStatus.status_date >= start_date,
        DailyStatus.status_date <= end_date
    ).order_by(DailyStatus.status_date.asc()).all()
    
    return render_template('manager_vendor_details.html',
                           manager=manager,
                           vendor=vendor,
                           statuses=statuses,
                           start_date=start_date,
                           end_date=end_date)

@app.route('/vendor/edit-status/<int:status_id>', methods=['GET', 'POST'])
@login_required
def vendor_edit_status(status_id):
    """Edit existing daily attendance status"""
    if current_user.role != UserRole.VENDOR:
        return jsonify({'error': 'Access denied'}), 403
    
    vendor = current_user.vendor_profile
    if not vendor:
        return jsonify({'error': 'Vendor profile not found'}), 404
    
    # Get the status record
    status_record = DailyStatus.query.get_or_404(status_id)
    
    # Verify ownership
    if status_record.vendor_id != vendor.id:
        return jsonify({'error': 'Access denied - not your status'}), 403
    
    # Check if status can be edited (only pending or rejected)
    if status_record.approval_status not in [ApprovalStatus.PENDING, ApprovalStatus.REJECTED]:
        return jsonify({'error': 'Cannot edit approved status'}), 400
    
    if request.method == 'GET':
        # Return status data for the edit form
        return jsonify({
            'id': status_record.id,
            'status_date': status_record.status_date.strftime('%Y-%m-%d'),
            'status': status_record.status.value,
            'location': status_record.location or '',
            'comments': status_record.comments or '',
            'approval_status': status_record.approval_status.value
        })
    
    # POST - Update the status
    try:
        status_value = request.form['status']
        location = request.form.get('location', '')
        comments = request.form.get('comments', '')
        
        # Convert string to enum
        status_map = {
            'in_office_full': AttendanceStatus.IN_OFFICE_FULL,
            'wfh_full': AttendanceStatus.WFH_FULL,
            'leave_full': AttendanceStatus.LEAVE_FULL,
            'mixed_half': AttendanceStatus.IN_OFFICE_HALF,  # Use IN_OFFICE_HALF as base for mixed
            'absent': AttendanceStatus.ABSENT
        }
        
        new_status = status_map.get(status_value)
        if not new_status:
            return jsonify({'error': 'Invalid status value'}), 400
        
        # Store old values for audit
        old_values = {
            'status': status_record.status.value,
            'location': status_record.location,
            'comments': status_record.comments,
            'approval_status': status_record.approval_status.value
        }
        
        # Update the status
        status_record.status = new_status
        status_record.location = location
        status_record.comments = comments
        status_record.submitted_at = datetime.utcnow()
        status_record.approval_status = ApprovalStatus.PENDING  # Reset to pending
        status_record.manager_comments = None  # Clear previous manager comments
        
        db.session.commit()
        
        # Create audit log
        new_values = {
            'status': status_value,
            'location': location,
            'comments': comments,
            'approval_status': 'pending'
        }
        create_audit_log(current_user.id, 'UPDATE', 'daily_statuses', status_id, old_values, new_values)
        
        return jsonify({
            'status': 'success',
            'message': 'Status updated successfully!',
            'data': {
                'id': status_record.id,
                'status': new_status.value,
                'location': location,
                'comments': comments,
                'approval_status': 'pending',
                'status_date': status_record.status_date.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Error updating status: {str(e)}'
        }), 500

@app.route('/vendor/submit-status', methods=['POST'])
@login_required
def vendor_submit_status():
    """Submit daily attendance status"""
    if current_user.role != UserRole.VENDOR:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    vendor = current_user.vendor_profile
    if not vendor:
        flash('Vendor profile not found', 'error')
        return redirect(url_for('login'))
    
    try:
        status_date = datetime.strptime(request.form['status_date'], '%Y-%m-%d').date()
        
        # Check if the date is a weekend or holiday - prevent attendance submission
        from utils import is_non_working_day, get_non_working_day_reason
        if is_non_working_day(status_date):
            reason = get_non_working_day_reason(status_date)
            flash(f'Attendance cannot be submitted for {reason}. No attendance is required on non-working days.', 'warning')
            return redirect(url_for('vendor_dashboard'))
        
        status_value = request.form['status']
        location = request.form.get('location', '')
        comments = request.form.get('comments', '')
        
        # Time tracking fields
        in_time = request.form.get('in_time')
        out_time = request.form.get('out_time')
        office_in_time = request.form.get('office_in_time')
        office_out_time = request.form.get('office_out_time')
        wfh_in_time = request.form.get('wfh_in_time')
        wfh_out_time = request.form.get('wfh_out_time')
        
        # Half-day specific time fields
        office_in_time_am = request.form.get('office_in_time_am')
        office_out_time_am = request.form.get('office_out_time_am')
        wfh_in_time_am = request.form.get('wfh_in_time_am')
        wfh_out_time_am = request.form.get('wfh_out_time_am')
        office_in_time_pm = request.form.get('office_in_time_pm')
        office_out_time_pm = request.form.get('office_out_time_pm')
        wfh_in_time_pm = request.form.get('wfh_in_time_pm')
        wfh_out_time_pm = request.form.get('wfh_out_time_pm')
        
        break_duration = int(request.form.get('break_duration', 0)) if request.form.get('break_duration') else 0
        total_hours = float(request.form.get('total_hours', 0)) if request.form.get('total_hours') else None
        
        # Half-day type fields
        half_am_type = request.form.get('half_am_type')
        half_pm_type = request.form.get('half_pm_type')
        am_enum = None
        pm_enum = None
        
        # Convert time strings to time objects
        def parse_time(time_str):
            if time_str:
                try:
                    return datetime.strptime(time_str, '%H:%M').time()
                except ValueError:
                    return None
            return None
        
        in_time_obj = parse_time(in_time)
        out_time_obj = parse_time(out_time)
        office_in_time_obj = parse_time(office_in_time)
        office_out_time_obj = parse_time(office_out_time)
        wfh_in_time_obj = parse_time(wfh_in_time)
        wfh_out_time_obj = parse_time(wfh_out_time)
        
        # Process half-day specific times based on status
        if status_value == 'mixed_half':
            # Consolidate half-day times into appropriate fields
            # AM times
            if half_am_type == 'in_office' and office_in_time_am and office_out_time_am:
                office_in_time_obj = parse_time(office_in_time_am)
                office_out_time_obj = parse_time(office_out_time_am) if not office_in_time_pm else None
            elif half_am_type == 'wfh' and wfh_in_time_am and wfh_out_time_am:
                wfh_in_time_obj = parse_time(wfh_in_time_am)
                wfh_out_time_obj = parse_time(wfh_out_time_am) if not wfh_in_time_pm else None
            
            # PM times
            if half_pm_type == 'in_office' and office_in_time_pm and office_out_time_pm:
                if not office_in_time_obj:
                    office_in_time_obj = parse_time(office_in_time_pm)
                office_out_time_obj = parse_time(office_out_time_pm)
            elif half_pm_type == 'wfh' and wfh_in_time_pm and wfh_out_time_pm:
                if not wfh_in_time_obj:
                    wfh_in_time_obj = parse_time(wfh_in_time_pm)
                wfh_out_time_obj = parse_time(wfh_out_time_pm)
            
            # Calculate total hours if not provided
            if not total_hours or total_hours == 0:
                total_mins = 0
                
                # Helper function to calculate minutes between times
                def calc_minutes(start_time, end_time):
                    if start_time and end_time:
                        start_mins = start_time.hour * 60 + start_time.minute
                        end_mins = end_time.hour * 60 + end_time.minute
                        if end_mins < start_mins:
                            end_mins += 24 * 60
                        return end_mins - start_mins
                    return 0
                
                # Calculate AM hours
                if half_am_type == 'in_office' and office_in_time_am and office_out_time_am:
                    total_mins += calc_minutes(parse_time(office_in_time_am), parse_time(office_out_time_am))
                elif half_am_type == 'wfh' and wfh_in_time_am and wfh_out_time_am:
                    total_mins += calc_minutes(parse_time(wfh_in_time_am), parse_time(wfh_out_time_am))
                
                # Calculate PM hours
                if half_pm_type == 'in_office' and office_in_time_pm and office_out_time_pm:
                    total_mins += calc_minutes(parse_time(office_in_time_pm), parse_time(office_out_time_pm))
                elif half_pm_type == 'wfh' and wfh_in_time_pm and wfh_out_time_pm:
                    total_mins += calc_minutes(parse_time(wfh_in_time_pm), parse_time(wfh_out_time_pm))
                
                # Subtract break time and convert to hours
                total_mins -= break_duration
                total_hours = max(0, total_mins / 60.0)
                print(f"Calculated total_hours for half-day: {total_hours}")
        else:
            # For full-day statuses, calculate total hours if not provided
            if (not total_hours or total_hours == 0) and in_time_obj and out_time_obj:
                def calc_minutes(start_time, end_time):
                    if start_time and end_time:
                        start_mins = start_time.hour * 60 + start_time.minute
                        end_mins = end_time.hour * 60 + end_time.minute
                        if end_mins < start_mins:
                            end_mins += 24 * 60
                        return end_mins - start_mins
                    return 0
                
                total_mins = calc_minutes(in_time_obj, out_time_obj)
                total_mins -= break_duration
                total_hours = max(0, total_mins / 60.0)
                print(f"Calculated total_hours for full-day: {total_hours}")
        
        # Convert string to enum
        status_map = {
            'in_office_full': AttendanceStatus.IN_OFFICE_FULL,
            'wfh_full': AttendanceStatus.WFH_FULL,
            'leave_full': AttendanceStatus.LEAVE_FULL,
            'mixed_half': AttendanceStatus.IN_OFFICE_HALF,  # Use IN_OFFICE_HALF as base for mixed
            'absent': AttendanceStatus.ABSENT
        }
        
        status = status_map.get(status_value)
        if not status:
            flash('Invalid status value', 'error')
            return redirect(url_for('vendor_dashboard'))
        
        # Handle mixed half-day submissions
        if status_value == 'mixed_half':
            if not half_am_type or not half_pm_type:
                flash('Both AM and PM activities must be specified for half-day status', 'error')
                return redirect(url_for('vendor_dashboard'))
            
            # Validate half-day type values
            valid_half_day_types = ['in_office', 'wfh', 'leave', 'absent']
            if half_am_type not in valid_half_day_types or half_pm_type not in valid_half_day_types:
                flash('Invalid half-day activity type', 'error')
                return redirect(url_for('vendor_dashboard'))
            
            # ALL COMBINATIONS ARE NOW ALLOWED - No restrictions
            
            # Convert to enums - ALL COMBINATIONS ALLOWED
            from models import HalfDayType
            half_day_map = {
                'in_office': HalfDayType.IN_OFFICE,
                'wfh': HalfDayType.WFH,
                'leave': HalfDayType.LEAVE,
                'absent': HalfDayType.ABSENT
            }
            
            am_enum = half_day_map[half_am_type]
            pm_enum = half_day_map[half_pm_type]
            
            # Log the combination for audit purposes
            print(f"âœ… Half-day submission: AM={half_am_type}, PM={half_pm_type} for vendor {vendor.vendor_id}")
        else:
            # Clear half-day types for full-day statuses
            am_enum = None
            pm_enum = None
        
        # Check if status already exists
        existing_status = DailyStatus.query.filter_by(
            vendor_id=vendor.id,
            status_date=status_date
        ).first()
        
        if existing_status:
            existing_status.status = status
            existing_status.location = location
            existing_status.comments = comments
            existing_status.in_time = in_time_obj
            existing_status.out_time = out_time_obj
            existing_status.office_in_time = office_in_time_obj
            existing_status.office_out_time = office_out_time_obj
            existing_status.wfh_in_time = wfh_in_time_obj
            existing_status.wfh_out_time = wfh_out_time_obj
            existing_status.break_duration = break_duration
            existing_status.total_hours = total_hours
            existing_status.half_am_type = am_enum
            existing_status.half_pm_type = pm_enum
            existing_status.submitted_at = datetime.utcnow()
            existing_status.approval_status = ApprovalStatus.PENDING
        else:
            new_status = DailyStatus(
                vendor_id=vendor.id,
                status_date=status_date,
                status=status,
                location=location,
                comments=comments,
                in_time=in_time_obj,
                out_time=out_time_obj,
                office_in_time=office_in_time_obj,
                office_out_time=office_out_time_obj,
                wfh_in_time=wfh_in_time_obj,
                wfh_out_time=wfh_out_time_obj,
                break_duration=break_duration,
                total_hours=total_hours,
                half_am_type=am_enum,
                half_pm_type=pm_enum
            )
            db.session.add(new_status)
        
        db.session.commit()
        
        create_audit_log(current_user.id, 'CREATE' if not existing_status else 'UPDATE', 
                       'daily_statuses', existing_status.id if existing_status else None, 
                       {}, {'status': status_value, 'location': location})
        
        # ðŸ†• NEW: Update Excel sheets immediately for Power Automate
        try:
            from scripts.daily_excel_updater import handle_vendor_status_submission
            handle_vendor_status_submission(vendor.vendor_id)
            print(f"âœ… Excel sheets updated for vendor {vendor.vendor_id}")
        except ImportError:
            try:
                from scripts.power_automate_excel_refresh import power_automate_excel_refresh
                power_automate_excel_refresh()
                print(f"âœ… Fallback Excel refresh completed for vendor {vendor.vendor_id}")
            except Exception as fallback_error:
                print(f"âš ï¸ Both Excel update methods failed for vendor {vendor.vendor_id}: {str(fallback_error)}")
        except Exception as excel_error:
            # Don't fail the main submission if Excel update fails
            print(f"âš ï¸ Excel update failed for vendor {vendor.vendor_id}: {str(excel_error)}")
        
        flash('Status submitted successfully!', 'success')
        return redirect(url_for('vendor_dashboard'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error submitting status: {str(e)}', 'error')
        return redirect(url_for('vendor_dashboard'))

@app.route('/manager/approve-status/<int:status_id>', methods=['POST'])
@login_required
def api_approve_status(status_id):
    """Approve or reject vendor status"""
    try:
        # Check if user is manager
        if current_user.role != UserRole.MANAGER:
            print(f"âŒ Access denied - User role: {current_user.role}")
            return jsonify({'error': 'Access denied - Manager role required'}), 403
        
        # Get manager profile
        manager = current_user.manager_profile
        if not manager:
            print(f"âŒ Manager profile not found for user: {current_user.username}")
            return jsonify({'error': 'Manager profile not found'}), 404
        
        # Get status record with vendor eagerly loaded to avoid detached/expired issues
        daily_status = DailyStatus.query.options(joinedload(DailyStatus.vendor)).get(status_id)
        if not daily_status:
            print(f"âŒ Status record not found: {status_id}")
            return jsonify({'error': 'Status record not found'}), 404
        
        # Verify vendor is in manager's team
        vendor = daily_status.vendor
        # Compare against manager.manager_id (string FK), not manager.id
        if not vendor or vendor.manager_id != manager.manager_id:
            vid = vendor.vendor_id if vendor else 'UNKNOWN'
            print(f"âŒ Vendor {vid} not in manager's team")
            return jsonify({'error': 'You can only approve statuses for your team members'}), 403
        
        # Get action and reason
        action = request.form.get('action')
        reason = request.form.get('reason', '')
        
        print(f"ðŸ“ Processing {action} for status {status_id} by {current_user.username}")
        
        # Precompute values for response to avoid accessing expired ORM attributes after commit
        if action == 'approve':
            daily_status.approval_status = ApprovalStatus.APPROVED
            manager_comments_value = reason or 'Approved'
            daily_status.manager_comments = manager_comments_value
            approval_status_value = 'approved'
            message = 'Status approved successfully'
            print(f"âœ… Approved status {status_id}")
        elif action == 'reject':
            daily_status.approval_status = ApprovalStatus.REJECTED
            manager_comments_value = reason or 'Rejected'
            daily_status.manager_comments = manager_comments_value
            approval_status_value = 'rejected'
            message = 'Status rejected'
            print(f"âŒ Rejected status {status_id} - Reason: {reason}")
        else:
            print(f"âŒ Invalid action: {action}")
            return jsonify({'error': 'Invalid action - must be approve or reject'}), 400
        
        # Update approval metadata
        daily_status.approved_at = datetime.utcnow()
        daily_status.approved_by = current_user.id
        
        # Cache vendor_id for post-commit tasks
        vendor_id_for_excel = vendor.vendor_id
        
        # Commit to database
        db.session.commit()
        print(f"âœ… Database updated successfully")
        
        # Create audit log
        create_audit_log(current_user.id, action.upper(), 'daily_statuses', status_id, 
                       {'approval_status': 'pending'}, 
                       {'approval_status': action, 'manager_comments': reason})
        
        # Update Excel sheets immediately for Power Automate (same logic as vendor submission)
        try:
            from scripts.daily_excel_updater import handle_vendor_status_submission
            handle_vendor_status_submission(vendor_id_for_excel)
            print(f"âœ… Excel sheets updated after manager {action} for vendor {vendor_id_for_excel}")
        except ImportError:
            try:
                from scripts.power_automate_excel_refresh import power_automate_excel_refresh
                power_automate_excel_refresh()
                print(f"âœ… Fallback Excel refresh completed after manager {action} for vendor {vendor_id_for_excel}")
            except Exception as fallback_error:
                print(f"âš ï¸ Both Excel update methods failed after manager {action} for vendor {vendor_id_for_excel}: {str(fallback_error)}")
        except Exception as excel_error:
            # Don't fail the main approval if Excel update fails
            print(f"âš ï¸ Excel update failed after manager action for vendor {vendor_id_for_excel}: {str(excel_error)}")
        
        return jsonify({
            'status': 'success',
            'message': message,
            'action': action,
            'status_id': status_id,
            'approval_status': approval_status_value,
            'manager_comments': manager_comments_value
        })
        
    except Exception as e:
        print(f"âŒ Error in approve/reject: {str(e)}")
        db.session.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Server error: {str(e)}'
        }), 500

@app.route('/manager/approve-status/bulk', methods=['POST'])
@login_required
def api_bulk_approve_status():
    """Bulk approve or reject vendor statuses for a manager's team"""
    try:
        if current_user.role != UserRole.MANAGER:
            return jsonify({'error': 'Access denied - Manager role required'}), 403
        
        manager = current_user.manager_profile
        if not manager:
            return jsonify({'error': 'Manager profile not found'}), 404
        
        data = request.get_json(silent=True) or {}
        action = data.get('action')
        status_ids = data.get('status_ids', [])
        reason = data.get('reason', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({'error': 'Invalid action'}), 400
        if not status_ids:
            return jsonify({'error': 'No status IDs provided'}), 400
        
        # Limit to manager's team
        team_vendor_ids = [v.id for v in manager.team_vendors.all()] if manager.team_vendors else []
        if not team_vendor_ids:
            return jsonify({'error': 'No team members found'}), 400
        
        # Eager-load vendor to avoid detached instances
        statuses = DailyStatus.query.options(joinedload(DailyStatus.vendor)).filter(
            DailyStatus.id.in_(status_ids),
            DailyStatus.vendor_id.in_(team_vendor_ids)
        ).all()
        
        updated = 0
        affected_vendor_ids = set()
        for ds in statuses:
            old_status = ds.approval_status.value
            if action == 'approve':
                ds.approval_status = ApprovalStatus.APPROVED
                ds.manager_comments = reason or 'Approved'
            else:
                ds.approval_status = ApprovalStatus.REJECTED
                ds.manager_comments = reason or 'Rejected'
            ds.approved_at = datetime.utcnow()
            ds.approved_by = current_user.id
            updated += 1
            if ds.vendor:
                affected_vendor_ids.add(ds.vendor.vendor_id)
            # Audit log per record
            create_audit_log(current_user.id, action.upper(), 'daily_statuses', ds.id,
                             {'approval_status': old_status},
                             {'approval_status': action, 'manager_comments': ds.manager_comments})
        
        db.session.commit()
        
        # Update Excel sheets for all affected vendors (same logic as individual approvals)
        try:
            from scripts.daily_excel_updater import handle_vendor_status_submission
            for vendor_id in affected_vendor_ids:
                try:
                    handle_vendor_status_submission(vendor_id)
                    print(f"âœ… Excel sheets updated for vendor {vendor_id} in bulk {action}")
                except Exception as vendor_excel_error:
                    print(f"âš ï¸ Excel update failed for vendor {vendor_id} in bulk operation: {str(vendor_excel_error)}")
        except ImportError:
            try:
                from scripts.power_automate_excel_refresh import power_automate_excel_refresh
                power_automate_excel_refresh()
                print(f"âœ… Fallback Excel refresh completed for bulk {action} operation")
            except Exception as fallback_error:
                print(f"âš ï¸ Bulk Excel update fallback failed: {str(fallback_error)}")
        except Exception as excel_error:
            print(f"âš ï¸ Bulk Excel update failed: {str(excel_error)}")
        
        return jsonify({'status': 'success', 'updated': updated, 'action': action})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/admin/add-holiday', methods=['POST'])
@login_required
def api_add_holiday():
    """Add a new holiday"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        holiday_date = datetime.strptime(request.form['holiday_date'], '%Y-%m-%d').date()
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if holiday already exists
        existing_holiday = Holiday.query.filter_by(holiday_date=holiday_date).first()
        if existing_holiday:
            return jsonify({'error': 'Holiday already exists for this date'}), 400
        
        new_holiday = Holiday(
            holiday_date=holiday_date,
            name=name,
            description=description,
            created_by=current_user.id
        )
        
        db.session.add(new_holiday)
        db.session.commit()
        
        # Optional: audit log
        try:
            create_audit_log(current_user.id, 'CREATE', 'holidays', new_holiday.id, {}, {
                'holiday_date': holiday_date.isoformat(),
                'name': name,
                'description': description
            })
        except Exception:
            # Audit log failure should not block the main operation
            pass
        
        return jsonify({
            'status': 'success',
            'message': 'Holiday added successfully',
            'data': {
                'date': holiday_date.isoformat(),
                'name': name,
                'description': description
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/holiday/<int:holiday_id>/update', methods=['POST'])
@login_required
def api_update_holiday(holiday_id):
    """Update an existing holiday"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        holiday = Holiday.query.get(holiday_id)
        if not holiday:
            return jsonify({'error': 'Holiday not found'}), 404
        
        # Support form posts
        holiday_date_str = request.form.get('holiday_date')
        name = request.form.get('name')
        description = request.form.get('description', '')
        
        if not holiday_date_str or not name:
            return jsonify({'error': 'holiday_date and name are required'}), 400
        
        new_date = datetime.strptime(holiday_date_str, '%Y-%m-%d').date()
        
        # Uniqueness check for date if changed
        if new_date != holiday.holiday_date:
            dup = Holiday.query.filter(Holiday.holiday_date == new_date, Holiday.id != holiday_id).first()
            if dup:
                return jsonify({'error': 'Another holiday already exists for this date'}), 400
        
        old_values = {
            'holiday_date': holiday.holiday_date.isoformat(),
            'name': holiday.name,
            'description': holiday.description or ''
        }
        
        holiday.holiday_date = new_date
        holiday.name = name
        holiday.description = description
        
        db.session.commit()
        
        try:
            create_audit_log(current_user.id, 'UPDATE', 'holidays', holiday.id, old_values, {
                'holiday_date': holiday.holiday_date.isoformat(),
                'name': holiday.name,
                'description': holiday.description or ''
            })
        except Exception:
            pass
        
        return jsonify({'status': 'success', 'message': 'Holiday updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/holiday/<int:holiday_id>/delete', methods=['POST'])
@login_required
def api_delete_holiday(holiday_id):
    """Delete an existing holiday"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        holiday = Holiday.query.get(holiday_id)
        if not holiday:
            return jsonify({'error': 'Holiday not found'}), 404
        
        old_values = {
            'holiday_date': holiday.holiday_date.isoformat(),
            'name': holiday.name,
            'description': holiday.description or ''
        }
        
        db.session.delete(holiday)
        db.session.commit()
        
        try:
            create_audit_log(current_user.id, 'DELETE', 'holidays', holiday_id, old_values, {})
        except Exception:
            pass
        
        return jsonify({'status': 'success', 'message': 'Holiday deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/vendor/mismatch/<int:mismatch_id>/explain', methods=['POST'])
@login_required
def submit_mismatch_explanation(mismatch_id):
    """Submit explanation for attendance mismatch"""
    if current_user.role != UserRole.VENDOR:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    mismatch = MismatchRecord.query.get_or_404(mismatch_id)
    
    # Verify this mismatch belongs to current vendor
    vendor = current_user.vendor_profile
    if not vendor or mismatch.vendor_id != vendor.id:
        flash('Access denied', 'error')
        return redirect(url_for('vendor_dashboard'))
    
    try:
        reason = request.form['reason']
        mismatch.vendor_reason = reason
        mismatch.vendor_submitted_at = datetime.utcnow()
        
        db.session.commit()
        
        create_audit_log(current_user.id, 'UPDATE', 'mismatch_records', mismatch.id, 
                       {'vendor_reason': mismatch.vendor_reason}, 
                       {'vendor_reason': reason})
        
        flash('Mismatch explanation submitted successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error submitting explanation: {str(e)}', 'error')
    
    return redirect(url_for('vendor_dashboard'))

@app.route('/profile')
@login_required
def profile():
    """User profile page - view personal information"""
    # Get user profile information based on role
    profile_data = {
        'user': current_user,
        'vendor_profile': None,
        'manager_profile': None
    }
    
    if current_user.role == UserRole.VENDOR:
        profile_data['vendor_profile'] = current_user.vendor_profile
    elif current_user.role == UserRole.MANAGER:
        profile_data['manager_profile'] = current_user.manager_profile
        # Get team size for managers
        if profile_data['manager_profile']:
            team_count = Vendor.query.filter_by(manager_id=profile_data['manager_profile'].manager_id).count()
            profile_data['team_count'] = team_count
    
    return render_template('profile.html', **profile_data)

@app.route('/profile/change-password', methods=['POST'])
@login_required
def change_password():
    """Change user password"""
    try:
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validation
        if not current_password:
            flash('Current password is required', 'error')
            return redirect(url_for('profile'))
        
        if not new_password:
            flash('New password is required', 'error')
            return redirect(url_for('profile'))
        
        if len(new_password) < 6:
            flash('New password must be at least 6 characters long', 'error')
            return redirect(url_for('profile'))
        
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return redirect(url_for('profile'))
        
        # Verify current password
        if not current_user.check_password(current_password):
            flash('Current password is incorrect', 'error')
            return redirect(url_for('profile'))
        
        # Update password
        old_values = {'password_changed': False}
        current_user.set_password(new_password)
        db.session.commit()
        
        # Create audit log
        try:
            create_audit_log(current_user.id, 'UPDATE', 'users', current_user.id, 
                           old_values, {'password_changed': True})
        except Exception:
            pass
        
        flash('Password changed successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error changing password: {str(e)}', 'error')
    
    return redirect(url_for('profile'))

@app.route('/api/profile/info')
@login_required
def api_profile_info():
    """API endpoint to get user profile information"""
    try:
        profile_info = {
            'id': current_user.id,
            'username': current_user.username,
            'email': current_user.email,
            'role': current_user.role.value,
            'is_active': current_user.is_active,
            'created_at': current_user.created_at.isoformat() if current_user.created_at else None,
            'last_login': current_user.last_login.isoformat() if current_user.last_login else None
        }
        
        # Add role-specific information
        if current_user.role == UserRole.VENDOR and current_user.vendor_profile:
            vendor = current_user.vendor_profile
            manager = Manager.query.filter_by(manager_id=vendor.manager_id).first() if vendor.manager_id else None
            
            profile_info['vendor_info'] = {
                'vendor_id': vendor.vendor_id,
                'full_name': vendor.full_name,
                'department': vendor.department,
                'company': vendor.company,
                'band': vendor.band,
                'location': vendor.location,
                'manager_name': manager.full_name if manager else 'Not Assigned',
                'created_at': vendor.created_at.isoformat() if vendor.created_at else None
            }
            
        elif current_user.role == UserRole.MANAGER and current_user.manager_profile:
            manager = current_user.manager_profile
            team_count = Vendor.query.filter_by(manager_id=manager.manager_id).count()
            
            profile_info['manager_info'] = {
                'manager_id': manager.manager_id,
                'full_name': manager.full_name,
                'department': manager.department,
                'team_name': manager.team_name,
                'email': manager.email,
                'phone': manager.phone,
                'team_count': team_count,
                'created_at': manager.created_at.isoformat() if manager.created_at else None
            }
        
        return jsonify(profile_info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# =============================================================================
# EXCEL SYNC FUNCTIONALITY
# =============================================================================

def excel_log_message(message):
    """Log message with timestamp for Excel sync"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[EXCEL SYNC {timestamp}] {message}")

def clear_notification_data(file_name=None):
    """Clear data rows from Excel notification files while preserving headers and structure
    
    Args:
        file_name (str, optional): Specific Excel file to clear. If None, clears all notification files.
    
    Returns:
        dict: Results of the clearing operation
    """
    if not app.config['EXCEL_NETWORK_FOLDER']:
        excel_log_message("âš ï¸ Network folder not configured")
        return {'success': False, 'message': 'Network folder not configured', 'cleared': 0}
    
    try:
        network_path = Path(app.config['EXCEL_NETWORK_FOLDER'])
        
        if not network_path.exists():
            excel_log_message(f"âš ï¸ Network folder not found: {network_path}")
            return {'success': False, 'message': f'Network folder not found: {network_path}', 'cleared': 0}
        
        # Determine which files to process
        if file_name:
            files_to_process = [network_path / file_name]
            if not files_to_process[0].exists():
                return {'success': False, 'message': f'File not found: {file_name}', 'cleared': 0}
        else:
            all_excel_files = list(network_path.glob("*.xlsx"))
            # Filter out temporary Excel lock files and backup files
            files_to_process = [p for p in all_excel_files if not p.name.startswith("~$") 
                              and not p.name.startswith("PA_backup_")
                              and not p.name.lower().startswith("pa_backup_")]
        
        if not files_to_process:
            excel_log_message("âš ï¸ No Excel files found in network folder")
            return {'success': False, 'message': 'No Excel files found', 'cleared': 0}
        
        excel_log_message(f"ðŸ§¹ Clearing notification data from {len(files_to_process)} files")
        
        files_cleared = 0
        cleared_files = []
        
        for file_path in files_to_process:
            try:
                # Load workbook and first worksheet
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Get the header row
                header_row = [cell.value for cell in ws[1]]
                
                # Find data rows (skip header)
                data_rows = ws.max_row - 1
                
                if data_rows > 0:
                    # Delete all rows except header
                    ws.delete_rows(2, data_rows)
                    
                    # Save the file
                    wb.save(file_path)
                    excel_log_message(f"âœ… Cleared {data_rows} rows from {file_path.name}")
                    files_cleared += 1
                    cleared_files.append(file_path.name)
                else:
                    excel_log_message(f"â„¹ï¸ No data rows to clear in {file_path.name}")
                
            except Exception as e:
                error_msg = f"âŒ Failed to clear {file_path.name}: {str(e)}"
                excel_log_message(error_msg)
                excel_sync_status['errors'].append(error_msg)
        
        result = {
            'success': True,
            'message': f'Successfully cleared {files_cleared} files',
            'cleared': files_cleared,
            'files': cleared_files
        }
        
        return result
        
    except Exception as e:
        error_msg = f"âŒ Failed to clear notification data: {str(e)}"
        excel_log_message(error_msg)
        excel_sync_status['errors'].append(error_msg)
        return {'success': False, 'message': str(e), 'cleared': 0}

def sync_excel_files():
    """Copy Excel files from local to network drive and apply Power Automate formatting with vendor data"""
    global excel_sync_status
    
    if not app.config['EXCEL_NETWORK_FOLDER']:
        excel_log_message("âš ï¸ Network folder not configured")
        return
    
    try:
        local_path = Path(app.config['EXCEL_LOCAL_FOLDER'])
        network_path = Path(app.config['EXCEL_NETWORK_FOLDER'])
        
        if not local_path.exists():
            raise FileNotFoundError(f"Local folder not found: {local_path}")
        
        # Create network folder if it doesn't exist
        network_path.mkdir(parents=True, exist_ok=True)
        
        # Find all Excel files
        excel_files = list(local_path.glob("*.xlsx"))
        
        if not excel_files:
            excel_log_message("âš ï¸ No Excel files found in local folder")
            return
        
        # Import the vendor data population script
        try:
            from scripts.populate_excel_from_vendors import get_vendor_data, create_notification_excel
            populate_with_vendors = True
            excel_log_message("âœ… Will populate network files with vendor data")
        except ImportError:
            populate_with_vendors = False
            excel_log_message("âš ï¸ Vendor population not available, using basic copy")
        
        files_copied = 0
        files_formatted = 0
        
        # Get vendor data if available
        if populate_with_vendors:
            vendors_data, vendor_emails = get_vendor_data()
            if not vendors_data:
                # Use sample data if no vendors in database
                vendors_data = [
                    ('VENDOR001', 'Jane Vendor', 'IT', 'Company A', 'MGR001'),
                    ('VENDOR002', 'Mike Vendor', 'IT', 'Company A', 'MGR001'),
                    ('VENDOR003', 'Sarah Vendor', 'Finance', 'Company B', 'MGR002'),
                    ('VENDOR004', 'David Vendor', 'Finance', 'Company B', 'MGR002'),
                    ('TEST001', 'Test User', 'Testing', 'Company C', 'MGR003')
                ]
                vendor_emails = {
                    'VENDOR001': 'jane.vendor@company.com',
                    'VENDOR002': 'mike.vendor@company.com',
                    'VENDOR003': 'sarah.vendor@company.com',
                    'VENDOR004': 'david.vendor@company.com',
                    'TEST001': 'test001@company.com'
                }
            excel_log_message(f"ðŸ“Š Found {len(vendors_data)} vendors to populate")
        
        # Define notification type mapping
        notification_type_map = {
            '01_daily_status_reminders.xlsx': 'Daily Reminder',
            '02_manager_summary_notifications.xlsx': 'Manager Summary',
            '03_manager_all_complete_notifications.xlsx': 'Manager Summary',
            '04_mismatch_notifications.xlsx': 'Mismatch Alert',
            '05_manager_feedback_notifications.xlsx': 'Manager Summary',
            '06_monthly_report_notifications.xlsx': 'Manager Summary',
            '07_admin_system_alerts.xlsx': 'System Alert',
            '08_holiday_reminder_notifications.xlsx': 'Holiday Reminder',
            '09_late_submission_alerts.xlsx': 'Late Submission'
        }
        
        for file_path in excel_files:
            try:
                dest_path = network_path / file_path.name
                
                if populate_with_vendors and file_path.name in notification_type_map:
                    # Create formatted file with vendor data for network folder
                    notification_type = notification_type_map[file_path.name]
                    create_notification_excel(dest_path, notification_type, vendors_data, vendor_emails)
                    excel_log_message(f"âœ… Created {file_path.name} with vendor data for Power Automate")
                    files_formatted += 1
                else:
                    # Just copy the file as-is
                    shutil.copy2(file_path, dest_path)
                    excel_log_message(f"ðŸ“‹ Copied: {file_path.name}")
                
                files_copied += 1
                
            except Exception as e:
                error_msg = f"âŒ Failed to process {file_path.name}: {str(e)}"
                excel_log_message(error_msg)
                excel_sync_status['errors'].append(error_msg)
        
        excel_sync_status['last_sync'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        excel_sync_status['files_synced'] = files_copied
        excel_sync_status['status'] = 'Running'
        
        if populate_with_vendors and files_formatted > 0:
            excel_log_message(f"ðŸ“Š Sync completed: {files_copied}/{len(excel_files)} files, {files_formatted} with vendor data")
        else:
            excel_log_message(f"ðŸ“Š Sync completed: {files_copied}/{len(excel_files)} files")
        
    except Exception as e:
        error_msg = f"âŒ Sync failed: {str(e)}"
        excel_log_message(error_msg)
        excel_sync_status['errors'].append(error_msg)
        excel_sync_status['status'] = 'Error'

def excel_sync_loop():
    """Main Excel sync loop that runs every 10 minutes"""
    global excel_sync_running, excel_sync_paused
    
    excel_log_message("ðŸš€ Excel sync service started")
    
    while excel_sync_running:
        if not excel_sync_paused:
            sync_excel_files()
        else:
            excel_sync_status['status'] = 'Paused'
        
        # Wait for next sync (check every 10 seconds for stop/pause commands)
        for _ in range(app.config['EXCEL_SYNC_INTERVAL'] // 10):
            if not excel_sync_running:
                break
            time.sleep(10)
    
    excel_log_message("ðŸ›‘ Excel sync service stopped")

def start_excel_sync():
    """Start the Excel sync service"""
    global excel_sync_running, excel_sync_thread
    
    if excel_sync_running:
        return "Excel sync is already running"
    
    if not app.config['EXCEL_NETWORK_FOLDER']:
        return "Network folder not configured. Please set it in Admin Settings."
    
    excel_sync_running = True
    excel_sync_paused = False
    excel_sync_thread = threading.Thread(target=excel_sync_loop, daemon=True)
    excel_sync_thread.start()
    return "Excel sync service started"

def stop_excel_sync():
    """Stop the Excel sync service"""
    global excel_sync_running
    excel_sync_running = False
    excel_sync_status['status'] = 'Stopped'
    return "Excel sync service stopped"

def pause_excel_sync():
    """Pause the Excel sync service"""
    global excel_sync_paused
    excel_sync_paused = True
    excel_sync_status['status'] = 'Paused'
    return "Excel sync paused"

def resume_excel_sync():
    """Resume the Excel sync service"""
    global excel_sync_paused
    excel_sync_paused = False
    excel_sync_status['status'] = 'Running'
    return "Excel sync resumed"

# Excel Sync Admin Routes
@app.route('/admin/excel-sync')
@login_required
def admin_excel_sync():
    """Excel sync admin dashboard"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    return render_template('admin_excel_sync.html',
                         sync_status=excel_sync_status,
                         network_folder=app.config['EXCEL_NETWORK_FOLDER'],
                         local_folder=app.config['EXCEL_LOCAL_FOLDER'],
                         sync_running=excel_sync_running,
                         sync_paused=excel_sync_paused)

@app.route('/api/excel-sync/control/<action>', methods=['POST'])
@login_required
def api_excel_sync_control(action):
    """Control Excel sync service via API"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if action == 'start':
        message = start_excel_sync()
    elif action == 'stop':
        message = stop_excel_sync()
    elif action == 'pause':
        message = pause_excel_sync()
    elif action == 'resume':
        message = resume_excel_sync()
    else:
        return jsonify({'message': 'Unknown action'}), 400
    
    return jsonify({'message': message})

@app.route('/api/excel-sync/force', methods=['POST'])
@login_required
def api_excel_sync_force():
    """Force immediate Excel sync"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    sync_excel_files()
    return jsonify({'message': 'Force sync completed'})

@app.route('/api/excel-sync/config', methods=['POST'])
@login_required
def api_excel_sync_config():
    """Configure Excel sync network folder"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        network_folder = data.get('network_folder', '').strip()
        
        if not network_folder:
            return jsonify({'error': 'Network folder path is required'}), 400
        
        # Validate the path
        try:
            Path(network_folder).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return jsonify({'error': f'Invalid network folder path: {str(e)}'}), 400
        
        app.config['EXCEL_NETWORK_FOLDER'] = network_folder
        
        # Save to system config
        set_system_config('excel_network_folder', network_folder, 'Excel sync network folder path', current_user.id)
        
        return jsonify({'message': 'Network folder configured successfully'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel-sync/status')
@login_required
def api_excel_sync_status():
    """Get Excel sync status"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    return jsonify({
        'running': excel_sync_running,
        'paused': excel_sync_paused,
        'status': excel_sync_status['status'],
        'last_sync': excel_sync_status['last_sync'],
        'files_synced': excel_sync_status['files_synced'],
        'error_count': len(excel_sync_status['errors']),
        'errors': excel_sync_status['errors'][-10:],  # Last 10 errors
        'network_folder': app.config['EXCEL_NETWORK_FOLDER'],
        'local_folder': app.config['EXCEL_LOCAL_FOLDER']
    })

@app.route('/api/excel-sync/clear-notifications', methods=['POST'])
@login_required
def api_clear_notifications():
    """Clear notification data from Excel files in network drive"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json() or {}
        file_name = data.get('file_name')  # Optional specific file
        
        # Clear notification data
        result = clear_notification_data(file_name)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'files_cleared': result['cleared'],
                'files': result['files']
            })
        else:
            return jsonify({
                'success': False,
                'error': result['message']
            }), 400
    
    except Exception as e:
        excel_log_message(f"âŒ API clear notifications error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel-sync/format-power-automate', methods=['POST'])
@login_required
def api_format_power_automate():
    """Format Excel files in network drive for Power Automate compatibility"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json() or {}
        file_name = data.get('file_name')  # Optional specific file
        create_backup = data.get('backup', False)  # Default to no backup
        
        if not app.config['EXCEL_NETWORK_FOLDER']:
            return jsonify({'error': 'Network folder not configured'}), 400
        
        # Import Power Automate formatter
        try:
            from scripts.excel_power_automate_formatter import excel_power_automate_formatter
        except ImportError:
            return jsonify({'error': 'Power Automate formatter module not available'}), 500
        
        network_path = Path(app.config['EXCEL_NETWORK_FOLDER'])
        
        if file_name:
            # Format specific file
            file_path = network_path / file_name
            if not file_path.exists():
                return jsonify({'error': f'File not found: {file_name}'}), 404
            
            success = excel_power_automate_formatter.validate_and_format_excel_file(file_path, create_backup)
            
            if success:
                excel_log_message(f"âœ… Manually formatted {file_name} for Power Automate")
                return jsonify({
                    'success': True,
                    'message': f'Successfully formatted {file_name} for Power Automate',
                    'formatted_files': [file_name]
                })
            else:
                return jsonify({
                    'success': False,
                    'error': f'Failed to format {file_name}'
                }), 400
        else:
            # Format all files in directory
            result = excel_power_automate_formatter.format_all_files_in_directory(network_path, create_backup)
            
            excel_log_message(f"ðŸ“Š Manual Power Automate formatting: {result['message']}")
            
            return jsonify({
                'success': result['success'],
                'message': result['message'],
                'formatted_count': result.get('formatted_count', 0),
                'total_files': result.get('total_files', 0),
                'formatted_files': result.get('formatted_files', []),
                'errors': result.get('errors', [])
            })
    
    except Exception as e:
        excel_log_message(f"âŒ API Power Automate formatting error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel-sync/populate-vendor-data', methods=['POST'])
@login_required
def api_populate_vendor_data():
    """Populate Excel files in network folder with vendor data from database"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json() or {}
        network_folder = data.get('network_folder', app.config.get('EXCEL_NETWORK_FOLDER'))
        
        if not network_folder:
            return jsonify({'error': 'Network folder not specified'}), 400
        
        # Import vendor population script
        try:
            from scripts.populate_excel_from_vendors import main as populate_vendors
        except ImportError:
            return jsonify({'error': 'Vendor population module not available'}), 500
        
        # Run the population
        excel_log_message(f"ðŸ”„ Populating Excel files with vendor data in: {network_folder}")
        
        try:
            populate_vendors(network_folder_path=network_folder)
            excel_log_message(f"âœ… Successfully populated Excel files with vendor data")
            
            return jsonify({
                'success': True,
                'message': 'Successfully populated Excel files with vendor data',
                'network_folder': network_folder
            })
        except Exception as populate_error:
            excel_log_message(f"âŒ Failed to populate vendor data: {str(populate_error)}")
            return jsonify({
                'success': False,
                'error': f'Failed to populate vendor data: {str(populate_error)}'
            }), 500
    
    except Exception as e:
        excel_log_message(f"âŒ API vendor population error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# System Issues API endpoints
@app.route('/api/system-issues')
@login_required
def api_get_system_issues():
    """Get all active system issues"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        from system_issues import SystemIssueManager
        active_issues = SystemIssueManager.get_active_issues()
        return jsonify({
            'success': True,
            'issues': [issue.to_dict() for issue in active_issues],
            'count': len(active_issues)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/system-issues/<int:issue_id>/resolve', methods=['POST'])
@login_required
def api_resolve_system_issue(issue_id):
    """Mark a system issue as resolved"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json() or {}
        resolution_notes = data.get('resolution_notes', '')
        
        from system_issues import SystemIssueManager
        success = SystemIssueManager.resolve_issue(issue_id, current_user.id, resolution_notes)
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Issue marked as resolved'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Issue not found or already resolved'
            }), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel-sync/list-files')
@login_required
def api_list_notification_files():
    """List all notification Excel files available for clearing"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        if not app.config['EXCEL_NETWORK_FOLDER']:
            return jsonify({'files': [], 'message': 'Network folder not configured'})
        
        network_path = Path(app.config['EXCEL_NETWORK_FOLDER'])
        
        if not network_path.exists():
            return jsonify({'files': [], 'message': 'Network folder not found'})
        
        excel_files = list(network_path.glob("*.xlsx"))
        # Exclude Power Automate backup files and Excel temporary lock files from admin listing
        excel_files = [p for p in excel_files if not p.name.startswith("PA_backup_") 
                      and not p.name.lower().startswith("pa_backup_")
                      and not p.name.startswith("~$")]
        
        file_info = []
        for file_path in excel_files:
            try:
                # Get file info
                stat = file_path.stat()
                
                # Try to count data rows
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                data_rows = max(0, ws.max_row - 1)  # Exclude header
                wb.close()
                
                file_info.append({
                    'name': file_path.name,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'data_rows': data_rows,
                    'display_name': file_path.stem.replace('_', ' ').title()
                })
                
            except Exception as e:
                file_info.append({
                    'name': file_path.name,
                    'size': stat.st_size if 'stat' in locals() else 0,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S') if 'stat' in locals() else 'Unknown',
                    'data_rows': 'Error',
                    'display_name': file_path.stem.replace('_', ' ').title(),
                    'error': str(e)
                })
        
        return jsonify({
            'success': True,
            'files': file_info,
            'count': len(file_info)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================
# NOTIFICATION SCHEDULER ROUTES
# ============================================

@app.route('/admin/notification-scheduler')
@login_required
def admin_notification_scheduler():
    """Admin page for notification scheduler control"""
    if current_user.role != UserRole.ADMIN:
        flash('Access denied', 'error')
        return redirect(url_for('index'))
    
    # Get current status
    status = None
    if notification_scheduler_available and notification_scheduler_service:
        status = notification_scheduler_service.get_status()
    
    return render_template('admin_notification_scheduler.html',
                         available=notification_scheduler_available,
                         status=status)

@app.route('/api/notification-scheduler/start', methods=['POST'])
@login_required
def api_notification_scheduler_start():
    """Start the notification scheduler"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if not notification_scheduler_available or not notification_scheduler_service:
        return jsonify({'error': 'Notification scheduler not available'}), 500
    
    result = notification_scheduler_service.start()
    return jsonify(result)

@app.route('/api/notification-scheduler/stop', methods=['POST'])
@login_required
def api_notification_scheduler_stop():
    """Stop the notification scheduler"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if not notification_scheduler_available or not notification_scheduler_service:
        return jsonify({'error': 'Notification scheduler not available'}), 500
    
    result = notification_scheduler_service.stop()
    return jsonify(result)

@app.route('/api/notification-scheduler/pause', methods=['POST'])
@login_required
def api_notification_scheduler_pause():
    """Pause the notification scheduler"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if not notification_scheduler_available or not notification_scheduler_service:
        return jsonify({'error': 'Notification scheduler not available'}), 500
    
    result = notification_scheduler_service.pause()
    return jsonify(result)

@app.route('/api/notification-scheduler/resume', methods=['POST'])
@login_required
def api_notification_scheduler_resume():
    """Resume the notification scheduler"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if not notification_scheduler_available or not notification_scheduler_service:
        return jsonify({'error': 'Notification scheduler not available'}), 500
    
    result = notification_scheduler_service.resume()
    return jsonify(result)

@app.route('/api/notification-scheduler/force-sync', methods=['POST'])
@login_required
def api_notification_scheduler_force_sync():
    """Force an immediate notification sync"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if not notification_scheduler_available or not notification_scheduler_service:
        return jsonify({'error': 'Notification scheduler not available'}), 500
    
    result = notification_scheduler_service.force_sync()
    return jsonify(result)

@app.route('/api/notification-scheduler/status', methods=['GET'])
@login_required
def api_notification_scheduler_status():
    """Get notification scheduler status"""
    if current_user.role != UserRole.ADMIN:
        return jsonify({'error': 'Access denied'}), 403
    
    if not notification_scheduler_available or not notification_scheduler_service:
        return jsonify({'available': False, 'error': 'Notification scheduler not available'}), 200
    
    status = notification_scheduler_service.get_status()
    return jsonify({'available': True, 'status': status})

if __name__ == '__main__':
    print("\n" + "="*70)
    print("ATTENDO - Starting Application...")
    print("="*70)
    
    # Create tables and initialize demo data
    with app.app_context():
        create_tables()
        print("Database initialized. Default Admin user ensured (Admin / admin123).")
    
    # Start notification scheduler (old system)
    start_notification_scheduler()
    print("Notification scheduler (old) started!")
    
    # Start enhanced notification scheduler service
    if notification_scheduler_available and notification_scheduler_service:
        try:
            notification_scheduler_service.init_app(app)
            result = notification_scheduler_service.start()
            if result.get('success'):
                print("âœ… Enhanced notification scheduler started successfully!")
                print("   - Checks every 10 minutes for scheduled notifications")
                print("   - Auto-removes sent notifications")
                print("   - Syncs to network folder for Power Automate")
            else:
                print(f"âš ï¸ Enhanced notification scheduler failed to start: {result.get('message')}")
        except Exception as e:
            print(f"âš ï¸ Enhanced notification scheduler error: {e}")
    
    # Load Excel sync network folder from system config
    try:
        with app.app_context():
            network_folder_config = get_system_config('excel_network_folder')
            if network_folder_config:
                app.config['EXCEL_NETWORK_FOLDER'] = network_folder_config
                print(f"Excel sync network folder loaded: {network_folder_config}")
                
                # Auto-start Excel sync if configured
                if app.config['EXCEL_SYNC_ENABLED'] and Path(app.config['EXCEL_LOCAL_FOLDER']).exists():
                    start_result = start_excel_sync()
                    print(f"Excel sync service: {start_result}")
    except Exception as e:
        print(f"Excel sync initialization warning: {e}")
    
    print("="*70)
    print("ATTENDO is now running!")
    print("="*70)
    print("\nAccess the application at:")
    print("   Web Interface: http://localhost:5000")
    print("   API Documentation: http://localhost:5000/api/docs")
    print("\nLogin Credentials:")
    print("   Admin:    Admin / admin123")
    print("   (Create Managers and Vendors via Admin > System Settings > Import)")
    print("\nPress CTRL+C to stop the server")
    print("="*70 + "\n")
    
    # Run the Flask app
    app.run(debug=True, host='0.0.0.0', port=5000)
